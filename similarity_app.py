import streamlit as st
import pandas as pd
import numpy as np
import io
import os
import json
import html
import logging
import subprocess
import threading
import shutil
import hashlib
import difflib
import re
import sys
import traceback
import zipfile
import xml.etree.ElementTree as ET
from collections import Counter

from openpyxl.utils import get_column_letter
import requests
import streamlit.components.v1 as components
from model_configs import get_default_provider_model, get_model_parameters, get_provider_model_options
from auth import ERROR_LOG_PATH, check_authentication, show_user_info, get_current_user_key


logger = logging.getLogger(__name__)


if not check_authentication():
    st.stop()
show_user_info()

DEFAULT_SYSTEM_PROMPT = (
    "You are a similarity scoring engine for answer matching. Compare semantic meaning "
    "and final answer equivalence, not formatting. Ignore casing, punctuation, filler "
    "phrases such as 'based on the provided context', and minor wording changes. Use the "
    "question to judge whether both answers address the same thing. If both answers "
    "honestly indicate the context does not provide the requested information, score them "
    "high. If one answer abstains but the other gives concrete content, score them lower. "
    "If one answer is more cautious about missing details but the shared facts still align, "
    "treat that as a moderate-to-high match rather than a contradiction. "
    "Return only one numeric score from 0 to 100."
)
DEFAULT_USER_PROMPT_TEMPLATE = (
    "Question:\n{question}\n\n"
    "Compare these two answers to the same question and score their semantic similarity from 0 to 100.\n"
    "Give a high score when they mean the same thing even if the wording is different.\n"
    "If both answers correctly say the context does not provide the information, give a high score.\n"
    "If one answer abstains while the other gives a concrete answer, lower the score.\n"
    "If one answer says some details are not stated but both answers still share the same available facts, score it as a moderate-to-high match instead of a low mismatch.\n"
    "Return only the numeric score.\n"
    "Answer 1:\n{answer1}\n\n"
    "Answer 2:\n{answer2}"
)

SIMILARITY_CONTEXT_PHRASES = (
    r"based on the provided context",
    r"from the context",
    r"from context",
    r"context",
    r"see context",
    r"as per context",
    r"per context",
    r"per the context",
    r"per the provided context",
    r"provided context",
    r"according to",
)

SIMILARITY_ABSTENTION_PATTERNS = (
    r"\bnot (?:explicitly )?(?:stated|described|provided|mentioned|available|included)\b",
    r"\bnot (?:fully )?(?:explained|detailed|defined|clarified)\b",
    r"\bnot explicitly (?:detailed|explained|defined|clarified)\b",
    r"\bnot described in detail\b",
    r"\bdoes not (?:explicitly )?(?:state|describe|provide|mention|specify|cover|include)\b",
    r"\bdoes not (?:fully )?(?:explain|detail|clarify|define)\b",
    r"\bnot enough information\b",
    r"\binsufficient information\b",
    r"\bno (?:specific )?(?:information|details?) (?:are|is) (?:provided|available)\b",
    r"\bthe provided context does not\b",
    r"\bthe context does not\b",
    r"\bthe provided context only (?:indicates|states|mentions|shows)\b",
    r"\bthe context only (?:indicates|states|mentions|shows)\b",
    r"\bnot available in the provided context\b",
)

SIMILARITY_NEGATION_TOKENS = {
    "no",
    "not",
    "never",
    "none",
    "cannot",
    "without",
    "false",
    "incorrect",
    "neither",
    "nor",
}

SIMILARITY_STRONG_NEGATION_TOKENS = {
    "no",
    "not",
    "never",
    "none",
    "cannot",
    "false",
    "incorrect",
    "neither",
    "nor",
}

SIMILARITY_TRUE_TOKENS = {
    "yes",
    "true",
    "approved",
    "allow",
    "allowed",
    "enable",
    "enabled",
    "success",
    "successful",
    "present",
}

SIMILARITY_FALSE_TOKENS = {
    "no",
    "false",
    "denied",
    "disallowed",
    "disable",
    "disabled",
    "fail",
    "failed",
    "absent",
}

SIMILARITY_STOPWORDS = {
    "a",
    "an",
    "and",
    "are",
    "as",
    "at",
    "be",
    "been",
    "by",
    "for",
    "from",
    "has",
    "have",
    "in",
    "is",
    "it",
    "its",
    "of",
    "on",
    "or",
    "that",
    "the",
    "their",
    "there",
    "this",
    "to",
    "was",
    "were",
    "will",
    "with",
}

SIMILARITY_MONTH_TOKENS = {
    "jan",
    "january",
    "feb",
    "february",
    "mar",
    "march",
    "apr",
    "april",
    "may",
    "jun",
    "june",
    "jul",
    "july",
    "aug",
    "august",
    "sep",
    "sept",
    "september",
    "oct",
    "october",
    "nov",
    "november",
    "dec",
    "december",
}

SIMILARITY_UNIT_SPECS = {
    "mg": ("mass", 0.001),
    "milligram": ("mass", 0.001),
    "milligrams": ("mass", 0.001),
    "g": ("mass", 1.0),
    "gram": ("mass", 1.0),
    "grams": ("mass", 1.0),
    "kg": ("mass", 1000.0),
    "kgs": ("mass", 1000.0),
    "kilogram": ("mass", 1000.0),
    "kilograms": ("mass", 1000.0),
    "lb": ("mass", 453.59237),
    "lbs": ("mass", 453.59237),
    "pound": ("mass", 453.59237),
    "pounds": ("mass", 453.59237),
    "oz": ("mass", 28.349523125),
    "ounce": ("mass", 28.349523125),
    "ounces": ("mass", 28.349523125),
    "mm": ("length", 0.001),
    "millimeter": ("length", 0.001),
    "millimeters": ("length", 0.001),
    "cm": ("length", 0.01),
    "centimeter": ("length", 0.01),
    "centimeters": ("length", 0.01),
    "m": ("length", 1.0),
    "meter": ("length", 1.0),
    "meters": ("length", 1.0),
    "km": ("length", 1000.0),
    "kilometer": ("length", 1000.0),
    "kilometers": ("length", 1000.0),
    "in": ("length", 0.0254),
    "inch": ("length", 0.0254),
    "inches": ("length", 0.0254),
    "ft": ("length", 0.3048),
    "foot": ("length", 0.3048),
    "feet": ("length", 0.3048),
    "yd": ("length", 0.9144),
    "yard": ("length", 0.9144),
    "yards": ("length", 0.9144),
    "mi": ("length", 1609.344),
    "mile": ("length", 1609.344),
    "miles": ("length", 1609.344),
    "ml": ("volume", 0.001),
    "milliliter": ("volume", 0.001),
    "milliliters": ("volume", 0.001),
    "millilitre": ("volume", 0.001),
    "millilitres": ("volume", 0.001),
    "l": ("volume", 1.0),
    "liter": ("volume", 1.0),
    "liters": ("volume", 1.0),
    "litre": ("volume", 1.0),
    "litres": ("volume", 1.0),
    "gal": ("volume", 3.785411784),
    "gallon": ("volume", 3.785411784),
    "gallons": ("volume", 3.785411784),
    "ms": ("duration", 0.001),
    "millisecond": ("duration", 0.001),
    "milliseconds": ("duration", 0.001),
    "sec": ("duration", 1.0),
    "secs": ("duration", 1.0),
    "second": ("duration", 1.0),
    "seconds": ("duration", 1.0),
    "min": ("duration", 60.0),
    "mins": ("duration", 60.0),
    "minute": ("duration", 60.0),
    "minutes": ("duration", 60.0),
    "hr": ("duration", 3600.0),
    "hrs": ("duration", 3600.0),
    "hour": ("duration", 3600.0),
    "hours": ("duration", 3600.0),
    "kb": ("storage", 1024.0),
    "mb": ("storage", 1024.0 * 1024.0),
    "gb": ("storage", 1024.0 * 1024.0 * 1024.0),
    "tb": ("storage", 1024.0 * 1024.0 * 1024.0 * 1024.0),
    "byte": ("storage", 1.0),
    "bytes": ("storage", 1.0),
}

SIMILARITY_UNIT_PATTERN = re.compile(
    r"(?<!\w)(?P<value>\d+(?:\.\d+)?)\s*(?P<unit>"
    + "|".join(sorted((re.escape(unit) for unit in SIMILARITY_UNIT_SPECS.keys()), key=len, reverse=True))
    + r")\b",
    re.IGNORECASE,
)


def _normalized_prompt_value(prompt_value, fallback):
    """Return fallback when prompt is missing/blank, otherwise keep user-entered content."""
    if isinstance(prompt_value, str) and prompt_value.strip():
        return prompt_value
    return fallback


def _normalize_similarity_text(text, relax_numbers=False):
    text = str(text or "").lower().strip()
    if not text:
        return ""
    text = re.sub(r"\[.*?\]", " ", text)
    for phrase in SIMILARITY_CONTEXT_PHRASES:
        text = re.sub(rf"\b{phrase}\b", " ", text)
    text = re.sub(r"[^\w\s]", " ", text)
    if relax_numbers:
        text = re.sub(r"\d+(?:\.\d+)?", " <num> ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def _strip_similarity_markup(text):
    cleaned = str(text or "")
    if not cleaned:
        return ""
    cleaned = re.sub(r"\[(?:\d+(?:\s*,\s*\d+)*)\]", " ", cleaned)
    cleaned = re.sub(r"\[(?:\d+)\]\[(?:\d+)\]", " ", cleaned)
    cleaned = re.sub(r"(?m)^\s{0,3}#{1,6}\s*", "", cleaned)
    cleaned = cleaned.replace("**", " ").replace("__", " ").replace("`", " ")
    cleaned = re.sub(r"(?m)^\s*[-*+]\s+", "", cleaned)
    cleaned = re.sub(r"(?m)^\s*\d+\.\s+", "", cleaned)
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    return cleaned


def _prepare_answer_for_similarity(text):
    return _strip_similarity_markup(text)


def _prepare_question_for_similarity(text):
    return _strip_similarity_markup(text)


def _has_abstention_signal(text):
    source = _prepare_answer_for_similarity(text).lower()
    if not source:
        return False
    return any(re.search(pattern, source, flags=re.IGNORECASE) for pattern in SIMILARITY_ABSTENTION_PATTERNS)


def _char_ngrams(text, n=3):
    if len(text) < n:
        return Counter([text]) if text else Counter()
    return Counter(text[i : i + n] for i in range(len(text) - n + 1))


def _counter_cosine_similarity(c1, c2):
    if not c1 or not c2:
        return 0.0
    keys = set(c1.keys()) | set(c2.keys())
    dot = sum(c1.get(k, 0) * c2.get(k, 0) for k in keys)
    n1 = sum(v * v for v in c1.values()) ** 0.5
    n2 = sum(v * v for v in c2.values()) ** 0.5
    if n1 == 0 or n2 == 0:
        return 0.0
    return dot / (n1 * n2)


def _lexical_similarity_percent_core(text1, text2, relax_numbers=False):
    """Hybrid lexical similarity fallback (0-100) with better quality than raw ratio."""
    a = _normalize_similarity_text(text1, relax_numbers=relax_numbers)
    b = _normalize_similarity_text(text2, relax_numbers=relax_numbers)
    if not a or not b:
        return 0.0

    seq_ratio = difflib.SequenceMatcher(None, a, b).ratio()
    a_tokens = a.split()
    b_tokens = b.split()
    a_set = set(a_tokens)
    b_set = set(b_tokens)
    union = a_set | b_set
    jaccard = (len(a_set & b_set) / len(union)) if union else 0.0

    # Token-order-insensitive overlap.
    sorted_ratio = difflib.SequenceMatcher(None, " ".join(sorted(a_tokens)), " ".join(sorted(b_tokens))).ratio()

    # Character n-gram cosine captures near matches/typos.
    char_cos = _counter_cosine_similarity(_char_ngrams(a, 3), _char_ngrams(b, 3))

    # Numeric consistency is important for answer matching.
    nums_a = set(re.findall(r"\d+(?:\.\d+)?", a))
    nums_b = set(re.findall(r"\d+(?:\.\d+)?", b))
    if nums_a or nums_b:
        num_union = nums_a | nums_b
        num_score = (len(nums_a & nums_b) / len(num_union)) if num_union else 0.0
    else:
        num_score = 1.0

    score = (
        0.40 * seq_ratio
        + 0.20 * jaccard
        + 0.20 * sorted_ratio
        + 0.15 * char_cos
        + 0.05 * num_score
    )
    return round(max(0.0, min(1.0, score)) * 100.0, 2)


def _lexical_similarity_percent(text1, text2):
    strict_score = _lexical_similarity_percent_core(text1, text2, relax_numbers=False)
    relaxed_score = _lexical_similarity_percent_core(text1, text2, relax_numbers=True)
    if relaxed_score <= strict_score:
        return strict_score

    strict_a = _normalize_similarity_text(text1)
    strict_b = _normalize_similarity_text(text2)
    nums_a = set(re.findall(r"\d+(?:\.\d+)?", strict_a))
    nums_b = set(re.findall(r"\d+(?:\.\d+)?", strict_b))

    if not nums_a and not nums_b:
        return strict_score

    boost_cap = 10.0 if nums_a and nums_b and not (nums_a & nums_b) else 14.0
    return round(min(relaxed_score, strict_score + boost_cap), 2)


def _has_negation_mismatch(text1, text2):
    if _has_abstention_signal(text1) or _has_abstention_signal(text2):
        return False
    a_tokens = set(_normalize_similarity_text(text1).split())
    b_tokens = set(_normalize_similarity_text(text2).split())
    a_has_negation = bool(a_tokens & SIMILARITY_STRONG_NEGATION_TOKENS)
    b_has_negation = bool(b_tokens & SIMILARITY_STRONG_NEGATION_TOKENS)
    return a_has_negation != b_has_negation


def _extract_numeric_tokens(text):
    return set(re.findall(r"\d+(?:\.\d+)?", _normalize_similarity_text(text)))


def _extract_percent_tokens(text):
    source = str(text or "").lower()
    return set(re.findall(r"(\d+(?:\.\d+)?)\s*(?:%|percent)", source))


def _extract_date_markers(text):
    source = str(text or "").lower()
    markers = set(re.findall(r"\b\d{1,2}[/-]\d{1,2}(?:[/-]\d{2,4})?\b", source))
    markers.update(re.findall(r"\b\d{4}[/-]\d{1,2}[/-]\d{1,2}\b", source))
    month_expr = r"(?:jan(?:uary)?|feb(?:ruary)?|mar(?:ch)?|apr(?:il)?|may|jun(?:e)?|jul(?:y)?|aug(?:ust)?|sep(?:t(?:ember)?)?|oct(?:ober)?|nov(?:ember)?|dec(?:ember)?)"
    markers.update(re.findall(rf"\b{month_expr}\s+\d{{1,2}}(?:,\s*\d{{2,4}})?\b", source))
    markers.update(re.findall(rf"\b\d{{1,2}}\s+{month_expr}(?:\s+\d{{2,4}})?\b", source))
    return markers


def _extract_boolean_label(text):
    significant = _significant_tokens(text)
    if len(significant) > 12:
        return None
    tokens = set(_normalize_similarity_text(text).split())
    has_true = bool(tokens & SIMILARITY_TRUE_TOKENS)
    has_false = bool(tokens & SIMILARITY_FALSE_TOKENS)
    if has_true and not has_false:
        return "true"
    if has_false and not has_true:
        return "false"
    return None


def _extract_quantity_units(text):
    quantities = []
    source = str(text or "").lower()
    for match in SIMILARITY_UNIT_PATTERN.finditer(source):
        raw_value = match.group("value")
        unit = match.group("unit").lower()
        spec = SIMILARITY_UNIT_SPECS.get(unit)
        if not spec:
            continue
        category, factor = spec
        try:
            numeric_value = float(raw_value)
        except Exception:
            continue
        quantities.append(
            {
                "raw_value": raw_value,
                "category": category,
                "unit": unit,
                "normalized": numeric_value * factor,
            }
        )
    return quantities


def _quantities_close(value_a, value_b):
    scale = max(1.0, abs(value_a), abs(value_b))
    return abs(value_a - value_b) <= (scale * 0.02)


def _analyze_quantity_unit_relationships(text1, text2):
    quantities_a = _extract_quantity_units(text1)
    quantities_b = _extract_quantity_units(text2)
    categories_a = {item["category"] for item in quantities_a}
    categories_b = {item["category"] for item in quantities_b}
    shared_categories = categories_a & categories_b
    matched_numeric_tokens_a = set()
    matched_numeric_tokens_b = set()
    unit_mismatch = False

    for category in shared_categories:
        cat_items_a = [item for item in quantities_a if item["category"] == category]
        cat_items_b = [item for item in quantities_b if item["category"] == category]
        matched_in_category = False
        for item_a in cat_items_a:
            for item_b in cat_items_b:
                if _quantities_close(item_a["normalized"], item_b["normalized"]):
                    matched_in_category = True
                    matched_numeric_tokens_a.add(item_a["raw_value"])
                    matched_numeric_tokens_b.add(item_b["raw_value"])
        if not matched_in_category:
            unit_mismatch = True

    return {
        "matched_numeric_tokens_a": matched_numeric_tokens_a,
        "matched_numeric_tokens_b": matched_numeric_tokens_b,
        "unit_mismatch": unit_mismatch,
    }


def _significant_tokens(text):
    tokens = []
    for token in _normalize_similarity_text(text, relax_numbers=True).split():
        if token == "<num>":
            tokens.append(token)
            continue
        if token in SIMILARITY_STOPWORDS:
            continue
        if len(token) <= 2:
            continue
        tokens.append(token)
    return tokens


def _significant_overlap_ratio(text1, text2):
    tokens_a = set(_significant_tokens(text1))
    tokens_b = set(_significant_tokens(text2))
    if not tokens_a or not tokens_b:
        return 0.0
    return len(tokens_a & tokens_b) / max(1, min(len(tokens_a), len(tokens_b)))


def _is_short_subset_match(text1, text2):
    tokens_a = set(_significant_tokens(text1))
    tokens_b = set(_significant_tokens(text2))
    if not tokens_a or not tokens_b:
        return False
    shorter, longer = (tokens_a, tokens_b) if len(tokens_a) <= len(tokens_b) else (tokens_b, tokens_a)
    if len(shorter) > 3:
        return False
    return shorter.issubset(longer)


def _detect_similarity_conflicts(text1, text2):
    numeric_a = _extract_numeric_tokens(text1)
    numeric_b = _extract_numeric_tokens(text2)
    percent_a = _extract_percent_tokens(text1)
    percent_b = _extract_percent_tokens(text2)
    date_a = _extract_date_markers(text1)
    date_b = _extract_date_markers(text2)
    bool_a = _extract_boolean_label(text1)
    bool_b = _extract_boolean_label(text2)
    abstain_a = _has_abstention_signal(text1)
    abstain_b = _has_abstention_signal(text2)
    quantity_relationships = _analyze_quantity_unit_relationships(text1, text2)
    residual_numeric_a = numeric_a - quantity_relationships["matched_numeric_tokens_a"]
    residual_numeric_b = numeric_b - quantity_relationships["matched_numeric_tokens_b"]

    return {
        "negation_mismatch": _has_negation_mismatch(text1, text2),
        "boolean_mismatch": bool(bool_a and bool_b and bool_a != bool_b),
        "numeric_mismatch": bool(residual_numeric_a and residual_numeric_b and not (residual_numeric_a & residual_numeric_b)),
        "percent_mismatch": bool(percent_a and percent_b and not (percent_a & percent_b)),
        "date_mismatch": bool(date_a and date_b and not (date_a & date_b)),
        "unit_mismatch": quantity_relationships["unit_mismatch"],
        "abstention_mismatch": bool(abstain_a != abstain_b),
        "both_abstain": bool(abstain_a and abstain_b),
    }


def _calibrate_similarity_score(score, text1, text2):
    if score is None:
        return None
    try:
        score_val = float(score)
    except Exception:
        return None

    score_val = max(0.0, min(100.0, score_val))
    strict_a = _normalize_similarity_text(text1)
    strict_b = _normalize_similarity_text(text2)
    if not strict_a or not strict_b:
        return round(score_val, 2)
    if strict_a == strict_b:
        return 100.0

    relaxed_a = _normalize_similarity_text(text1, relax_numbers=True)
    relaxed_b = _normalize_similarity_text(text2, relax_numbers=True)
    strict_score = _lexical_similarity_percent_core(text1, text2, relax_numbers=False)
    anchor_score = _lexical_similarity_percent(text1, text2)
    nums_a = set(re.findall(r"\d+(?:\.\d+)?", strict_a))
    nums_b = set(re.findall(r"\d+(?:\.\d+)?", strict_b))
    has_disjoint_numbers = bool(nums_a and nums_b and not (nums_a & nums_b))
    overlap_ratio = _significant_overlap_ratio(text1, text2)
    short_subset_match = _is_short_subset_match(text1, text2)
    conflicts = _detect_similarity_conflicts(text1, text2)
    negation_mismatch = conflicts["negation_mismatch"]
    hard_value_conflict = any(
        conflicts[key]
        for key in ("boolean_mismatch", "percent_mismatch", "date_mismatch", "unit_mismatch", "numeric_mismatch")
    )
    shared_fact_abstention = (
        conflicts["abstention_mismatch"]
        and not hard_value_conflict
        and overlap_ratio >= 0.5
    )
    adjusted = score_val

    if not negation_mismatch and relaxed_a == relaxed_b and strict_a != strict_b:
        adjusted = max(adjusted, 78.0 if has_disjoint_numbers else 86.0)
    elif not negation_mismatch:
        if anchor_score >= 90.0 and adjusted < 78.0:
            adjusted = max(adjusted, round((adjusted * 0.60) + (anchor_score * 0.40), 2))
        elif anchor_score >= 82.0 and adjusted < 70.0:
            adjusted = max(adjusted, round((adjusted * 0.70) + (anchor_score * 0.30), 2))
        elif anchor_score >= 74.0 and (anchor_score - adjusted) >= 18.0:
            adjusted = max(adjusted, round((adjusted * 0.80) + (anchor_score * 0.20), 2))

    # Hard contradiction caps prevent wording overlap from producing false 90+ matches.
    if conflicts["boolean_mismatch"]:
        adjusted = min(adjusted, 52.0)
    if conflicts["percent_mismatch"]:
        adjusted = min(adjusted, 58.0)
    if conflicts["date_mismatch"]:
        adjusted = min(adjusted, 60.0)
    if conflicts["unit_mismatch"]:
        adjusted = min(adjusted, 60.0 if overlap_ratio >= 0.6 else 68.0)
    if conflicts["numeric_mismatch"]:
        adjusted = min(adjusted, 68.0 if overlap_ratio >= 0.6 else 75.0)
    if conflicts["abstention_mismatch"]:
        if shared_fact_abstention and (anchor_score >= 32.0 or score_val >= 72.0):
            adjusted = min(adjusted, 80.0 if overlap_ratio >= 0.65 else 76.0)
            if score_val >= 72.0:
                adjusted = max(adjusted, 72.0)
        elif overlap_ratio >= 0.6 or anchor_score >= 48.0:
            adjusted = min(adjusted, 74.0)
        elif overlap_ratio >= 0.45 or anchor_score >= 34.0:
            adjusted = min(adjusted, 68.0)
        else:
            adjusted = min(adjusted, 62.0)
    if negation_mismatch:
        adjusted = min(adjusted, strict_score + 6.0, 62.0)
    if conflicts["both_abstain"] and not has_disjoint_numbers and not hard_value_conflict:
        if score_val >= 90.0 and overlap_ratio >= 0.4:
            adjusted = max(adjusted, 94.0 if anchor_score >= 45.0 else 92.0)
        else:
            adjusted = max(adjusted, 88.0 if anchor_score >= 50.0 else 82.0)
    if not hard_value_conflict and not negation_mismatch:
        if (
            shared_fact_abstention
            and adjusted < 68.0
            and overlap_ratio >= 0.55
            and (anchor_score >= 25.0 or score_val >= 60.0)
        ):
            adjusted = max(adjusted, 72.0 if overlap_ratio >= 0.68 and anchor_score >= 32.0 else 68.0)
        elif (
            not conflicts["abstention_mismatch"]
            and not conflicts["both_abstain"]
            and adjusted < 66.0
            and overlap_ratio >= 0.72
            and anchor_score >= 30.0
        ):
            adjusted = max(adjusted, 68.0)

    # Partial-answer risk: one answer is much longer, with limited shared content.
    sig_a = _significant_tokens(text1)
    sig_b = _significant_tokens(text2)
    length_ratio = (max(len(sig_a), len(sig_b)) / max(1, min(len(sig_a), len(sig_b)))) if sig_a and sig_b else 1.0
    if (
        adjusted > 84.0
        and not short_subset_match
        and overlap_ratio < 0.5
        and length_ratio >= 1.8
        and anchor_score < 88.0
    ):
        adjusted = min(adjusted, 84.0)

    # Balanced high-score gate: allow 90+ only for strong semantic agreement with no conflicts.
    has_conflict = any(conflicts.values())
    semantic_high_equivalence = (
        not has_conflict
        and length_ratio <= 3.6
        and (
            (score_val >= 96.0 and overlap_ratio >= 0.68)
            or (score_val >= 94.0 and overlap_ratio >= 0.75)
            or (score_val >= 92.0 and overlap_ratio >= 0.82)
            or (short_subset_match and score_val >= 90.0 and overlap_ratio >= 0.58)
        )
    )
    if semantic_high_equivalence:
        adjusted = max(adjusted, min(score_val, 96.0 if overlap_ratio >= 0.85 else 94.0))

    both_abstain_high_equivalence = (
        conflicts["both_abstain"]
        and not has_disjoint_numbers
        and not hard_value_conflict
        and score_val >= 90.0
        and overlap_ratio >= 0.4
        and length_ratio <= 3.2
    )
    allow_ninety_plus = (
        (
            not has_conflict
            and (
                strict_a == strict_b
                or (relaxed_a == relaxed_b and not has_disjoint_numbers)
                or semantic_high_equivalence
                or (short_subset_match and score_val >= 88.0 and overlap_ratio >= 0.5)
                or (score_val >= 92.0 and anchor_score >= 84.0 and overlap_ratio >= 0.5)
                or (score_val >= 95.0 and anchor_score >= 80.0 and overlap_ratio >= 0.4)
            )
        )
        or both_abstain_high_equivalence
    )
    if adjusted > 90.0 and not allow_ninety_plus:
        adjusted = min(adjusted, 89.0)

    return round(max(0.0, min(100.0, adjusted)), 2)


def _calibrate_similarity_series(scores, left_texts, right_texts):
    calibrated = []
    for score, lhs, rhs in zip(scores, left_texts, right_texts):
        calibrated.append(_calibrate_similarity_score(score, lhs, rhs))
    return calibrated


def _build_similarity_system_prompt(system_prompt):
    base_prompt = _normalized_prompt_value(system_prompt, DEFAULT_SYSTEM_PROMPT).strip()
    guidance = (
        "Focus on semantic meaning and final answer equivalence. Use the question to judge "
        "whether both answers respond to the same requirement. Ignore punctuation, casing, "
        "boilerplate/context phrases, citation markers, and minor rewording. If both answers "
        "correctly indicate the source does not provide the requested information, treat them "
        "as a strong match. If one answer is more cautious about missing detail but the shared "
        "facts still align, keep the score in a moderate-to-high range rather than treating it "
        "like a contradiction. Return only a single numeric score from 0 to 100 with no explanation."
    )
    if guidance.lower() in base_prompt.lower():
        return base_prompt
    return f"{base_prompt}\n\n{guidance}".strip()


def _build_similarity_user_template(user_template):
    prompt_template = _normalized_prompt_value(user_template, DEFAULT_USER_PROMPT_TEMPLATE)
    if "{answer1}" not in prompt_template or "{answer2}" not in prompt_template:
        prompt_template = DEFAULT_USER_PROMPT_TEMPLATE
    if "{question}" not in prompt_template:
        prompt_template = f"Question:\n{{question}}\n\n{prompt_template.lstrip()}"
    output_rule = (
        "\n\nScoring rules: compare semantic meaning, use the question as context, ignore filler "
        "phrases/citation markers/formatting, treat matching abstentions as a strong match, keep "
        "partial abstention with aligned facts in a moderate-to-high band, and "
        "output only one number from 0 to 100."
    )
    if "output only one number" in prompt_template.lower() or "return only the numeric score" in prompt_template.lower():
        return prompt_template
    return f"{prompt_template.rstrip()}{output_rule}"


def _get_available_local_models():
    """Return locally installed sentence-transformer model options only."""
    local_cache_root = os.path.join(
        os.path.expanduser("~"),
        ".cache",
        "torch",
        "sentence_transformers",
    )
    cached_mpnet = os.path.join(local_cache_root, "sentence-transformers_all-mpnet-base-v2")
    cached_minilm = os.path.join(local_cache_root, "sentence-transformers_all-MiniLM-L6-v2")
    options = {}
    if os.path.isdir(cached_mpnet):
        options["all-mpnet-base-v2 (Accurate)"] = cached_mpnet
    if os.path.isdir(cached_minilm):
        options["all-MiniLM-L6-v2 (Fast)"] = cached_minilm
    return options


def _has_sentence_transformers() -> bool:
    try:
        import importlib.util
        return importlib.util.find_spec("sentence_transformers") is not None
    except Exception:
        return False


SPREADSHEET_EXTENSIONS = {".xlsx", ".xls", ".csv"}
TEXT_FILE_EXTENSIONS = {".txt", ".md", ".log"}
JSON_FILE_EXTENSIONS = {".json"}
PDF_FILE_EXTENSIONS = {".pdf"}
WORD_FILE_EXTENSIONS = {".docx"}
LEGACY_WORD_FILE_EXTENSIONS = {".doc"}
COMPARE_ANY_TWO_EXTENSIONS = (
    SPREADSHEET_EXTENSIONS
    | TEXT_FILE_EXTENSIONS
    | JSON_FILE_EXTENSIONS
    | PDF_FILE_EXTENSIONS
    | WORD_FILE_EXTENSIONS
)


def _uploaded_file_extension(uploaded_file):
    if uploaded_file is None:
        return ""
    name = getattr(uploaded_file, "name", "") or ""
    return os.path.splitext(name.lower())[1]


def _is_spreadsheet_upload(uploaded_file):
    return _uploaded_file_extension(uploaded_file) in SPREADSHEET_EXTENSIONS


def _read_csv_with_encodings(uploaded_file):
    last_exc = None
    for enc in ("utf-8", "utf-8-sig", "cp1252", "latin1"):
        try:
            uploaded_file.seek(0)
            return pd.read_csv(uploaded_file, encoding=enc, sep=None, engine="python")
        except Exception as exc:
            last_exc = exc
    if last_exc is None:
        raise RuntimeError("Failed to read CSV: unknown error")
    raise last_exc


def _read_excel_upload(uploaded_file, sheet_name=None):
    uploaded_file.seek(0)
    xl = pd.read_excel(uploaded_file, sheet_name=sheet_name)
    if isinstance(xl, dict):
        if not xl:
            raise ValueError("Excel file does not contain any readable sheets.")
        if sheet_name is None:
            return next(iter(xl.values()))
        if sheet_name in xl:
            return xl[sheet_name]
        str_key = str(sheet_name)
        if str_key in xl:
            return xl[str_key]
        return next(iter(xl.values()))
    return xl


def _decode_uploaded_text(uploaded_file):
    last_exc = None
    for enc in ("utf-8", "utf-8-sig", "cp1252", "latin1"):
        try:
            uploaded_file.seek(0)
            raw = uploaded_file.read()
            if isinstance(raw, str):
                return raw
            return raw.decode(enc)
        except Exception as exc:
            last_exc = exc
    if last_exc is None:
        raise RuntimeError("File content could not be decoded as text.")
    raise last_exc


def _build_segment_dataframe(segments, label_prefix):
    clean_segments = [re.sub(r"\s+", " ", str(segment)).strip() for segment in segments if str(segment).strip()]
    if not clean_segments:
        raise ValueError("The uploaded file did not contain readable text to compare.")
    return pd.DataFrame(
        {
            "Section": [f"{label_prefix} {idx + 1}" for idx in range(len(clean_segments))],
            "Content": clean_segments,
        }
    )


def _split_text_segments(text):
    normalized = str(text or "").replace("\r\n", "\n").replace("\r", "\n").strip()
    if not normalized:
        return [], "Section"

    paragraphs = [
        re.sub(r"\s+", " ", chunk).strip()
        for chunk in re.split(r"\n\s*\n+", normalized)
        if chunk.strip()
    ]
    if len(paragraphs) >= 2:
        return paragraphs, "Paragraph"

    lines = [line.strip() for line in normalized.split("\n") if line.strip()]
    if lines:
        return lines, "Line"

    return [normalized], "Section"


def _stringify_nested_value(value):
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return value


def _normalize_structured_dataframe(df, index_label="Row"):
    normalized_df = df.copy()
    normalized_df.columns = [
        str(col).strip() if str(col).strip() else f"Column {idx + 1}"
        for idx, col in enumerate(normalized_df.columns)
    ]
    for col in normalized_df.columns:
        normalized_df[col] = normalized_df[col].apply(_stringify_nested_value)

    if normalized_df.shape[1] == 0:
        raise ValueError("The uploaded file did not contain any comparable data.")

    if normalized_df.shape[1] == 1:
        only_col = normalized_df.columns[0]
        normalized_df.insert(0, index_label, [f"{index_label} {idx + 1}" for idx in range(len(normalized_df))])
        if only_col == index_label:
            normalized_df = normalized_df.rename(columns={only_col: "Value"})

    return normalized_df


def _flatten_json_pairs(value, prefix="root"):
    if isinstance(value, dict):
        rows = []
        if not value:
            return [(prefix, "")]
        for key, item in value.items():
            next_prefix = f"{prefix}.{key}" if prefix else str(key)
            rows.extend(_flatten_json_pairs(item, next_prefix))
        return rows

    if isinstance(value, list):
        rows = []
        if not value:
            return [(prefix, "[]")]
        for idx, item in enumerate(value):
            next_prefix = f"{prefix}[{idx}]"
            rows.extend(_flatten_json_pairs(item, next_prefix))
        return rows

    return [(prefix, _stringify_nested_value(value))]


def _find_json_record_list(value):
    if isinstance(value, list) and value and all(isinstance(item, dict) for item in value):
        return value
    if isinstance(value, dict):
        for item in value.values():
            found = _find_json_record_list(item)
            if found:
                return found
    if isinstance(value, list):
        for item in value:
            found = _find_json_record_list(item)
            if found:
                return found
    return None


def _build_json_upload_dataframe(uploaded_file):
    uploaded_file.seek(0)
    payload = json.loads(_decode_uploaded_text(uploaded_file))

    record_list = _find_json_record_list(payload)
    if record_list:
        return _normalize_structured_dataframe(pd.DataFrame(record_list))

    if isinstance(payload, dict):
        rows = _flatten_json_pairs(payload)
        return pd.DataFrame(rows, columns=["Field", "Value"])

    if isinstance(payload, list):
        if all(not isinstance(item, (dict, list)) for item in payload):
            return pd.DataFrame(
                {
                    "Item": [f"Item {idx + 1}" for idx in range(len(payload))],
                    "Value": [_stringify_nested_value(item) for item in payload],
                }
            )
        rows = _flatten_json_pairs(payload)
        return pd.DataFrame(rows, columns=["Field", "Value"])

    return pd.DataFrame({"Item": ["Item 1"], "Value": [_stringify_nested_value(payload)]})


def _build_text_upload_dataframe(uploaded_file):
    text = _decode_uploaded_text(uploaded_file)
    segments, label_prefix = _split_text_segments(text)
    return _build_segment_dataframe(segments, label_prefix)


def _build_docx_upload_dataframe(uploaded_file):
    uploaded_file.seek(0)
    try:
        with zipfile.ZipFile(uploaded_file) as archive:
            xml_payload = archive.read("word/document.xml")
    except KeyError as exc:
        raise ValueError("DOCX file does not contain a readable document body.") from exc
    except zipfile.BadZipFile as exc:
        raise ValueError("The uploaded DOCX file is invalid or corrupted.") from exc

    root = ET.fromstring(xml_payload)
    namespace = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    paragraphs = []
    for paragraph in root.findall(".//w:body/w:p", namespace):
        text_parts = [node.text for node in paragraph.findall(".//w:t", namespace) if node.text]
        paragraph_text = re.sub(r"\s+", " ", "".join(text_parts)).strip()
        if paragraph_text:
            paragraphs.append(paragraph_text)

    return _build_segment_dataframe(paragraphs, "Paragraph")


def _extract_pdf_text_pages(file_bytes):
    reader_errors = []

    try:
        from pypdf import PdfReader  # type: ignore

        reader = PdfReader(io.BytesIO(file_bytes))
        return [(page.extract_text() or "").strip() for page in reader.pages]
    except Exception as exc:
        reader_errors.append(exc)

    try:
        from PyPDF2 import PdfReader  # type: ignore

        reader = PdfReader(io.BytesIO(file_bytes))
        return [(page.extract_text() or "").strip() for page in reader.pages]
    except Exception as exc:
        reader_errors.append(exc)

    raise RuntimeError(
        "PDF text extraction requires 'pypdf' or 'PyPDF2' in the app environment."
    ) from (reader_errors[-1] if reader_errors else None)


def _build_pdf_upload_dataframe(uploaded_file):
    uploaded_file.seek(0)
    file_bytes = uploaded_file.read()
    pages = _extract_pdf_text_pages(file_bytes)
    clean_pages = [page for page in pages if page and page.strip()]
    if not clean_pages:
        raise ValueError(
            "The PDF appears to be scanned or image-only, so no selectable text could be extracted. "
            "Please run OCR on the PDF or convert it to DOCX/TXT and retry."
        )
    return _build_segment_dataframe(clean_pages, "Page")


def _default_compare_columns(df):
    if df is None or not hasattr(df, "columns") or len(df.columns) == 0:
        raise ValueError("No comparable columns are available.")
    question_col = df.columns[0]
    answer_col = df.columns[1] if len(df.columns) > 1 else df.columns[0]
    return question_col, answer_col


def _should_use_best_match_alignment(comparison_mode, uploaded_file1, uploaded_file2):
    if comparison_mode != "Compare Any Two Files":
        return False
    return (
        (uploaded_file1 is not None and not _is_spreadsheet_upload(uploaded_file1))
        or (uploaded_file2 is not None and not _is_spreadsheet_upload(uploaded_file2))
    )


def _should_use_question_context(comparison_mode, uploaded_file1=None, uploaded_file2=None):
    if comparison_mode == "Compare Two Columns in Same Excel File":
        return True
    if comparison_mode == "Compare Two Excel Files":
        return True
    if comparison_mode == "Compare Any Two Files":
        return bool(
            uploaded_file1 is not None
            and uploaded_file2 is not None
            and _is_spreadsheet_upload(uploaded_file1)
            and _is_spreadsheet_upload(uploaded_file2)
        )
    return False


def _alignment_seed_score(text1, text2):
    left = str(text1 or "").strip()
    right = str(text2 or "").strip()
    if not left and not right:
        return 0.0
    if not left or not right:
        return -1.0

    lexical_score = _lexical_similarity_percent(left, right)
    overlap_score = _significant_overlap_ratio(left, right) * 100.0
    seed_score = (0.85 * lexical_score) + (0.15 * overlap_score)

    relaxed_left = _normalize_similarity_text(left, relax_numbers=True)
    relaxed_right = _normalize_similarity_text(right, relax_numbers=True)
    if relaxed_left and relaxed_left == relaxed_right:
        seed_score += 10.0
    elif _is_short_subset_match(left, right):
        seed_score += 6.0

    return round(max(0.0, min(100.0, seed_score)), 2)


def _build_best_match_alignment(question_labels1, answers1, question_labels2, answers2):
    questions1 = [str(value or "") for value in (question_labels1 or [])]
    answers1 = [str(value or "") for value in (answers1 or [])]
    questions2 = [str(value or "") for value in (question_labels2 or [])]
    answers2 = [str(value or "") for value in (answers2 or [])]

    candidate_pairs = []
    for left_index, left_answer in enumerate(answers1):
        for right_index, right_answer in enumerate(answers2):
            seed_score = _alignment_seed_score(left_answer, right_answer)
            candidate_pairs.append((seed_score, left_index, right_index))

    candidate_pairs.sort(
        key=lambda item: (item[0], -abs(item[1] - item[2]), -item[1], -item[2]),
        reverse=True,
    )

    left_to_right = {}
    used_left = set()
    used_right = set()

    for seed_score, left_index, right_index in candidate_pairs:
        if left_index in used_left or right_index in used_right:
            continue
        left_to_right[left_index] = (right_index, seed_score)
        used_left.add(left_index)
        used_right.add(right_index)
        if len(used_left) == min(len(answers1), len(answers2)):
            break

    aligned_rows = []
    for left_index in range(len(answers1)):
        if left_index in left_to_right:
            right_index, seed_score = left_to_right[left_index]
            aligned_rows.append(
                {
                    "question1": questions1[left_index],
                    "answer1": answers1[left_index],
                    "question2": questions2[right_index],
                    "answer2": answers2[right_index],
                    "alignment_seed_score": seed_score,
                }
            )
        else:
            aligned_rows.append(
                {
                    "question1": questions1[left_index],
                    "answer1": answers1[left_index],
                    "question2": "",
                    "answer2": "",
                    "alignment_seed_score": None,
                }
            )

    for right_index in range(len(answers2)):
        if right_index in used_right:
            continue
        aligned_rows.append(
            {
                "question1": "",
                "answer1": "",
                "question2": questions2[right_index],
                "answer2": answers2[right_index],
                "alignment_seed_score": None,
            }
        )

    return aligned_rows


def read_uploaded_file(uploaded_file, sheet_name=None):
    """Read spreadsheet and non-spreadsheet uploads into a comparable DataFrame."""
    if uploaded_file is None:
        return None

    extension = _uploaded_file_extension(uploaded_file)

    if extension == ".csv":
        return _read_csv_with_encodings(uploaded_file)

    if extension in {".xlsx", ".xls"}:
        return _read_excel_upload(uploaded_file, sheet_name=sheet_name)

    if extension in JSON_FILE_EXTENSIONS:
        return _build_json_upload_dataframe(uploaded_file)

    if extension in TEXT_FILE_EXTENSIONS:
        return _build_text_upload_dataframe(uploaded_file)

    if extension in WORD_FILE_EXTENSIONS:
        return _build_docx_upload_dataframe(uploaded_file)

    if extension in PDF_FILE_EXTENSIONS:
        return _build_pdf_upload_dataframe(uploaded_file)

    if extension in LEGACY_WORD_FILE_EXTENSIONS:
        raise ValueError("Legacy .doc files are not supported directly. Save the file as .docx, .txt, or .pdf and retry.")

    try:
        return _read_excel_upload(uploaded_file, sheet_name=sheet_name)
    except Exception as excel_exc:
        try:
            return _read_csv_with_encodings(uploaded_file)
        except Exception:
            raise ValueError(
                f"Unsupported or unreadable file type '{extension or 'unknown'}'. "
                "Use Excel/CSV, JSON, TXT, PDF, or DOCX."
            ) from excel_exc


def _build_export_summary(df, threshold, primary_sim_col=None):
    if df is None or df.empty:
        return {
            "total_pairs": 0,
            "above_threshold": 0,
            "between_40_threshold": 0,
            "below_40": 0,
            "average_similarity": 0,
            "threshold": threshold,
        }

    sim_col = primary_sim_col if primary_sim_col in df.columns else None
    if sim_col is None:
        sim_cols = [c for c in df.columns if "Similarity" in str(c)]
        sim_col = sim_cols[0] if sim_cols else None

    if sim_col and sim_col in df.columns:
        numeric_sim = pd.to_numeric(df[sim_col], errors="coerce")
        total_pairs = int(numeric_sim.notna().sum())
        above_thresh = int((numeric_sim >= threshold).sum())
        between_40_thresh = int(((numeric_sim >= 40) & (numeric_sim < threshold)).sum())
        below_40 = int((numeric_sim < 40).sum())
        avg_similarity = round(float(numeric_sim.mean()), 2) if total_pairs > 0 else 0
    else:
        total_pairs = len(df)
        above_thresh = 0
        between_40_thresh = 0
        below_40 = 0
        avg_similarity = 0

    return {
        "total_pairs": total_pairs,
        "above_threshold": above_thresh,
        "between_40_threshold": between_40_thresh,
        "below_40": below_40,
        "average_similarity": avg_similarity,
        "threshold": threshold,
    }


def _build_non_excel_export_df(df):
    export_df = df.copy()
    drop_cols = [c for c in export_df.columns if "(diff)" in str(c)]
    if drop_cols:
        export_df = export_df.drop(columns=drop_cols, errors="ignore")

    for col in export_df.columns:
        if "Similarity" in str(col):
            numeric = pd.to_numeric(export_df[col], errors="coerce")
            export_df[col] = numeric.apply(lambda v: f"{float(v):.2f}%" if pd.notnull(v) else "")
    return export_df


def _build_non_excel_json_payload(df, summary, uploaded_file1=None, uploaded_file2=None):
    payload = {
        "export_type": "similarity_results",
        "input_files": [
            getattr(uploaded_file1, "name", "") if uploaded_file1 is not None else "",
            getattr(uploaded_file2, "name", "") if uploaded_file2 is not None else "",
        ],
        "summary": summary,
        "rows": json.loads(df.to_json(orient="records", force_ascii=False)),
    }
    return json.dumps(payload, ensure_ascii=False, indent=2)


def _build_non_excel_html_payload(df, summary, title):
    safe_title = html.escape(title)
    table_html = df.to_html(index=False, escape=True)
    summary_rows = [
        ("Total Pairs", summary["total_pairs"]),
        (f"Above {summary['threshold']}%", summary["above_threshold"]),
        (f"Between 40-{summary['threshold']}%", summary["between_40_threshold"]),
        ("Below 40%", summary["below_40"]),
        ("Average Similarity (%)", summary["average_similarity"]),
    ]
    summary_items = "".join(
        f"<li><strong>{html.escape(str(label))}:</strong> {html.escape(str(value))}</li>"
        for label, value in summary_rows
    )
    return f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <title>{safe_title}</title>
  <style>
    body {{ font-family: 'Segoe UI', Arial, sans-serif; margin: 24px; color: #0f172a; background: #f8fafc; }}
    h1 {{ margin-bottom: 8px; }}
    .summary {{ background: #ffffff; border: 1px solid #cbd5e1; border-radius: 10px; padding: 16px 20px; margin-bottom: 20px; }}
    .summary ul {{ margin: 0; padding-left: 20px; }}
    table {{ width: 100%; border-collapse: collapse; background: #ffffff; }}
    th, td {{ border: 1px solid #dbe4ee; padding: 10px 12px; text-align: left; vertical-align: top; }}
    th {{ background: #e2e8f0; }}
    tr:nth-child(even) td {{ background: #f8fafc; }}
  </style>
</head>
<body>
  <h1>{safe_title}</h1>
  <div class="summary">
    <ul>{summary_items}</ul>
  </div>
  {table_html}
</body>
</html>"""




# Modern CSS with gradient backgrounds, card-based design, and smooth animations
st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
    
    * {
        font-family: 'Inter', -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
    }
    
    /* Background - slate-50 to blue-50 gradient */
    .main, .stApp {
        background: linear-gradient(to bottom right, #f8fafc, #eff6ff);
        background-attachment: fixed;
    }
    
    .block-container {
        padding-top: 1rem;
        padding-left: 3rem;
        padding-right: 3rem;
        padding-bottom: 3rem;
        max-width: none;
        width: 100%;
    }
    
    /* Header */
    .header-card {
        background: transparent;
        padding: 0 1rem;
        margin-bottom: 1rem;
        margin-top: 0;
        border: none;
        box-shadow: none;
        text-align: center;
    }
    
    .header-title {
        font-size: 2.25rem;
        font-weight: 700;
        color: #1e293b;
        margin-bottom: 0.5rem;
        text-align: center;
    }
    
    .header-subtitle {
        text-align: center;
        /* default text color */
        font-size: 1rem;
        margin-bottom: 0;
    }
    
    /* Step Cards - white bg with subtle shadow */
    .step-card {
        background: white;
        padding: 1.5rem;
        border-radius: 0.75rem;
        margin-bottom: 1.5rem;
        box-shadow: 0 1px 3px rgba(0, 0, 0, 0.1);
        border: 1px solid #e2e8f0;
        transition: none;
    }

    /* Compact variant for smaller step boxes */
    .step-card.small {
        padding: 0.65rem 0.9rem;
        border-radius: 0.6rem;
        margin-bottom: 1rem;
    }
    
    .step-card:hover {
        transform: none;
        box-shadow: 0 1px 3px rgba(0, 0, 0, 0.1);
    }
    
    /* Step number - blue circle */
    .step-number {
        display: inline-flex;
        align-items: center;
        justify-content: center;
        background: #f1f5f9;
        /* default text color */
        width: 28px;
        height: 28px;
        border-radius: 50%;
        font-weight: 600;
        font-size: 0.95rem;
        margin-right: 0.5rem;
        vertical-align: middle;
        box-shadow: none;
    }

    /* Progress bar neutral gray overrides */
    /* Streamlit progress bar has different DOM across versions; provide several selectors */
    /* Use Streamlit's default progress styling */
    
    .step-title {
        font-size: 1.05rem; /* moderate heading for readability */
        font-weight: 600;
        color: #1e293b;
        display: inline-block;
        vertical-align: middle;
    }

    /* Step card inner body: keep the radio control and description inside the white card */
    .step-card-body {
        background: white;
        padding-top: 0.5rem;
        margin-top: 0.5rem;
    }

    /* Make the very next radio widget and info-box appear visually inside the step card */
    .step-card + .stRadio, .step-card + .info-box {
    /* Compare row spacing and larger buttons */
    .compare-row { margin-top: 1.25rem; }
    .compare-row .stButton>button, .compare-row .stDownloadButton>button {
        padding: 0.9rem 1.2rem !important;
        font-size: 1.03rem !important;
        border-radius: 0.6rem !important;
    }
        background: white !important;
        padding: 0.75rem 1rem !important;
        border-radius: 0 0 12px 12px !important;
        margin-top: -0.5rem !important;
        border-top: 1px solid #e2e8f0 !important;
    }

    /* Ensure the radio labels use full width inside the card */
    .step-card + .stRadio > div > label {
        background: transparent !important;
        border-color: transparent !important;
        box-shadow: none !important;
    }
    
    /* Info boxes - blue-50 bg */
    .info-box {
        background: #eff6ff;
        border: 1px solid #bfdbfe;
        border-left: none;
        padding: 1rem;
        border-radius: 0.5rem;
        margin: 1rem 0;
        color: #1e3a8a;
        font-size: 0.9rem; /* readable body text */
    }
    
    /* Primary buttons - blue-600 */
    .stButton>button {
        color: white;
        border: none;
        border-radius: 0.5rem;
        padding: 0.75rem 1.5rem;
        font-weight: 600;
        font-size: 1rem;
        transition: all 0.2s ease;
        box-shadow: 0 4px 6px -1px rgba(37, 99, 235, 0.3);
        width: 100%;
    }
    
    .stButton>button:hover {
        transform: translateY(-1px);
        box-shadow: 0 10px 15px -3px rgba(37, 99, 235, 0.4);
    }
    
    .stButton>button:disabled {
        /* default background */
        cursor: not-allowed;
        transform: none;
        box-shadow: none;
    }
    
    /* Radio buttons - segmented control style */
    .stRadio > div {
        background: transparent;
        padding: 0;
        border-radius: 0;
        box-shadow: none;
        display: flex;
        gap: 0.75rem;
        flex-wrap: wrap;
    }
    
    .stRadio > div > label {
        background: #f8fafc;
        padding: 0.75rem 1rem;
        border-radius: 0.5rem;
        transition: all 0.2s ease;
        cursor: pointer;
        border: 1px solid #e2e8f0;
        font-size: 0.875rem;
        font-weight: 500;
        color: #475569;
        margin: 0;
    }
    
    .stRadio > div > label:hover {
        background: #f1f5f9;
        border-color: #e2e8f0;
    }
    
    .stRadio > div > label[data-baseweb="radio"] > div:first-child {
        display: none;
    }
    
    /* Selected radio - softened mid-contrast blue with white text */
    div[role="radiogroup"] label:has(input[type="radio"]:checked) {
        background: linear-gradient(90deg, #2b6cb0 0%, #3b82f6 100%) !important; /* mid-blue */
        color: #ffffff !important; /* keep white text for legibility */
        border: 1px solid rgba(59,130,246,0.85) !important; /* medium border */
        box-shadow: 0 6px 18px -8px rgba(59,130,246,0.12) !important; /* lighter shadow */
        font-weight: 700 !important;
        transform: translateY(-2px) !important;
    }

    /* Nested text and title should be white and bold for contrast but not oversized */
    div[role="radiogroup"] label:has(input[type="radio"]:checked) span,
    div[role="radiogroup"] label:has(input[type="radio"]:checked) .step-title {
        color: #ffffff !important;
        font-weight: 700 !important;
        font-size: 0.98rem !important;
    }
    
    /* Select boxes */
    .stSelectbox > div > div {
        background: #e3f2fd;
        border-radius: 6px;
        margin-bottom: 0.5em;
        min-height: 36px;
    }
    
    .stSelectbox select {
        font-size: 0.9em;
        padding: 0.3em 0.5em;
    }

    /* Keep help icon visible regardless of upstream SVG color overrides */
    [data-testid="stWidgetLabelHelp"] {
        position: relative !important;
        width: 18px !important;
        height: 18px !important;
        min-width: 18px !important;
        min-height: 18px !important;
        display: inline-flex !important;
        align-items: center !important;
        justify-content: center !important;
        border: 1px solid #94a3b8 !important;
        border-radius: 50% !important;
        background: transparent !important;
    }
    [data-testid="stWidgetLabelHelp"]::before {
        content: "?" !important;
        color: #334155 !important;
        font-size: 11px !important;
        font-weight: 700 !important;
        line-height: 1 !important;
        pointer-events: none !important;
    }
    [data-testid="stWidgetLabelHelp"] button {
        position: absolute !important;
        inset: 0 !important;
        width: 100% !important;
        height: 100% !important;
        opacity: 0 !important;
        background: transparent !important;
        border: none !important;
        margin: 0 !important;
        padding: 0 !important;
    }
    
    /* Text inputs */
    .stTextInput > div > div > input {
        background: white;
        border: 1px solid #e2e8f0;
        border-radius: 0.5rem;
        padding: 0.75rem;
        transition: all 0.2s ease;
        font-size: 1rem;
    }
    
    .stTextInput > div > div > input:focus {
        border-color: #e2e8f0;
        box-shadow: 0 0 0 3px rgba(148,163,184,0.12);
        outline: none;
    }
    
    /* File uploader */
    [data-testid="stFileUploader"] {
        background: white;
        padding: 0;
        border-radius: 0.5rem;
    }
    
    [data-testid="stFileUploader"] section {
        border: 2px dashed #e2e8f0;
        border-radius: 0.5rem;
        padding: 2rem;
        background: #f8fafc;
        transition: all 0.2s ease;
    }
    
    [data-testid="stFileUploader"] section:hover {
        border-color: #e2e8f0;
        background: #f1f5f9;
    }
    
    /* Metrics - colored stat cards */
    [data-testid="stMetricValue"] {
        font-size: 1.875rem;
        font-weight: 700;
    }
    
    [data-testid="stMetricLabel"] {
        font-size: 0.875rem;
        font-weight: 500;
        /* default color */
        text-transform: none;
        letter-spacing: normal;
        margin-top: 0.25rem;
    }
    
    /* Metric containers with colored backgrounds */
    [data-testid="metric-container"] {
        background: white;
        padding: 1rem;
        border-radius: 0.5rem;
        box-shadow: none;
        border: 1px solid #e2e8f0;
    }
    
    [data-testid="metric-container"]:nth-child(1) {
        background: #f8fafc;
    }
    
    [data-testid="metric-container"]:nth-child(1) [data-testid="stMetricValue"] {
        color: #1e293b;
    }
    
    [data-testid="metric-container"]:nth-child(2) {
        background: #f0fdf4;
    }
    
    [data-testid="metric-container"]:nth-child(2) [data-testid="stMetricValue"] {
        color: #15803d;
    }
    
    [data-testid="metric-container"]:nth-child(2) [data-testid="stMetricLabel"] {
        color: #16a34a;
    }
    
    [data-testid="metric-container"]:nth-child(3) {
        background: #fffbeb;
    }
    
    [data-testid="metric-container"]:nth-child(3) [data-testid="stMetricValue"] {
        color: #b45309;
    }
    
    [data-testid="metric-container"]:nth-child(3) [data-testid="stMetricLabel"] {
        color: #d97706;
    }
    
    [data-testid="metric-container"]:nth-child(4) {
        background: #fef2f2;
    }
    
    [data-testid="metric-container"]:nth-child(4) [data-testid="stMetricValue"] {
        color: #b91c1c;
    }
    
    [data-testid="metric-container"]:nth-child(4) [data-testid="stMetricLabel"] {
        color: #dc2626;
    }
    
    /* Expander */
    .streamlit-expanderHeader {
        background: white;
        border-radius: 0.5rem;
        padding: 1rem;
        font-weight: 600;
        color: #1e293b;
        border: 1px solid #e2e8f0;
    }
    
    .streamlit-expanderHeader:hover {
        border-color: #e2e8f0;
        background: #f8fafc;
    }
    
    /* Progress bar */
    .stProgress > div > div {
        border-radius: 0.5rem;
    }
    
    /* Slider */
    .stSlider > div > div > div > div {
    }
    
    /* Download buttons */
    .stDownloadButton > button {
        color: white;
        border: none;
        border-radius: 0.5rem;
        padding: 0.5rem 1rem;
        font-weight: 600;
        font-size: 0.875rem;
        transition: all 0.2s ease;
        width: 100%;
        display: flex;
        align-items: center;
        justify-content: center;
        gap: 0.5rem;
    }
    
    .stDownloadButton > button:hover {
        /* default background */
        transform: translateY(-1px);
        box-shadow: 0 4px 6px -1px rgba(15, 23, 35, 0.06);
    }
    
    /* Secondary download buttons (warning colors) */
    .stDownloadButton:nth-child(2) > button {
        background: #f59e0b;
    }
    
    .stDownloadButton:nth-child(2) > button:hover {
        background: #d97706;
    }
    
    .stDownloadButton:nth-child(3) > button {
        background: #ef4444;
    }
    
    .stDownloadButton:nth-child(3) > button:hover {
        background: #dc2626;
    }
    
    /* Number inputs */
    .stNumberInput > div > div > input {
        background: white;
        border: 1px solid #e2e8f0;
        border-radius: 0.375rem;
        padding: 0.5rem 0.75rem;
        font-size: 0.875rem;
    }
    
    /* Text areas (smaller padding for compact layout) */
    .stTextArea > div > div > textarea {
        background: white;
        border: 1px solid #e2e8f0;
        border-radius: 0.5rem;
        padding: 0.45rem 0.5rem;
        min-height: 56px;
    }
    
    /* Success/Info messages */
    .stSuccess {
        background: #f0fdf4;
        border: 1px solid #86efac;
        border-radius: 0.5rem;
        padding: 1rem;
        color: #15803d;
    }
    
    .stInfo {
        background: #eff6ff;
        border: 1px solid #93c5fd;
        border-radius: 0.5rem;
        padding: 1rem;
        color: #1e40af;
    }
    
    .stWarning {
        background: #fffbeb;
        border: 1px solid #fcd34d;
        border-radius: 0.5rem;
        padding: 1rem;
        color: #b45309;
    }
    
    /* Dataframe */
    .dataframe {
        border-radius: 0.5rem;
        overflow: hidden;
        border: 1px solid #e2e8f0;
    }
    
    /* Hide Streamlit branding */
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    
    /* Divider */
    hr {
        border: none;
        height: 1px;
        background: #e2e8f0;
        margin: 2rem 0;
    }
        /* Hide any generated sibling elements inside the numeric column wrappers so the rounded box only contains the number input */
        #gpt-numeric-row .stColumns > div > .css-1l02zno > :not(.stNumberInput),
        #gpt-numeric-row .stColumns > div > [data-testid="stNumberInput"] > :not(.stNumberInput) {
            display: none !important;
            width: 0 !important;
            margin: 0 !important;
            padding: 0 !important;
        }
        /* Ensure the visible rounded box fits tightly */
        #gpt-numeric-row .stColumns > div > .css-1l02zno, #gpt-numeric-row .stColumns > div > [data-testid="stNumberInput"] {
            padding-right: 8px !important;
            padding-left: 8px !important;
            justify-content: flex-start !important;
        }
    </style>
""", unsafe_allow_html=True)

# Header
st.markdown("""
<div class="header-card">
    <h1 class="header-title">Similarity Answer Matcher</h1>
    <p class="header-subtitle">Compare answers across files with AI-powered similarity analysis</p>
</div>
""", unsafe_allow_html=True)

# --- Ensure AI Advanced settings defaults exist early ---
# This prevents Streamlit errors when code attempts to set session_state
# after a widget with the same key has already been created.
def _ensure_ai_defaults():
    st.session_state['gpt_system_prompt'] = _normalized_prompt_value(
        st.session_state.get('gpt_system_prompt'),
        DEFAULT_SYSTEM_PROMPT,
    )
    st.session_state['gpt_user_template'] = _normalized_prompt_value(
        st.session_state.get('gpt_user_template'),
        DEFAULT_USER_PROMPT_TEMPLATE,
    )
    st.session_state.setdefault('gpt_temperature', 0.0)
    st.session_state.setdefault('gpt_top_p', 1.0)
    st.session_state.setdefault('gpt_max_tokens', 20)
    
    # Initialize model-specific parameters
    from ai_providers import PROVIDERS
    for model_name in PROVIDERS.keys():
        model_params = get_model_parameters(model_name)
        for param_name, param_config in model_params.items():
            key = f'{model_name}_{param_name}'
            st.session_state.setdefault(key, param_config.get('default', 0.0))
    # transient status key used by save/reset callbacks
    if '_gpt_save_status' not in st.session_state:
        st.session_state.setdefault('_gpt_save_status', None)


# Call early to avoid ordering issues with widget creation
_ensure_ai_defaults()

# --- Advanced AI provider settings persistence (moved early) ---
default_system = DEFAULT_SYSTEM_PROMPT
default_user_tpl = DEFAULT_USER_PROMPT_TEMPLATE

# Use working directory for persistent file (Streamlit may run from a temp copy)
SETTINGS_PATH = os.path.join(os.getcwd(), '.gpt_settings.json')

def _settings_user_key():
    try:
        return get_current_user_key()
    except Exception:
        raw = st.session_state.get("username") or "default"
        return str(raw).strip().lower() or "default"

def _load_settings_root():
    if not os.path.exists(SETTINGS_PATH):
        return {}
    try:
        with open(SETTINGS_PATH, 'r', encoding='utf-8') as f:
            data = json.load(f)
        root = data if isinstance(data, dict) else {}
        # Security migration: quarantine legacy global settings so they are not
        # auto-visible to any authenticated user.
        users = root.get('users')
        has_legacy = any(k in root for k in ('current', 'history', 'api_keys'))
        if has_legacy and not isinstance(users, dict):
            legacy_blob = {k: root.get(k) for k in ('current', 'history', 'api_keys') if k in root}
            root = {'users': {'__legacy_global__': legacy_blob}}
            with open(SETTINGS_PATH, 'w', encoding='utf-8') as wf:
                json.dump(root, wf, ensure_ascii=False, indent=2)
        return root
    except Exception:
        return {}

def _extract_user_settings(root, user_key):
    users = root.get('users')
    if isinstance(users, dict):
        data = users.get(user_key, {})
        return data if isinstance(data, dict) else {}
    return {}

def _persist_user_settings(user_settings):
    try:
        root = _load_settings_root()
        users = root.get('users')
        if not isinstance(users, dict):
            users = {}
        users[_settings_user_key()] = user_settings if isinstance(user_settings, dict) else {}
        root['users'] = users
        # Drop legacy top-level keys once user-scoped structure exists.
        root.pop('current', None)
        root.pop('history', None)
        root.pop('api_keys', None)
        with open(SETTINGS_PATH, 'w', encoding='utf-8') as f:
            json.dump(root, f, ensure_ascii=False, indent=2)
        return True
    except Exception:
        return False

def load_gpt_settings():
    root = _load_settings_root()
    return _extract_user_settings(root, _settings_user_key())

def save_gpt_settings(data):
    try:
        existing = load_gpt_settings() or {}
        new_struct = existing if isinstance(existing, dict) else {}
        cur = new_struct.get('current')
        hist = new_struct.get('history', []) if isinstance(new_struct.get('history', []), list) else []
        if cur:
            hist = [cur] + hist
            hist = hist[:2]

        new_struct['history'] = hist
        if 'api_keys' not in new_struct or not isinstance(new_struct.get('api_keys'), dict):
            new_struct['api_keys'] = {}

        current_model = data.get('matching_method')
        current_api_key = data.get('api_key')
        if current_model:
            if current_api_key:
                new_struct['api_keys'][current_model] = current_api_key
            elif current_model in new_struct['api_keys']:
                # Keep persisted key unless explicitly cleared by reset.
                pass

        new_struct['current'] = {
            'gpt_system_prompt': _normalized_prompt_value(
                data.get('gpt_system_prompt'),
                DEFAULT_SYSTEM_PROMPT,
            ),
            'gpt_user_template': _normalized_prompt_value(
                data.get('gpt_user_template'),
                DEFAULT_USER_PROMPT_TEMPLATE,
            ),
            'gpt_temperature': data.get('gpt_temperature'),
            'gpt_top_p': data.get('gpt_top_p'),
            'gpt_max_tokens': data.get('gpt_max_tokens'),
            'matching_method': current_model,
        }

        return _persist_user_settings(new_struct)
    except Exception:
        return False

def reset_gpt_settings():
    try:
        return _persist_user_settings({})
    except Exception:
        return False

# Load saved settings now (before widgets) and assign into session_state so they persist
try:
    import json
    saved = load_gpt_settings()
    if saved:
        current = saved.get('current') if isinstance(saved.get('current'), dict) else saved
        # assign (do not use setdefault) so saved values override the early defaults
        st.session_state['gpt_system_prompt'] = _normalized_prompt_value(
            current.get('gpt_system_prompt'),
            default_system,
        )
        st.session_state['gpt_user_template'] = _normalized_prompt_value(
            current.get('gpt_user_template'),
            default_user_tpl,
        )
        st.session_state['gpt_temperature'] = current.get('gpt_temperature', 0.0)
        st.session_state['gpt_top_p'] = current.get('gpt_top_p', 1.0)
        st.session_state['gpt_max_tokens'] = current.get('gpt_max_tokens', 20)
        
        # Load saved API keys for all models
        if 'api_keys' in saved and isinstance(saved.get('api_keys'), dict):
            st.session_state['_saved_api_keys'] = saved['api_keys']
except Exception:
    pass

# Apply model-scoped advanced reset before Step 2 widgets are created.
# This avoids Streamlit errors when updating widget-bound session-state keys.
if st.session_state.pop('_advanced_reset_pending', False):
    reset_ok = True
    reset_error = None
    reset_model_for_status = st.session_state.get('matching_method', 'Azure OpenAI GPT-4o')
    try:
        reset_model = st.session_state.pop(
            '_advanced_reset_model',
            st.session_state.get('matching_method', 'Azure OpenAI GPT-4o')
        )
        reset_model_for_status = reset_model

        # Reset only selected model parameters.
        _cfg = get_model_parameters(reset_model)
        for _param_name, _param_cfg in _cfg.items():
            _k = f'{reset_model}_{_param_name}'
            st.session_state[_k] = _param_cfg.get('default', 0.0)

        # Clear API key for selected model only.
        _api_key_state = f"api_key_{reset_model.replace(' ', '_')}"
        st.session_state[_api_key_state] = ""
        _saved_api_keys = st.session_state.get('_saved_api_keys', {})
        if isinstance(_saved_api_keys, dict) and reset_model in _saved_api_keys:
            del _saved_api_keys[reset_model]
        st.session_state['_saved_api_keys'] = _saved_api_keys if isinstance(_saved_api_keys, dict) else {}

        # Reset model/deployment chooser for selected provider only.
        _model_select_key = f"provider_model_select_{reset_model.replace(' ', '_')}"
        _custom_model_key = f"provider_model_custom_{reset_model.replace(' ', '_')}"
        _resolved_model_key = f"{reset_model}_resolved_model_name"
        _default_model = get_default_provider_model(reset_model)
        if _default_model:
            st.session_state[_model_select_key] = _default_model
            st.session_state[_resolved_model_key] = _default_model
            st.session_state['active_provider_model_name'] = _default_model
        st.session_state[_custom_model_key] = ""
        st.session_state['gpt_system_prompt'] = DEFAULT_SYSTEM_PROMPT
        st.session_state['gpt_user_template'] = DEFAULT_USER_PROMPT_TEMPLATE

        # Remove only selected model API key from persisted settings for current user.
        _saved = load_gpt_settings() or {}
        _saved = _saved if isinstance(_saved, dict) else {}
        if isinstance(_saved.get('api_keys'), dict) and reset_model in _saved['api_keys']:
            del _saved['api_keys'][reset_model]
        if 'api_keys' in _saved and not _saved.get('api_keys'):
            del _saved['api_keys']
        _saved_current = _saved.get('current')
        if not isinstance(_saved_current, dict):
            _saved_current = {}
            _saved['current'] = _saved_current
        _saved_current['gpt_system_prompt'] = DEFAULT_SYSTEM_PROMPT
        _saved_current['gpt_user_template'] = DEFAULT_USER_PROMPT_TEMPLATE
        if reset_model:
            _saved_current['matching_method'] = reset_model
        _persist_user_settings(_saved)
    except Exception as e:
        reset_ok = False
        reset_error = str(e)

    st.session_state['_advanced_action_status'] = 'reset_ok' if reset_ok else 'reset_fail'
    st.session_state['_advanced_action_error'] = reset_error
    st.session_state['_advanced_action_model'] = reset_model_for_status


# --- Comparison Mode ---
st.markdown("""
<div class="step-card small">
    <span class="step-number">1</span>
    <span class="step-title">Select Comparison Mode</span>
</div>
""", unsafe_allow_html=True)

# Display shorter/compact labels in a single row but map to the internal mode strings
display_mode_map = {
    "Compare Two Excel Files": "Compare Two Excel Files",
    "Compare Any Two Files": "Compare Any Two Files",
    "Compare Two Columns (Same File)": "Compare Two Columns in Same Excel File",
    "One Column vs Two Targets": "Compare One Column with Two Targets (Same Excel)",
}

if 'comparison_mode' not in st.session_state:
    # default internal mode
    st.session_state['comparison_mode'] = 'Compare Two Columns in Same Excel File'

selected_display_mode = st.selectbox(
    "Choose how you want to compare your data:",
    list(display_mode_map.keys()),
    index=list(display_mode_map.values()).index(st.session_state.get('comparison_mode', 'Compare Two Columns in Same Excel File')),
    help="Select the type of comparison you want to perform",
    key='comparison_mode_select'
)

# Map the compact display label back to the internal comparison_mode string and persist
comparison_mode = display_mode_map.get(selected_display_mode, "Compare Two Excel Files")
st.session_state['comparison_mode'] = comparison_mode

# Dynamic description
desc = {
    "Compare Two Excel Files": "Upload two Excel files with questions in the first column and answers in the second column.",
    "Compare Any Two Files": "Upload any two files (JSON, TXT, PDF, DOCX, or structured sheets). Select columns to compare answers.",
    "Compare Two Columns in Same Excel File": "Upload a single Excel file with at least three columns: one for questions and two for answers.",
    "Compare One Column with Two Targets (Same Excel)": "Upload a single Excel file: one base column compared separately with two target columns."
}

st.markdown(f'<div class="info-box">💡 {desc.get(comparison_mode, desc["Compare Two Excel Files"])}</div>', unsafe_allow_html=True)

IS_TWO_FILE_MODE = comparison_mode in ("Compare Two Excel Files", "Compare Any Two Files")

# --- Matching Method ---
st.markdown("""
<div class="step-card small">
    <span class="step-number">2</span>
    <span class="step-title">Choose AI Similarity Model</span>
</div>
""", unsafe_allow_html=True)

# Add CSS to align columns and form elements vertically
st.markdown("""
<style>
/* Align col_left and col_right on same baseline */
[data-testid="column"] > div:first-child > div:first-child {
    display: flex;
    flex-direction: column;
    justify-content: flex-start;
}

/* Make selectbox and text input have same height */
.stSelectbox > div > div > div, 
.stTextInput > div > div > input {
    min-height: 38px;
    height: 38px;
}

/* Align selectbox label with input label */
.stSelectbox label, 
.stTextInput label {
    margin-bottom: 0.5rem;
    font-weight: 500;
}

/* Uniform label style for Step 2 top-row fields */
.step2-field-label {
    font-size: 1.04rem;
    font-weight: 500;
    color: #0f172a;
    margin-bottom: 0.5rem;
    line-height: 1.2;
    min-height: 1.2rem;
}

/* Fix number inputs to align with text inputs */
.stNumberInput > div > div > input {
    height: 38px;
    min-height: 38px;
}

/* Align buttons vertically with number inputs */
.stButton > button {
    height: 38px;
    min-height: 38px;
    margin-top: 1.8rem;
}

/* Fix textarea alignment */
.stTextArea > div > div > textarea {
    margin-bottom: 0;
}

/* Ensure numeric row buttons align properly */
#gpt-numeric-row [data-testid="column"] {
    align-items: flex-end;
}

/* Remove excessive margins that cause misalignment */
.stSelectbox, .stTextInput, .stNumberInput {
    margin-bottom: 0;
}

/* Align first label of number input row with others */
.stNumberInput > label {
    margin-bottom: 0.5rem;
}
</style>
""", unsafe_allow_html=True)

method_col, api_col_main, model_col_main, icon_col_main = st.columns([3.8, 3.1, 1.9, 0.7], gap='small')

# Read matching choice from query params (cards use ?matching=... links)
# Only use URL param to set an initial default if the session state has no selection yet.
provider_aliases = {
    "OpenAI GPT-4o": "OpenAI GPT-4o",
    "Groq": "Groq",
    "OpenRouter": "OpenRouter",
    "xAI Grok": "xAI Grok",
    "Google Gemini": "Google Gemini",
    "Anthropic Claude": "Anthropic Claude",
    "OpenAI GPT-4o (API)": "OpenAI GPT-4o",
    "Groq (API)": "Groq",
    "OpenRouter (API)": "OpenRouter",
    "xAI Grok (API)": "xAI Grok",
    "Google Gemini (API)": "Google Gemini",
    "Anthropic Claude (API)": "Anthropic Claude",
    "Gemini": "Google Gemini",
    "Grok": "xAI Grok",
    "Claude 3 Opus": "Anthropic Claude",
    "OpenAI GPT-4o-mini": "OpenAI GPT-4o-mini",
}

matching_display_map = {
    "Azure OpenAI GPT-4o": "Azure OpenAI GPT-4o",
    "OpenAI GPT-4o": "OpenAI GPT-4o",
    "Groq": "Groq",
    "OpenRouter": "OpenRouter",
    "xAI Grok": "xAI Grok",
    "Google Gemini": "Google Gemini",
    "Anthropic Claude": "Anthropic Claude",
}

params = st.query_params
if 'matching' in params and params.get('matching'):
    val = provider_aliases.get(params.get('matching')[0], params.get('matching')[0])
    from ai_providers import PROVIDERS
    if val in PROVIDERS:
        # Only set from URL if session state has no value yet (don't override user actions)
        if 'matching_method' not in st.session_state:
            st.session_state['matching_method'] = val
            try:
                st.experimental_rerun()
            except Exception:
                pass

# Ensure a default selection exists
if 'matching_method' not in st.session_state:
    st.session_state['matching_method'] = 'Azure OpenAI GPT-4o'

# Step 2 method selector (left-most control in the same top row)
with method_col:
    # Use the internal labels as radio options; we'll inject CSS to show badges and subtitles
    from ai_providers import PROVIDERS
    matching_display_options = list(matching_display_map.keys())
    # Migrate older stored display labels (with "(API)") to the new labels.
    legacy_display_map = {
        "OpenAI GPT-4o (API)": "OpenAI GPT-4o",
        "Groq (API)": "Groq",
        "OpenRouter (API)": "OpenRouter",
        "xAI Grok (API)": "xAI Grok",
        "Google Gemini (API)": "Google Gemini",
        "Anthropic Claude (API)": "Anthropic Claude",
    }
    if st.session_state.get("matching_method_select") in legacy_display_map:
        st.session_state["matching_method_select"] = legacy_display_map[st.session_state["matching_method_select"]]
    # Determine default index from session_state (default to Azure)
    current = st.session_state.get('matching_method', 'Azure OpenAI GPT-4o')
    current = provider_aliases.get(current, current)
    current_display = next((label for label, value in matching_display_map.items() if value == current), None)
    if current_display not in matching_display_options:
        current_display = matching_display_options[0]
        current = matching_display_map[current_display]
        st.session_state['matching_method'] = current
    # Determine default index from internal value
    try:
        default_idx = matching_display_options.index(current_display)
    except ValueError:
        default_idx = 0  # Default to first provider (Azure)

    # Provider selectbox for similarity backend
    st.markdown("<div class='step2-field-label'>Choose AI Similarity Model</div>", unsafe_allow_html=True)
    sel = st.selectbox(
        "Choose AI Similarity Model",
        matching_display_options,
        index=default_idx,
        key='matching_method_select',
        help='Select the AI model backend used for similarity scoring.',
        label_visibility="collapsed",
    )
    selected_provider = matching_display_map.get(sel, sel)
    st.session_state['matching_method'] = selected_provider
    st.session_state['matching_radio'] = selected_provider

# Allow selection via the cards above: use session_state to store selection
if 'matching_method' not in st.session_state:
    st.session_state['matching_method'] = 'Azure OpenAI GPT-4o'

matching_method = st.session_state.get('matching_method', 'Azure OpenAI GPT-4o')
selected_model = None
local_backend_ready = _has_sentence_transformers()

if matching_method == "Local Model":
    matching_method = "Azure OpenAI GPT-4o"
    st.session_state['matching_method'] = matching_method
    st.session_state['matching_method_select'] = matching_method

if matching_method == "Local Model":
    with api_col_main:
        available_local_models = _get_available_local_models() if local_backend_ready else {}
        st.markdown("<div class='step2-field-label'>Select Local Language Model</div>", unsafe_allow_html=True)
        if not local_backend_ready:
            selected_model = None
            st.error(
                "Local Model backend is not available in the Python runtime used by Streamlit."
            )
            st.caption(
                f"Install in this runtime: `{sys.executable} -m pip install sentence-transformers torch`"
            )
        elif not available_local_models:
            selected_model = None
            st.error(
                "No installed local models found. Install/cache a local sentence-transformer model first "
                "(e.g., all-mpnet-base-v2), then retry."
            )
        else:
            model_options = list(available_local_models.keys())
            # default to all-mpnet-base-v2 (Accurate) when available
            try:
                default_idx = model_options.index("all-mpnet-base-v2 (Accurate)")
            except ValueError:
                default_idx = 0
            selected_model_name = st.selectbox(
                "Select Local Language Model",
                model_options,
                index=default_idx,
                key="local_model_select_main",
                help="Only locally installed models are listed.",
                label_visibility="collapsed",
            )
            selected_model = available_local_models[selected_model_name]
        # Selecting a local model should close any cloud provider advanced settings
        st.session_state['show_advanced_gpt'] = False
else:
        # Cloud controls in a single row: API key | model | advanced settings
        model_select_key = f"provider_model_select_{matching_method.replace(' ', '_')}"
        custom_model_key = f"provider_model_custom_{matching_method.replace(' ', '_')}"
        resolved_model_key = f"{matching_method}_resolved_model_name"
        custom_option_label = "Custom deployment/model..."
        default_model = get_default_provider_model(matching_method)

        with api_col_main:
            # Load saved API key for this model if available
            saved_api_keys = st.session_state.get('_saved_api_keys', {})
            saved_key = saved_api_keys.get(matching_method, '')
            
            # Use per-model session key to avoid cross-contamination
            model_key = f'api_key_{matching_method.replace(" ", "_")}'
            if model_key not in st.session_state:
                st.session_state[model_key] = saved_key
            
            st.markdown("<div class='step2-field-label'>API Key *</div>", unsafe_allow_html=True)
            api_key = st.text_input(
                "API Key *",
                value=st.session_state.get(model_key, ''),
                type="password",
                placeholder=f"Enter your {matching_method} API key",
                help=f"Enter your {matching_method} API key. Click 'Save' to persist this key.",
                label_visibility="collapsed",
            )
            st.session_state[model_key] = api_key

        with model_col_main:
            provider_model_options = get_provider_model_options(matching_method)
            if provider_model_options:
                saved_model = st.session_state.get(resolved_model_key, default_model)
                model_choices = provider_model_options + [custom_option_label]

                if model_select_key not in st.session_state:
                    st.session_state[model_select_key] = saved_model if saved_model in provider_model_options else custom_option_label
                if st.session_state.get(model_select_key) not in model_choices:
                    st.session_state[model_select_key] = provider_model_options[0]

                model_label = "Model/Deployment"
                st.markdown(f"<div class='step2-field-label'>{model_label}</div>", unsafe_allow_html=True)
                selected_provider_option = st.selectbox(
                    model_label,
                    model_choices,
                    key=model_select_key,
                    help="Choose the deployment/model used for similarity scoring.",
                    label_visibility="collapsed",
                )

                resolved_provider_model = selected_provider_option
                if selected_provider_option == custom_option_label:
                    if custom_model_key not in st.session_state:
                        st.session_state[custom_model_key] = saved_model if saved_model not in provider_model_options else ""
                    custom_model_input = st.text_input(
                        "Custom deployment/model",
                        key=custom_model_key,
                        placeholder="Enter deployment/model name",
                    )
                    resolved_provider_model = custom_model_input.strip() or default_model

                st.session_state[resolved_model_key] = resolved_provider_model
                st.session_state["active_provider_model_name"] = resolved_provider_model
            else:
                st.session_state[resolved_model_key] = ""
                st.session_state["active_provider_model_name"] = ""
        # Inline icon column
        with icon_col_main:
            st.markdown(
                """
                <style>
                /* Small inline gear: use flexbox and relative positioning to align with input's center */
                .adv-inline-small {
                    display: flex !important;
                    align-items: center !important;
                    justify-content: center !important;
                    height: 100% !important;
                    padding: 0 !important;
                    margin: 0 !important;
                }
                .adv-inline-small .stButton > button {
                    background: transparent !important;
                    border: none !important;
                    box-shadow: none !important;
                    padding: 0px !important;
                    margin: 0 0 0 6px !important;
                    font-size: 0.95rem !important;
                    /* default color */
                    height: 32px !important;
                    width: 32px !important;
                    display: inline-flex !important;
                    align-items: center !important;
                    justify-content: center !important;
                    border-radius: 6px !important;
                    line-height: 1 !important;
                    transform: translateY(3px) !important;
                    position: relative !important;
                    top: 1px !important;
                }
                .adv-inline-small .stButton > button:hover {
                    background: rgba(100,116,139,0.06) !important;
                }
                /* Ensure the container cell vertically centers its content */
                .css-1l02zno.e1tzin5v1, .css-1l02zno { display: flex !important; align-items: center !important; }
                """,
                unsafe_allow_html=True,
            )
            st.markdown('<div class="adv-inline-small">', unsafe_allow_html=True)
            if st.button("⚙️", key="adv_settings_icon_inline", help="Advanced AI Settings"):
                st.session_state['show_advanced_gpt'] = not st.session_state.get('show_advanced_gpt', False)
            st.markdown('</div>', unsafe_allow_html=True)

        # Initialize session state for advanced settings visibility
        if 'show_advanced_gpt' not in st.session_state:
            st.session_state['show_advanced_gpt'] = False

        # Clickable text link for advanced settings using form button (avoids Streamlit button styling issues)
        st.markdown("""
        <style>
        /* Hide the form button container borders and styling and remove gaps */
        .stForm {
            border: none !important;
            padding: 0 !important;
            margin: 0 !important;
            display: inline-block !important;
        }
        /* Ensure text input has no bottom margin so the link sits flush */
        .stTextInput {
            margin-bottom: 0 !important;
        }
        /* Style form submit button as a simple blue text link with gear */
        .stForm button[kind="formSubmit"] {
            background: transparent !important;
            border: none !important;
            box-shadow: none !important;
            padding: 0 !important;
            margin: 0 !important;
            /* default color */
            font-size: 0.9rem !important;
            font-weight: 500 !important;
            height: auto !important;
            min-height: 0 !important;
            width: auto !important;
            display: inline-flex !important;
            align-items: center !important;
            gap: 0.4rem !important;
        }
        .stForm button[kind="formSubmit"]:hover {
            background: transparent !important;
            text-decoration: underline !important;
            /* default color */
            border: none !important;
            box-shadow: none !important;
            transform: none !important;
        }
        .stForm button[kind="formSubmit"]:focus,
        .stForm button[kind="formSubmit"]:active {
            background: transparent !important;
            border: none !important;
            box-shadow: none !important;
        }
        /* Hide any small separators Streamlit may insert between widgets */
        .stTextInput + .stForm, .stTextInput + form, .stTextInput + div > .stForm { margin-top: 0 !important; padding-top: 0 !important; }
        /* Remove extra vertical gaps after the API key field */
        .stTextInput, .stTextInput > div, .stTextInput > div > div { margin-bottom: 0 !important; padding-bottom: 0 !important; }
        /* Hide hr separators and small spacer divs often rendered between sections */
        hr, .block-container > hr, div[role="separator"] { display: none !important; height: 0 !important; margin: 0 !important; padding: 0 !important; }
        .stDivider { margin: 0 !important; padding: 0 !important; height: 0 !important; }
        /* Make expander header and advanced controls sit flush under the API key */
        .streamlit-expanderHeader, .stExpander { margin-top: 0 !important; padding-top: 0 !important; }
        </style>
        """, unsafe_allow_html=True)
        
        # (previous below-icon layout removed; icon now inline to the right of input)

    # Removed explicit spacer to eliminate extra vertical gap under API key

# Show advanced settings section if toggled (outside columns to span full width)
if st.session_state.get('show_advanced_gpt', False):
    st.markdown(
        """
    <style>
      /* Target the nearby Streamlit containers to remove card/separator appearance */
      .stTextInput, .stForm, .stNumberInput, .stTextArea, .stSelectbox {
          background: transparent !important;
          border: none !important;
          box-shadow: none !important;
          padding: 0 !important;
          margin: 0 !important;
      }
      /* Remove any horizontal rules and separators that create visual gaps */
      .stTextInput + hr, .stForm + hr, hr, .stForm + .streamlit-expanderHeader {
          display: none !important;
          height: 0 !important;
          margin: 0 !important;
          padding: 0 !important;
          border: none !important;
      }
      /* Ensure the advanced toggle button appears as inline link and flush with API field */
      .stForm {
          margin: 0 !important;
          padding: 0 !important;
      }
      .stForm button[kind="formSubmit"] {
          margin: 0 !important;
          padding: 0 !important;
          background: transparent !important;
          border: none !important;
          box-shadow: none !important;
          /* default color */
          font-size: 0.9rem !important;
      }
      /* Make any nearby step-card separators invisible to avoid boxed look */
      .step-card, .step-card + .stRadio, .info-box {
          border: none !important;
          box-shadow: none !important;
          background: transparent !important;
      }
      /* Header styling for Advanced GPT label: blue text, no border */
    .adv-header {font-weight:600;font-size:0.95rem;margin:0;padding:0;border:none;background:transparent}
    </style>
        """,
        unsafe_allow_html=True,
    )

    default_system = (
        "You are a helpful assistant, Provided the similarity score by comparing text 1 and"
        " text 2, just provide only similarity score without explaination"
    )

    default_user_tpl = (
        "Compare the following two texts and provide a similarity score as a percentage."
        " Text 1: {answer1} Text 2: {answer2}"
    )

    # (Persistence helpers are defined once earlier and are user-scoped.)

    ta_col_left, ta_col_right = st.columns([1, 1])
    with ta_col_left:
        st.markdown(
            "<div style='font-weight:600;font-size:0.85rem;color:#0f172a;margin-bottom:0.12rem;'>System Prompt (All AI Providers)</div>",
            unsafe_allow_html=True,
        )
        gpt_system_prompt = st.text_area(
            "System Prompt",
            height=64,
            key="gpt_system_prompt",
            label_visibility="collapsed",
            help="System-level instructions for all cloud providers: Azure OpenAI, OpenAI, Gemini, Claude, Groq, OpenRouter, and xAI Grok."
        )
    with ta_col_right:
        st.markdown(
            "<div style='font-weight:600;font-size:0.85rem;color:#0f172a;margin-bottom:0.12rem;'>User Prompt Template (All AI Providers)</div>",
            unsafe_allow_html=True,
        )
        gpt_user_template = st.text_area(
            "User Prompt Template",
            height=64,
            key="gpt_user_template",
            label_visibility="collapsed",
            help="Template for the user message. Use {question}, {answer1}, and {answer2} as placeholders. Applies to all providers."
        )

    # (Save/Reset buttons moved next to Max Tokens input for visibility)

    # keep minimal separation but remove visible spacer elements
    # spacer removed to keep advanced controls flush with the API key

    # Handle anchor-based +/- clicks via URL query params so controls render inline
    # No query-param handlers — use native number inputs and steppers

    # Add CSS to align numeric inputs and buttons vertically
    st.markdown("""
    <style>
    /* Fix alignment of numeric inputs row */
    #gpt-numeric-row [data-testid="column"] {
        display: flex;
        flex-direction: column;
        justify-content: flex-end;
        padding-top: 0 !important;
    }
    
    #gpt-numeric-row .stNumberInput {
        margin-bottom: 0 !important;
    }
    
    #gpt-numeric-row .stNumberInput > div > div > input {
        height: 38px;
        min-height: 38px;
    }
    
    #gpt-numeric-row [data-testid="column"]:last-child {
        justify-content: flex-end;
    }
    
    #gpt-numeric-row .stButton > button {
        height: 38px;
        min-height: 38px;
        margin-top: 1.65rem;
    }
    </style>
    """, unsafe_allow_html=True)
    
    # Get the current model to determine which parameters to show
    from ai_providers import PROVIDERS
    matching_method = st.session_state.get('matching_method', 'Azure OpenAI GPT-4o')
    model_params = get_model_parameters(matching_method)
    
    # Handle reset request (set flag before widgets are created)
    if st.session_state.get('_reset_model_params', False):
        _target_model = st.session_state.get('matching_method', matching_method)
        model_config = get_model_parameters(_target_model)
        for param_name, param_config in model_config.items():
            key = f'{_target_model}_{param_name}'
            st.session_state[key] = param_config.get('default', 0.0)
        st.session_state['_reset_model_params'] = False
        st.session_state['_advanced_action_status'] = 'reset_ok'
        st.session_state['_advanced_action_model'] = _target_model
    
    # Display dynamic model-specific parameters
    if model_params:
        st.markdown(
            f"<div style='font-weight:600;font-size:0.9rem;color:#0f172a;margin-bottom:0.5rem;'>⚙️ {matching_method} Advanced Settings</div>",
            unsafe_allow_html=True,
        )

        # Fixed feedback location for Advanced Settings actions.
        adv_status = st.session_state.pop('_advanced_action_status', None)
        adv_error = st.session_state.pop('_advanced_action_error', None)
        adv_model = st.session_state.pop('_advanced_action_model', None)
        if adv_status == 'save_ok':
            st.success(f"Saved settings for {adv_model or matching_method}.")
        elif adv_status == 'save_fail':
            st.error(f"Save failed: {adv_error or 'could not persist settings.'}")
        elif adv_status == 'reset_ok':
            st.success(f"Reset complete for {adv_model or matching_method}: settings and API key cleared.")
        elif adv_status == 'reset_fail':
            st.error(f"Reset failed: {adv_error or 'some settings could not be cleared.'}")

        # Create one compact row for all parameters
        param_list = list(model_params.items())
        num_params = len(param_list)

        if num_params > 0:
            advanced_row_cols = st.columns(num_params + 1, gap="small")
            for col_idx, (param_name, param_config) in enumerate(param_list):
                with advanced_row_cols[col_idx]:
                    session_key = f'{matching_method}_{param_name}'
                    # Ensure session state is initialized
                    if session_key not in st.session_state:
                        st.session_state[session_key] = param_config.get('default', 0.0)

                    st.markdown(
                        f"<div style='font-weight:600;font-size:0.82rem;color:#0f172a;margin-bottom:0.12rem;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;' title='{param_config.get('label', param_name)}'>{param_config.get('label', param_name)}</div>",
                        unsafe_allow_html=True,
                    )

                    st.number_input(
                        param_config.get('label', param_name),
                        min_value=param_config.get('min', 0.0),
                        max_value=param_config.get('max', 100.0),
                        step=param_config.get('step', 0.01),
                        key=session_key,
                        label_visibility="collapsed",
                        help=param_config.get('help', '')
                    )

            with advanced_row_cols[num_params]:
                st.markdown(
                    "<div style='font-weight:600;font-size:0.82rem;color:#0f172a;margin-bottom:0.12rem;visibility:hidden;'>Action</div>",
                    unsafe_allow_html=True,
                )
                save_col, reset_col = st.columns([1, 1], gap="small")
                with save_col:
                    if st.button("Save", key="save_model_params_btn", help="Save model parameters"):
                        current_model = st.session_state.get('matching_method', matching_method)
                        model_key = f"api_key_{current_model.replace(' ', '_')}" if current_model else "api_key_default"
                        payload = {
                            'gpt_system_prompt': _normalized_prompt_value(
                                st.session_state.get('gpt_system_prompt'),
                                default_system,
                            ),
                            'gpt_user_template': _normalized_prompt_value(
                                st.session_state.get('gpt_user_template'),
                                default_user_tpl,
                            ),
                            'gpt_temperature': st.session_state.get('gpt_temperature', 0.0),
                            'gpt_top_p': st.session_state.get('gpt_top_p', 1.0),
                            'gpt_max_tokens': st.session_state.get('gpt_max_tokens', 20),
                            'matching_method': current_model,
                            'api_key': st.session_state.get(model_key, ''),
                        }
                        ok = save_gpt_settings(payload)
                        st.session_state['_advanced_action_status'] = 'save_ok' if ok else 'save_fail'
                        st.session_state['_advanced_action_model'] = current_model
                        st.rerun()
                with reset_col:
                    if st.button("Reset", key="reset_model_params_btn", help="Reset to defaults"):
                        st.session_state['_advanced_reset_model'] = st.session_state.get('matching_method', matching_method)
                        st.session_state['_advanced_reset_pending'] = True
                        st.rerun()
    
    st.markdown('</div>', unsafe_allow_html=True)

    def _save_gpt_callback():
        matching_method = st.session_state.get('matching_method')
        model_key = f'api_key_{matching_method.replace(" ", "_")}' if matching_method else 'api_key_default'
        data = {
            'gpt_system_prompt': _normalized_prompt_value(
                st.session_state.get('gpt_system_prompt'),
                default_system,
            ),
            'gpt_user_template': _normalized_prompt_value(
                st.session_state.get('gpt_user_template'),
                default_user_tpl,
            ),
            'gpt_temperature': st.session_state.get('gpt_temperature', 0.0),
            'gpt_top_p': st.session_state.get('gpt_top_p', 1.0),
            'gpt_max_tokens': st.session_state.get('gpt_max_tokens', 20),
            'matching_method': matching_method,
            'api_key': st.session_state.get(model_key, ''),
        }
        ok = save_gpt_settings(data)
        # Set a short-lived status key to show feedback after rerun
        st.session_state['_gpt_save_status'] = 'ok' if ok else 'fail'
        # Rerun so widget states refresh from session_state/file
        try:
            st.experimental_rerun()
        except Exception:
            pass

    def _reset_gpt_callback():
        ok = reset_gpt_settings()
        # Update session state defaults in a way that's safe for callbacks
        st.session_state.update({
            'gpt_system_prompt': DEFAULT_SYSTEM_PROMPT,
            'gpt_user_template': DEFAULT_USER_PROMPT_TEMPLATE,
            'gpt_temperature': 0.0,
            'gpt_top_p': 1.0,
            'gpt_max_tokens': 20,
        })
        st.session_state['_gpt_save_status'] = 'reset'
        try:
            st.experimental_rerun()
        except Exception:
            pass

    # Make sure these inline buttons are visually prominent (override earlier transparent rules)
    st.markdown(
        """
    <style>
    /* Advanced GPT inline buttons - visible blue style (refined) */
    .stButton > button {
        background: linear-gradient(90deg,#2563eb,#3b82f6) !important;
        color: #ffffff !important;
        border: none !important;
        box-shadow: none !important;
        padding: 0.32rem 0.9rem !important;
        border-radius: 8px !important;
        font-weight: 600 !important;
        height: 34px !important;
        display: inline-flex !important;
        align-items: center !important;
        justify-content: center !important;
        vertical-align: middle !important;
        font-size: 0.95rem !important;
    }
    .stButton > button:active, .stButton > button:focus { outline: none !important; box-shadow:none !important; }

    /* Make number inputs compact and visually subtle so buttons fit on same row */
    /* Target several possible DOM shapes Streamlit could render */
    .stNumberInput input[type=number], .stNumberInput > div > div > input[type=number],
    div[data-baseweb="input"] input[type=number] {
        -webkit-appearance: none !important;
        -moz-appearance: textfield !important;
        appearance: none !important;
        padding: 6px 8px !important;
        height:30px !important;
        max-width:110px !important;
        width:110px !important;
        background: #ffffff !important;
        border: 1px solid rgba(15,23,42,0.06) !important;
        border-radius: 6px !important;
        box-shadow: none !important;
        font-size: 0.95rem !important;
        color: #0f172a !important;
    }
    .stNumberInput > div > div { width:110px !important; max-width:110px !important; }
    .stNumberInput label { display:none !important; }

    /* Allow steppers rendered by Streamlit to be visible (no hiding) */

    /* Ensure buttons sit cleanly aligned with inputs: small left margin and centered vertically */
    .stColumns > div .stButton > button { margin-left: 10px !important; }
    .stColumns > div .stButton { padding: 0 !important; margin: 0 !important; }
        /* Compact button style tweaks specifically for inline Save/Reset */
        #save_gpt_settings_inline, #reset_gpt_settings_inline, button[key="save_gpt_settings_inline"], button[key="reset_gpt_settings_inline"] {
            padding: 0.28rem 0.7rem !important;
            font-size: 0.9rem !important;
            border-radius: 8px !important;
            height: 32px !important;
        }

        /* Remove any red/invalid outline that may appear on focused inputs */
        .stNumberInput input[type=number]:focus, .stTextArea textarea:focus, .stTextInput input:focus {
            outline: none !important;
            box-shadow: 0 0 0 1px rgba(59,130,246,0.12) !important; /* subtle blue focus */
            border-color: rgba(59,130,246,0.20) !important;
        }

        /* Prevent form validation-style red border on some browsers */
        input:invalid, textarea:invalid { box-shadow: none !important; border-color: rgba(15,23,42,0.06) !important; }
        /* Try to center the button cell vertically */
        .stColumns > div > div[role="group"], .stColumns > div > div { display:flex !important; align-items:center !important; }
        /* Prevent buttons from wrapping onto two lines and make them uniform */
        .stButton > button { white-space: nowrap !important; min-width: 88px !important; font-size: 0.9rem !important; }
        button[key="save_gpt_settings_inline"], button[key="reset_gpt_settings_inline"] { min-width: 88px !important; white-space: nowrap !important; }

        /* Force number inputs to keep subtle border color and remove red validation outlines */
        .stNumberInput input[type=number], .stNumberInput > div > div > input[type=number] {
            border-color: rgba(15,23,42,0.06) !important;
            box-shadow: none !important;
        }
        .stNumberInput input[type=number]:invalid, input:invalid, textarea:invalid {
            border-color: rgba(15,23,42,0.06) !important;
            box-shadow: none !important;
        }
        /* Remove the visible rounded box around the entire inputs row */
        /* Target the columns that contain the number inputs and their immediate containers */
        .stColumns > div { background: transparent !important; border: none !important; box-shadow: none !important; padding: 0 0.08rem !important; margin: 0 !important; }
        .stColumns > div .stNumberInput, .stColumns > div .stNumberInput > div, .stColumns > div .stNumberInput > div > div {
            background: transparent !important; border: none !important; box-shadow: none !important; padding: 0 !important; margin: 0 !important;
        }
        /* Also remove any extra rounded container appearing directly around the three-number row */
        .stColumns > div > .css-1l02zno, .stColumns > div > .css-1l02zno.e1tzin5v1 { background: transparent !important; border: none !important; box-shadow: none !important; padding:0 !important; margin:0 !important; }
        /* Additional overrides: remove any red highlight around the group/row */
        .stNumberInput, .stNumberInput > div, .stNumberInput > div > div, .stNumberInput * { box-shadow: none !important; outline: none !important; border-color: rgba(15,23,42,0.06) !important; }
        /* Remove border/outline from the column containers that hold the inputs */
        .stColumns > div, .stColumns > div > div, .stColumns > div > div[role="group"] { border: none !important; box-shadow: none !important; outline: none !important; }
        /* Ensure focused/active containers do not receive red outlines */
        .stColumns > div:focus-within, .stColumns > div > div:focus-within, .stNumberInput:focus-within { box-shadow: none !important; outline: none !important; border: none !important; }
        /* Final safety: override common red border colors used by browsers */
        .stNumberInput, .stNumberInput input, .stNumberInput textarea, input[aria-invalid="true"], input:invalid {
            border-color: rgba(15,23,42,0.06) !important;
            box-shadow: none !important;
            outline: none !important;
        }
        /* Tighten gap between input and inline buttons */
        #gpt-numeric-row .stColumns { gap:4px !important; }
        #gpt-numeric-row .stColumns > div { padding:0 !important; }
        """,
        unsafe_allow_html=True,
    )

    # close wrapper for numeric row
    st.markdown('</div>', unsafe_allow_html=True)

    # Scoped CSS for the numeric row: fix the three input columns to pixel widths
    st.markdown(
        '''
    <style>
    /* Scoped fixed widths for the three numeric columns so outer rounded boxes stay narrow */
    #gpt-numeric-row .stColumns > div:nth-child(1),
    #gpt-numeric-row .stColumns > div:nth-child(2),
    #gpt-numeric-row .stColumns > div:nth-child(3) {
        min-width: 140px !important;
        max-width: 140px !important;
        width: 140px !important;
        flex: 0 0 140px !important;
        padding-left: 2px !important;
        padding-right: 2px !important;
    }
    /* Ensure the internal generated wrapper fills that fixed width and left-aligns content
       but keep it transparent so only the number input is visible (removes empty box area) */
    #gpt-numeric-row .stColumns > div > .css-1l02zno, #gpt-numeric-row .stColumns > div > [data-testid="stNumberInput"] {
        width: 100% !important; display:flex !important; align-items:center !important; justify-content:flex-start !important; gap:6px !important;
        padding: 2px 2px 2px 2px !important; box-sizing: border-box !important; overflow: hidden !important; max-width: 140px !important;
        background: transparent !important; border: none !important; border-radius: 0 !important;
    }
    /* Inner number input area: keep compact */
    #gpt-numeric-row .stNumberInput > div > div { width: 72px !important; max-width:72px !important; padding: 0 !important; margin: 0 !important; }
    #gpt-numeric-row .stNumberInput input[type=number] { width: calc(100% - 44px) !important; box-sizing: border-box !important; padding: 6px 8px !important; height: 28px !important; background: #fff !important; border: 1px solid rgba(200,200,200,0.35) !important; border-radius:6px !important; }
    /* Button column small */
    #gpt-numeric-row .stColumns > div:nth-child(4) { min-width: 80px !important; max-width: 100px !important; flex: 0 0 90px !important; padding-left:6px !important; padding-right:6px !important; }
    /* Normalize heights */
    #gpt-numeric-row .stNumberInput > div, #gpt-numeric-row .stNumberInput > div > div { height:36px !important; display:flex !important; align-items:center !important; }
    /* Make the outer wrapper tight and fixed width matching columns */
    #gpt-numeric-row {
        display: flex !important;
        align-items: center !important;
        justify-content: flex-start !important;
        gap: 8px !important;
        padding: 0 !important;
        margin: 0 !important;
        background: transparent !important;
        border: none !important;
        max-width: 580px !important; /* sum of fixed column widths */
    }
    #gpt-numeric-row .stColumns { width: auto !important; display:flex !important; gap:6px !important; align-items:center !important; }
    /* Force each immediate column container to keep the fixed pixel width */
    #gpt-numeric-row .stColumns > div { box-sizing: border-box !important; flex: 0 0 140px !important; max-width: 140px !important; min-width: 140px !important; padding: 0 !important; }
    /* Remove excess outer wrappers while keeping internal small rounded box */
    #gpt-numeric-row, #gpt-numeric-row *::before, #gpt-numeric-row *::after {
        background: transparent !important; outline: none !important; margin: 0 !important; padding: 0 !important;
    }
    /* Hide any empty generated children that create spacer areas */
    #gpt-numeric-row .stColumns > div > *:empty { display: none !important; }
    /* Ensure inner wrappers don't overflow and occupy exactly the column width */
    #gpt-numeric-row .stColumns > div, #gpt-numeric-row .stColumns > div > * { overflow: hidden !important; box-sizing: border-box !important; }
    /* Reduce internal gaps and padding further to remove visible empty area */
    #gpt-numeric-row .stColumns > div > .css-1l02zno, #gpt-numeric-row .stColumns > div > [data-testid="stNumberInput"] { gap:4px !important; padding: 6px 4px !important; }
    /* Safeguard: hide any child elements that are purely structural spacers */
    #gpt-numeric-row .stColumns > div > *[style*="min-width: 0px"] { display: none !important; }
    /* Hide any sibling elements that come after the main inner wrapper to prevent extra visible gaps */
    #gpt-numeric-row .stColumns > div > .css-1l02zno ~ *,
    #gpt-numeric-row .stColumns > div > [data-testid="stNumberInput"] ~ * { display: none !important; }
    /* Force the visible number input wrapper to fixed width and above others */
    #gpt-numeric-row .stColumns > div > .css-1l02zno, #gpt-numeric-row .stColumns > div > [data-testid="stNumberInput"] { z-index: 10 !important; width: 140px !important; max-width:140px !important; }
    /* Aggressive: hide any generated spacer and keep only essential parts (labels, inputs, buttons) */
    #gpt-numeric-row .stColumns > div > * { display: none !important; }
    #gpt-numeric-row .stColumns > div .stNumberInput, #gpt-numeric-row .stColumns > div .stNumberInput * { display: block !important; }
    #gpt-numeric-row .stColumns > div .stMarkdown, #gpt-numeric-row .stColumns > div label, #gpt-numeric-row .stColumns > div .stButton { display: block !important; }
    /* Keep Save/Reset visible in the final column */
    #gpt-numeric-row .stColumns > div:nth-child(4) .stButton, #gpt-numeric-row .stColumns > div:nth-child(4) .stButton * { display: inline-block !important; }
    </style>
        ''',
        unsafe_allow_html=True,
    )


# Match Quality Threshold (slider only, heading removed per request)
threshold = st.slider("Set High Match Threshold (%)", min_value=50, max_value=100, value=85, help="Adjust the percentage above which matches are considered 'High'.")

# Compact summary of key settings to reduce vertical noise and keep users informed
try:
    _api_flag = bool(globals().get('api_key', '') )
except Exception:
    _api_flag = False
current_method = st.session_state.get('matching_method', 'Azure OpenAI GPT-4o')
azure_display = 'Set' if _api_flag else 'Not Set'
key_label = 'API Key:'
is_dark_theme = (st.session_state.get("user_theme", "light") == "dark")
summary_bg = "#111827" if is_dark_theme else "#fffef0"
summary_border = "#334155" if is_dark_theme else "#e5e7eb"
summary_text = "#e5e7eb" if is_dark_theme else "#0f172a"
summary_muted = "#cbd5e1" if is_dark_theme else "#334155"
summary_html = f"""
<div class='settings-summary-box' style='background:{summary_bg};border:1px solid {summary_border};color:{summary_text};padding:0.7em 1em;border-radius:8px;margin-bottom:1.2em;max-width:100%;'>
    <div style='display:flex;align-items:center;gap:0.75rem;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;font-size:0.95rem;color:{summary_muted};'>
        <div style='flex:0 0 auto;'><b style="color:{summary_text};">Mode:</b> {comparison_mode}</div>
        <div style='flex:0 0 auto;'><b style="color:{summary_text};">Method:</b> {current_method}</div>
        <div style='flex:0 0 auto;'><b style="color:{summary_text};">Threshold:</b> {threshold}%</div>
        <div style='flex:0 0 auto;'><b style="color:{summary_text};">{key_label}</b> {azure_display}</div>
    </div>
</div>
"""
st.markdown(summary_html, unsafe_allow_html=True)

# Theme-aware override for widget help icons + tooltip readability.
help_icon_border = "#64748b" if is_dark_theme else "#94a3b8"
help_icon_text = "#e2e8f0" if is_dark_theme else "#334155"
help_tt_bg = "#0f172a" if is_dark_theme else "#ffffff"
help_tt_text = "#e2e8f0" if is_dark_theme else "#0f172a"
help_tt_border = "#334155" if is_dark_theme else "#cbd5e1"
st.markdown(
    f"""
<style>
[data-testid="stWidgetLabelHelp"] {{
    border: 1px solid {help_icon_border} !important;
    background: transparent !important;
}}
[data-testid="stWidgetLabelHelp"]::before {{
    color: {help_icon_text} !important;
}}
[role="tooltip"], div[data-baseweb="tooltip"], div[data-baseweb="tooltip"] * {{
    background: {help_tt_bg} !important;
    color: {help_tt_text} !important;
    border-color: {help_tt_border} !important;
}}
</style>
""",
    unsafe_allow_html=True,
)

# --- Upload Files (Modified based on mode) ---
# Visible heading should preserve numbering (3) but not include the word 'Step'
st.markdown("""
<div class="step-card small">
    <span class="step-number">3</span>
    <span class="step-title">Upload File(s)</span>
</div>
""", unsafe_allow_html=True)

# Initialize column selectors to avoid undefined-variable errors
col1_q_val = None
col1_a_val = None
col2_q_val = None
col2_a_val = None

# File upload logic - "Compare Any Two Files" supports spreadsheet + document/text formats
if IS_TWO_FILE_MODE:
    col1, col2 = st.columns(2)
    if comparison_mode == "Compare Any Two Files":
        supported_compare_any_types = [ext.lstrip(".") for ext in sorted(COMPARE_ANY_TWO_EXTENSIONS)]
        with col1:
            uploaded_file1 = st.file_uploader(
                "Upload First File (JSON / TXT / PDF / DOCX)",
                type=supported_compare_any_types,
                key="file1",
            )
        with col2:
            uploaded_file2 = st.file_uploader(
                "Upload Second File (JSON / TXT / PDF / DOCX)",
                type=supported_compare_any_types,
                key="file2",
            )
    else:
        # Excel-specific modes only support Excel and CSV
        with col1:
            uploaded_file1 = st.file_uploader("Upload First File (Excel / CSV)", type=['xlsx', 'xls', 'csv'], key="file1")
        with col2:
            uploaded_file2 = st.file_uploader("Upload Second File (Excel / CSV)", type=['xlsx', 'xls', 'csv'], key="file2")
else:
    # Single file modes (Excel-specific) only support Excel and CSV
    uploaded_file1 = st.file_uploader("Upload File (Excel / CSV)", type=['xlsx', 'xls', 'csv'], key="single_file")
    uploaded_file2 = uploaded_file1  # Use same file for both comparisons

# After upload, read files and show sheet/column selectors
df1 = None
df2 = None
skip_file_read = False
if uploaded_file1:
    try:
        # Detect sheets for Excel files
        sheets = []
        try:
            fn1 = getattr(uploaded_file1, 'name', '') or ''
            if fn1.lower().endswith(('.xlsx', '.xls')):
                xls1 = pd.ExcelFile(uploaded_file1)
                sheets = xls1.sheet_names
            else:
                sheets = []
        except Exception:
            sheets = []

        if not IS_TWO_FILE_MODE and sheets:
            # For single-file modes, let user pick one sheet
            if comparison_mode == "Compare Two Columns in Same Excel File":
                selected_sheet = st.selectbox("Select sheet to compare", sheets, index=0, key="single_sheet_select")
            else:
                selected_sheet = st.selectbox("Select sheet to compare (base vs two targets)", sheets, index=0, key="single_sheet_select_2")
            df1 = read_uploaded_file(uploaded_file1, sheet_name=selected_sheet)
            if df1 is None or not hasattr(df1, 'columns'):
                st.error("Uploaded file could not be read. Please upload a valid Excel or CSV file.")
                df1 = pd.DataFrame()
                skip_file_read = True
            else:
                st.session_state['original_df'] = df1.copy()
                st.session_state['selected_sheet_singlefile'] = selected_sheet
        else:
            allow_excel_files = comparison_mode != "Compare Any Two Files"
            df1 = read_uploaded_file(uploaded_file1)
            if df1 is None or not hasattr(df1, 'columns'):
                if comparison_mode == "Compare Any Two Files":
                    st.error("Uploaded file could not be read. Please upload a valid Excel/CSV, JSON, TXT, PDF, or DOCX file.")
                else:
                    st.error("Uploaded file could not be read. Please upload a valid Excel or CSV file.")
                df1 = pd.DataFrame()
                skip_file_read = True
            else:
                st.session_state['original_df'] = df1.copy()

        if IS_TWO_FILE_MODE:
            allow_excel_files = comparison_mode != "Compare Any Two Files"
            if uploaded_file2 is None:
                st.info("Upload the second file to continue with two-file comparison.")
                df2 = pd.DataFrame()
                skip_file_read = True
            else:
            # detect sheets for file2 as well
                sheets2 = []
                try:
                    fn2 = getattr(uploaded_file2, 'name', '') or ''
                    if fn2.lower().endswith(('.xlsx', '.xls')):
                        xls2 = pd.ExcelFile(uploaded_file2)
                        sheets2 = xls2.sheet_names
                except Exception:
                    sheets2 = []

                # If either file has multiple sheets, allow per-file sheet selection laid out side-by-side
                if sheets or sheets2:
                    sheet_col1, sheet_col2 = st.columns(2)
                    with sheet_col1:
                        if sheets:
                            sel1 = st.selectbox("Select sheet for File 1", sheets, index=0, key="file1_sheet_select")
                        else:
                            sel1 = None
                    with sheet_col2:
                        if sheets2:
                            sel2 = st.selectbox("Select sheet for File 2", sheets2, index=0, key="file2_sheet_select")
                        else:
                            sel2 = None
                else:
                    sel1 = sel2 = None

                # Re-read the files using chosen sheets (if any)
                df1 = read_uploaded_file(uploaded_file1, sheet_name=sel1 if sel1 else None)
                df2 = read_uploaded_file(uploaded_file2, sheet_name=sel2 if sel2 else None)

                # Validate read results
                if df1 is None or not hasattr(df1, 'columns'):
                    st.error("Could not read File 1. Please check the file and re-upload.")
                    df1 = pd.DataFrame()
                    skip_file_read = True
                if df2 is None or not hasattr(df2, 'columns'):
                    st.error("Could not read File 2. Please check the file and re-upload.")
                    df2 = pd.DataFrame()
                    skip_file_read = True

                # Only set original_df to df2 if it wasn't already set from df1 and df2 is valid
                if st.session_state.get('original_df') is None and not df2.empty:
                    st.session_state['original_df'] = df2.copy()

                if not skip_file_read:
                    file1_is_spreadsheet = _is_spreadsheet_upload(uploaded_file1)
                    file2_is_spreadsheet = _is_spreadsheet_upload(uploaded_file2)

                    if file1_is_spreadsheet or file2_is_spreadsheet:
                        st.markdown("<b>Select columns to compare:</b>", unsafe_allow_html=True)
                    else:
                        st.info(
                            "Document/text uploads are mapped automatically using the extracted "
                            "`Section` and `Content` fields, then aligned by best text match before scoring."
                        )

                    col_file1, col_file2 = st.columns(2)
                    with col_file1:
                        st.markdown("<b>File 1</b>", unsafe_allow_html=True)
                        if file1_is_spreadsheet:
                            col1_q, col1_a = st.columns(2)
                            with col1_q:
                                col1_q_val = st.selectbox("Question", df1.columns, index=0, key="file1_q_sel", help="Select the question column in File 1")
                            with col1_a:
                                col1_a_val = st.selectbox("Answer (File 1)", df1.columns, index=1, key="file1_a_sel", help="Select the answer column in File 1")
                        else:
                            col1_q_val, col1_a_val = _default_compare_columns(df1)
                            st.caption(f"Using extracted fields automatically: `{col1_q_val}` and `{col1_a_val}`")
                    with col_file2:
                        st.markdown("<b>File 2</b>", unsafe_allow_html=True)
                        if file2_is_spreadsheet:
                            col2_q, col2_a = st.columns(2)
                            with col2_q:
                                col2_q_val = st.selectbox("Question", df2.columns, index=0, key="file2_q_sel", help="Select the question column in File 2")
                            with col2_a:
                                col2_a_val = st.selectbox("Answer (File 2)", df2.columns, index=1, key="file2_a_sel", help="Select the answer column in File 2")
                        else:
                            col2_q_val, col2_a_val = _default_compare_columns(df2)
                            st.caption(f"Using extracted fields automatically: `{col2_q_val}` and `{col2_a_val}`")
        else:
            df2 = df1
            if df1.shape[1] < 3:
                st.error("Error: File must have at least three columns to compare. Please check your file.")
            else:
                # Two different single-file modes: keep existing UI for the old mode; add a new UI for base->two-targets
                if comparison_mode == "Compare Two Columns in Same Excel File":
                    st.markdown("<b>Select two columns to compare within the same file:</b>", unsafe_allow_html=True)
                    col_sel1, col_sel2, col_sel3 = st.columns(3)
                    with col_sel1:
                        st.markdown("<small><b>Column A Question Column</b></small>", unsafe_allow_html=True)
                        col1_q_val = st.selectbox("Question", df1.columns, index=0, key="same_file_q_sel", help="Select the question column")
                    with col_sel2:
                        st.markdown("<small><b>Column B Answer 1</b></small>", unsafe_allow_html=True)
                        col1_a_val = st.selectbox("Base Answer", df1.columns, index=1, key="same_file_a1_sel", help="First answer column to compare (Base Answer)")
                    with col_sel3:
                        st.markdown("<small><b>Column C Answer 2</b></small>", unsafe_allow_html=True)
                        col2_a_val = st.selectbox("Target Column Ans", df1.columns, index=2 if df1.shape[1] > 2 else 1, key="same_file_a2_sel", help="Second answer column to compare (Target Column Ans)")
                    col2_q_val = col1_q_val
                elif comparison_mode == "Compare One Column with Two Targets (Same Excel)":
                    st.markdown("<b>Select base column and two target columns:</b>", unsafe_allow_html=True)
                    col_base, col_target1, col_target2 = st.columns(3)
                    with col_base:
                        st.markdown("<small><b>Base Column</b></small>", unsafe_allow_html=True)
                        col1_q_val = st.selectbox("Base Answer", df1.columns, index=0, key="base_col_sel", help="Select the base answer column")
                    with col_target1:
                        st.markdown("<small><b>Target Column A</b></small>", unsafe_allow_html=True)
                        col1_a_val = st.selectbox("Target Column Ans", df1.columns, index=1, key="target_a_sel", help="Select target answer column A")
                    with col_target2:
                        st.markdown("<small><b>Target Column B</b></small>", unsafe_allow_html=True)
                        col2_a_val = st.selectbox("Target Column Ans2", df1.columns, index=2 if df1.shape[1] > 2 else 1, key="target_b_sel", help="Select target answer column B")
                    # In this mode, treat base column as question-like column for pairing
                    col2_q_val = col1_q_val
    except Exception as e:
        error_msg = str(e)
        if 'utf-8' in error_msg.lower() or 'codec' in error_msg.lower() or 'decode' in error_msg.lower():
            st.error(f"**File Encoding Error**: The uploaded file appears to use a different text encoding. Please try:\n\n"
                    f"1. Save your file as UTF-8 encoded\n"
                    f"2. Use Excel to open and re-save as CSV (UTF-8)\n"
                    f"3. If using Excel, try saving as .xlsx format instead\n\n"
                    f"Technical details: {error_msg}")
        elif "scanned or image-only" in error_msg.lower() or "no selectable text could be extracted" in error_msg.lower():
            st.error(
                "**PDF Text Extraction Error**: This PDF looks like a scanned/image-only file, so the app could not extract text.\n\n"
                "Try one of these:\n"
                "1. Run OCR on the PDF, then upload it again\n"
                "2. Export the PDF text into `.docx` or `.txt`\n"
                "3. Use a text-based PDF instead of a scanned copy\n\n"
                f"Technical details: {error_msg}"
            )
        else:
            st.error(f"Error reading uploaded file(s): {error_msg}")

# --- Model Loading (with error handling) ---
class _NumpyTensorWrapper:
    """Tiny tensor-like wrapper supporting .diagonal().cpu().numpy() chain."""

    def __init__(self, arr):
        self._arr = np.asarray(arr, dtype=float)

    def diagonal(self):
        return _NumpyTensorWrapper(np.diagonal(self._arr))

    def cpu(self):
        return self

    def numpy(self):
        return self._arr


class _FallbackUtil:
    """Minimal cosine-sim helper compatible with existing local-model call sites."""

    @staticmethod
    def cos_sim(a, b):
        a_arr = np.asarray(a, dtype=float)
        b_arr = np.asarray(b, dtype=float)
        if a_arr.ndim == 1:
            a_arr = a_arr.reshape(1, -1)
        if b_arr.ndim == 1:
            b_arr = b_arr.reshape(1, -1)

        a_norm = np.linalg.norm(a_arr, axis=1, keepdims=True)
        b_norm = np.linalg.norm(b_arr, axis=1, keepdims=True)
        a_norm[a_norm == 0] = 1.0
        b_norm[b_norm == 0] = 1.0
        sim_matrix = (a_arr / a_norm) @ (b_arr / b_norm).T
        return _NumpyTensorWrapper(sim_matrix)


class _FallbackSentenceModel:
    """Lightweight local embedding fallback when sentence-transformers is unavailable."""

    def __init__(self, dim=384):
        self.dim = dim

    def _embed_text(self, text):
        vec = np.zeros(self.dim, dtype=float)
        tokens = str(text).lower().split()
        for token in tokens:
            idx = int(hashlib.md5(token.encode("utf-8")).hexdigest(), 16) % self.dim
            vec[idx] += 1.0
        return vec

    def encode(self, texts, convert_to_tensor=True):
        if isinstance(texts, str):
            texts = [texts]
        texts = texts or []
        if not texts:
            return np.zeros((0, self.dim), dtype=float)
        return np.vstack([self._embed_text(t) for t in texts])


@st.cache_resource
def load_main_model(selected_model):
    # Strict local mode: selected model must load successfully.
    if not selected_model:
        raise RuntimeError("No local model is available. Install at least one local sentence-transformer model and retry.")
    try:
        import importlib
        st_mod = importlib.import_module('sentence_transformers')
        SentenceTransformer = getattr(st_mod, 'SentenceTransformer')
    except Exception as e:
        raise RuntimeError(
            "Local Model requires 'sentence-transformers' in the same Python runtime used by Streamlit. "
            f"Runtime: {sys.executable}. Install with: {sys.executable} -m pip install sentence-transformers torch"
        ) from e
    try:
        return SentenceTransformer(selected_model)
    except Exception as e:
        raise RuntimeError(f"Failed to load selected local model: {selected_model}") from e


@st.cache_resource
def load_cross_encoder_model(selected_model):
    try:
        import importlib
        ce_mod = importlib.import_module('sentence_transformers.cross_encoder')
        CrossEncoder = getattr(ce_mod, 'CrossEncoder')
    except Exception:
        st.warning("Cross-encoder support is not available because the required package could not be imported. Falling back to embedding-only similarity.")
        return None
    try:
        return CrossEncoder(selected_model)
    except Exception as e:
        st.warning(f"Could not load cross-encoder model: {e}")
        return None

CROSS_ENCODER_MODELS = [
    "cross-encoder/stsb-roberta-base",
    "cross-encoder/ms-marco-MiniLM-L6-v2"
]

def get_ai_similarity(provider_name, answer1, answer2, api_key, system_prompt=None, user_template=None, temperature=0.0, top_p=1.0, max_tokens=20, model_name=None, question=None):
    """Call any AI provider for similarity matching."""
    from ai_providers import get_provider

    # Pass a template string; providers perform the final format(answer1, answer2).
    prompt_template = _build_similarity_user_template(user_template)
    sys_content = _build_similarity_system_prompt(system_prompt)
    safe_max_tokens = max(16, int(max_tokens or 20))
    prepared_answer1 = _prepare_answer_for_similarity(answer1)
    prepared_answer2 = _prepare_answer_for_similarity(answer2)
    prepared_question = _prepare_question_for_similarity(question)

    try:
        provider = get_provider(provider_name, api_key)
        score, explanation = provider.get_similarity(
            prepared_answer1,
            prepared_answer2,
            sys_content,
            prompt_template,
            temperature,
            top_p,
            safe_max_tokens,
            model_name=model_name,
            question=prepared_question,
        )
        if score is None:
            return None, (explanation or "Provider returned no score.")
        try:
            score_val = float(score)
        except Exception:
            return None, "Provider returned non-numeric score."
        if 0.0 <= score_val <= 1.0 and score_val not in (0.0, 1.0):
            score_val = round(score_val * 100.0, 2)
        score_val = max(0.0, min(100.0, score_val))
        if score_val == 0.0 and isinstance(explanation, str):
            low_exp = explanation.lower()
            if any(tok in low_exp for tok in ["error", "failed", "timeout", "invalid", "unauthorized", "forbidden"]):
                return None, explanation
        return score_val, (explanation or "")
    except Exception as e:
        return None, f"Provider error: {e}"


def _notify_local_backend(main_model):
    """Show one-time notice about which local backend is active."""
    if isinstance(main_model, _FallbackSentenceModel):
        raise RuntimeError("Local fallback backend is disabled in strict local mode.")
    else:
        if not st.session_state.get("_local_backend_notice_transformer", False):
            st.info("Local Model is running with sentence-transformers backend.")
            st.session_state["_local_backend_notice_transformer"] = True


# Legacy function for backward compatibility
def get_gpt4o_similarity(answer1, answer2, api_key, system_prompt=None, user_template=None, temperature=0.0, top_p=1.0, max_tokens=20, question=None):
    return get_ai_similarity("Azure OpenAI GPT-4o", answer1, answer2, api_key, system_prompt, user_template, temperature, top_p, max_tokens, question=question)

# Determine if columns have been selected
cols_selected = False
if IS_TWO_FILE_MODE:
    cols_selected = all([col1_q_val is not None, col1_a_val is not None, col2_q_val is not None, col2_a_val is not None])
else:
    cols_selected = all([col1_q_val is not None, col1_a_val is not None, col2_a_val is not None])

# Ensure api_key variable exists (get from per-model session state)
matching_method = st.session_state.get('matching_method', 'Azure OpenAI GPT-4o')
model_key = f'api_key_{matching_method.replace(" ", "_")}'
api_key = st.session_state.get(model_key, '')
provider_model_name = st.session_state.get(
    f"{matching_method}_resolved_model_name",
    get_default_provider_model(matching_method),
)

# Only show Compare button when files are uploaded and columns selected
ready_to_compare = bool(uploaded_file1 and (uploaded_file2 or comparison_mode == "Compare Two Columns in Same Excel File") and cols_selected)
# Position the Compare and Cancel buttons close to each other
if ready_to_compare:
    st.markdown('<div class="compare-row">', unsafe_allow_html=True)
    col_left, col_compare, col_cancel, col_right = st.columns([3, 1, 1, 3])
    with col_compare:
        # Reset cancel flag when Compare is pressed
        if "cancel_requested" in st.session_state:
            st.session_state["cancel_requested"] = False
        # Disable Compare for cloud providers when API key is missing.
        is_local_mode = (matching_method == "Local Model")
        local_ready = (local_backend_ready and bool(selected_model)) if is_local_mode else True
        disable_compare = ((not is_local_mode) and (not api_key)) or (is_local_mode and (not local_ready))
        compare_clicked = st.button("Compare", help="Click to start the similarity comparison", disabled=disable_compare)
        if disable_compare:
            if is_local_mode and not local_backend_ready:
                st.caption("Install sentence-transformers in the Streamlit runtime to enable Local Model.")
            elif is_local_mode and not selected_model:
                st.caption("Install/cache at least one local model to enable Compare.")
            else:
                st.caption(f"Enter {matching_method} API Key to enable Compare")
    with col_cancel:
        # Cancel button sets a request flag; cancellation is best-effort and works for loops
        if st.button("Cancel", help="Request cancellation of running comparison"):
            st.session_state["cancel_requested"] = True
    st.markdown('</div>', unsafe_allow_html=True)
else:
    compare_clicked = False

if compare_clicked:
    # ensure cancel flag is set to False at start
    st.session_state["cancel_requested"] = False
    
    import re, string
    import difflib
    SentenceTransformer = None
    util = _FallbackUtil()
    CrossEncoder = None
    # Import local-model dependencies only when the Local Model path is actually used.
    if matching_method == "Local Model":
        try:
            import importlib
            st_mod = importlib.import_module('sentence_transformers')
            SentenceTransformer = getattr(st_mod, 'SentenceTransformer', None)
            util = getattr(st_mod, 'util', None)
            if util is None:
                util = _FallbackUtil()
        except Exception:
            SentenceTransformer = None
            util = _FallbackUtil()
        try:
            ce_mod = importlib.import_module('sentence_transformers.cross_encoder')
            CrossEncoder = getattr(ce_mod, 'CrossEncoder', None)
        except Exception:
            CrossEncoder = None
    # --- 3.1: Column Selection and Validation (use earlier selections) ---
    try:
        with st.spinner("Loading models and processing files..."):
                # Ensure df1/df2 are loaded (re-read to get fresh file buffer if needed)
                allow_excel_files = comparison_mode != "Compare Any Two Files"
                if df1 is None:
                    df1 = read_uploaded_file(uploaded_file1)
                if IS_TWO_FILE_MODE:
                    if df2 is None:
                        df2 = read_uploaded_file(uploaded_file2)

                # Capture source filenames and sheet names
                file1_name = getattr(uploaded_file1, 'name', 'file1') if uploaded_file1 is not None else ''
                try:
                    fn1 = getattr(uploaded_file1, 'name', '') or ''
                    if fn1.lower().endswith('.xlsx'):
                        sheet1_name = pd.ExcelFile(uploaded_file1, engine='openpyxl').sheet_names[0]
                    else:
                        sheet1_name = ''
                except Exception:
                    sheet1_name = ''
                # Defaults for second file
                file2_name = getattr(uploaded_file2, 'name', 'file2') if uploaded_file2 is not None else ''
                sheet2_name = ''

                # Validate file format
                if IS_TWO_FILE_MODE:
                    if df1.shape[1] < 2 or df2.shape[1] < 2:
                        st.error("Error: Both files must have at least two columns (question and answer). Please check your files.")
                        st.stop()

                    # Extract selected columns from earlier UI
                    questions1 = df1[col1_q_val].astype(str).fillna("").tolist()
                    answers1 = df1[col1_a_val].astype(str).fillna("").tolist()
                    questions2 = df2[col2_q_val].astype(str).fillna("").tolist()
                    answers2 = df2[col2_a_val].astype(str).fillna("").tolist()
                    # source file 2 info
                    file2_name = getattr(uploaded_file2, 'name', 'file2') if uploaded_file2 is not None else ''
                    try:
                        fn2 = getattr(uploaded_file2, 'name', '') or ''
                        if fn2.lower().endswith('.xlsx'):
                            sheet2_name = pd.ExcelFile(uploaded_file2, engine='openpyxl').sheet_names[0]
                        else:
                            sheet2_name = ''
                    except Exception:
                        sheet2_name = ''
                    # ---- Compare Two Excel Files: run the comparison now ----
                    # Prepare lists and truncate to minimum length
                    questions1 = questions1 if 'questions1' in locals() else df1[col1_q_val].astype(str).fillna("").tolist()
                    answers1 = answers1 if 'answers1' in locals() else df1[col1_a_val].astype(str).fillna("").tolist()
                    questions2 = questions2 if 'questions2' in locals() else df2[col2_q_val].astype(str).fillna("").tolist()
                    answers2 = answers2 if 'answers2' in locals() else df2[col2_a_val].astype(str).fillna("").tolist()

                    def clean_answer(ans):
                        ans = ans.lower()
                        ans = re.sub(r'\[.*?\]', '', ans)
                        ans = re.sub(r'\bbased on the provided context,?\s*', '', ans)
                        ans = ans.translate(str.maketrans('', '', string.punctuation))
                        ans = ans.strip()
                        return ans

                    def aggressive_clean(ans):
                        ans = clean_answer(ans)
                        ans = _prepare_answer_for_similarity(ans)
                        # Preserve numeric tokens so "25" vs "70" are not treated as identical.
                        context_phrases = [
                            r'based on the provided context', r'from the context', r'from context', r'context', r'see context',
                            r'as per context', r'per context', r'per the context', r'per the provided context', r'provided context', r'according to'
                        ]
                        for phrase in context_phrases:
                            ans = re.sub(rf'\b{phrase}\b', '', ans, flags=re.IGNORECASE)
                        ans = re.sub(r'\s+', ' ', ans)
                        return ans.strip()

                    use_best_match_alignment = _should_use_best_match_alignment(
                        comparison_mode,
                        uploaded_file1,
                        uploaded_file2,
                    )
                    alignment_rows = None
                    alignment_seed_scores = []

                    if use_best_match_alignment:
                        alignment_rows = _build_best_match_alignment(
                            questions1,
                            answers1,
                            questions2,
                            answers2,
                        )
                        questions1 = [row["question1"] for row in alignment_rows]
                        answers1 = [row["answer1"] for row in alignment_rows]
                        questions2 = [row["question2"] for row in alignment_rows]
                        answers2 = [row["answer2"] for row in alignment_rows]
                        alignment_seed_scores = [row["alignment_seed_score"] for row in alignment_rows]
                        min_len = len(alignment_rows)
                    else:
                        # Normalize lengths
                        min_len = min(len(questions1), len(questions2), len(answers1), len(answers2))
                        questions1 = questions1[:min_len]
                        questions2 = questions2[:min_len]
                        answers1 = answers1[:min_len]
                        answers2 = answers2[:min_len]

                    cleaned1 = [aggressive_clean(a) for a in answers1]
                    cleaned2 = [aggressive_clean(a) for a in answers2]
                    cleaned1 = list(map(str, cleaned1))
                    cleaned2 = list(map(str, cleaned2))
                    prepared_answers1 = [_prepare_answer_for_similarity(a) for a in answers1]
                    prepared_answers2 = [_prepare_answer_for_similarity(a) for a in answers2]
                    use_question_context = _should_use_question_context(
                        comparison_mode,
                        uploaded_file1,
                        uploaded_file2,
                    )
                    question_contexts = [
                        _prepare_question_for_similarity(q) if use_question_context else ""
                        for q in questions1
                    ]
                    semantic_texts1 = [
                        f"Question: {q}\nAnswer: {a}" if q else a
                        for q, a in zip(question_contexts, cleaned1)
                    ]
                    semantic_texts2 = [
                        f"Question: {q}\nAnswer: {a}" if q else a
                        for q, a in zip(question_contexts, cleaned2)
                    ]

                    explanations = [""] * min_len

                    # Choose method: Cloud provider or Local Model chunked
                    if matching_method != "Local Model" and api_key:
                        gpt_scores = []
                        gpt_explanations = []
                        progress = st.progress(0, text=f"Comparing answers with {matching_method}...")
                        for idx, (a1, a2, q_ctx) in enumerate(zip(answers1, answers2, question_contexts)):
                            if st.session_state.get("cancel_requested", False):
                                st.info("Comparison cancelled by user.")
                                break
                            if not a1.strip() or not a2.strip():
                                score, explanation = 0, "One or both answers are empty."
                            else:
                                # Get model-specific parameters
                                model_params = get_model_parameters(matching_method)
                                temp_key = f'{matching_method}_temperature'
                                top_p_key = f'{matching_method}_top_p'
                                max_tokens_key = f'{matching_method}_max_tokens'
                                
                                temp = st.session_state.get(temp_key, model_params.get('temperature', {}).get('default', 0.0))
                                top_p = st.session_state.get(top_p_key, model_params.get('top_p', {}).get('default', 1.0))
                                max_tokens = st.session_state.get(max_tokens_key, model_params.get('max_tokens', {}).get('default', 20))
                                
                                score, explanation = get_ai_similarity(
                                    matching_method,
                                    a1,
                                    a2,
                                    api_key,
                                    system_prompt=st.session_state.get('gpt_system_prompt', None),
                                    user_template=st.session_state.get('gpt_user_template', None),
                                    temperature=temp,
                                    top_p=top_p,
                                    max_tokens=max_tokens,
                                    model_name=provider_model_name,
                                    question=q_ctx,
                                )
                                if score is None:
                                    raise RuntimeError(
                                        f"{matching_method} ({provider_model_name}) failed at row {idx + 1}: {explanation}"
                                    )
                            gpt_scores.append(score if score is not None else 0)
                            gpt_explanations.append(explanation)
                            progress.progress((idx+1)/min_len, text=f"Compared {idx+1}/{min_len} pairs")
                        progress.empty()
                        final_percent_sim = gpt_scores
                        explanations = gpt_explanations
                        raw_sim = [None] * min_len
                        fuzzy_scores = [None] * min_len
                    else:
                        if matching_method == "Local Model":
                            if st.session_state.get("cancel_requested", False):
                                st.info("Comparison cancelled before model loading.")
                                raise Exception("Cancelled")
                            main_model = load_main_model(selected_model)
                            _notify_local_backend(main_model)
                            cross_encoder = None
                            n = min_len
                            chunk_size = 64
                            progress = st.progress(0, text="Encoding and computing local similarities...")
                            if isinstance(main_model, _FallbackSentenceModel):
                                percent_sim_mpnet = []
                                raw_sims = []
                                for idx, (lhs, rhs) in enumerate(zip(cleaned1, cleaned2)):
                                    if st.session_state.get("cancel_requested", False):
                                        st.info("Comparison cancelled by user.")
                                        break
                                    s = _lexical_similarity_percent(lhs, rhs)
                                    percent_sim_mpnet.append(s)
                                    raw_sims.append(round(s / 100.0, 4))
                                    progress.progress((idx + 1) / max(1, n), text=f"Compared {idx+1}/{n} pairs")
                                raw_sim_mpnet = np.array(raw_sims)
                                cross_scores = percent_sim_mpnet
                                progress.empty()
                            else:
                                cross_encoder = load_cross_encoder_model(selected_model) if selected_model in CROSS_ENCODER_MODELS else None
                                sims = []
                                raw_sims = []
                                processed = 0
                                for i in range(0, n, chunk_size):
                                    if st.session_state.get("cancel_requested", False):
                                        st.info("Comparison cancelled by user.")
                                        break
                                    end = min(i + chunk_size, n)
                                    emb1 = main_model.encode(semantic_texts1[i:end], convert_to_tensor=True)
                                    emb2 = main_model.encode(semantic_texts2[i:end], convert_to_tensor=True)
                                    sim_chunk = util.cos_sim(emb1, emb2).diagonal().cpu().numpy()
                                    sims.extend(sim_chunk.tolist())
                                    raw_sims.extend(sim_chunk.round(4).tolist())
                                    processed = end
                                    progress.progress(processed / n, text=f"Encoded and compared {processed}/{n} pairs")
                                progress.progress(1.0, text=f"Encoded and compared {n}/{n} pairs")
                                progress.empty()
                                similarities = np.array(sims)
                                percent_sim_mpnet = (similarities * 100).round(2)
                                raw_sim_mpnet = np.array(raw_sims)
                                cross_scores = None
                                if cross_encoder is not None:
                                    try:
                                        pairs = list(zip(semantic_texts1, semantic_texts2))
                                        cross_sim_list = []
                                        for i in range(0, n, chunk_size):
                                            if st.session_state.get("cancel_requested", False):
                                                st.info("Comparison cancelled by user during cross-encoder step.")
                                                break
                                            end = min(i + chunk_size, n)
                                            pred = cross_encoder.predict(pairs[i:end], show_progress_bar=False)
                                            cross_sim_list.extend(pred.tolist())
                                            progress.progress(80 + int(end / n * 20), text=f"Cross-encoder processed {end}/{n} pairs")
                                        cross_sim = np.array(cross_sim_list)
                                        if cross_sim.size and np.max(cross_sim) - np.min(cross_sim) > 0:
                                            cross_sim = (cross_sim - np.min(cross_sim)) / (np.max(cross_sim) - np.min(cross_sim))
                                        cross_scores = (cross_sim * 100).round(2) if cross_sim.size else percent_sim_mpnet
                                    except Exception as e:
                                        st.warning(f"Cross-encoder failed: {e}")
                                        cross_scores = percent_sim_mpnet
                                else:
                                    cross_scores = percent_sim_mpnet
                                    try:
                                        progress.progress(100, text="Local model comparison complete")
                                    except Exception:
                                        pass

                                from difflib import SequenceMatcher
                            def fuzzy_ratio(a, b):
                                return int(SequenceMatcher(None, a, b).ratio() * 100)
                            fuzzy_scores = [fuzzy_ratio(a1, a2) for a1, a2 in zip(cleaned1, cleaned2)]
                            lexical_scores = [_lexical_similarity_percent(a1, a2) for a1, a2 in zip(prepared_answers1, prepared_answers2)]
                            if cross_encoder is not None and cross_scores is not None:
                                final_percent_sim = cross_scores
                            else:
                                final_percent_sim = [
                                    max(float(mpnet), float(fuzz), float(lex))
                                    for mpnet, fuzz, lex in zip(percent_sim_mpnet, fuzzy_scores, lexical_scores)
                                ]
                            raw_sim = raw_sim_mpnet
                        else:
                            final_percent_sim = [None] * min_len
                            raw_sim = [None] * min_len
                            fuzzy_scores = [None] * min_len

                    final_percent_sim = _calibrate_similarity_series(final_percent_sim, prepared_answers1, prepared_answers2)

                    match_quality = [
                        "High" if s and s > threshold else ("Medium" if s and s > 60 else "Low")
                        for s in final_percent_sim
                    ]

                    # Highlight differences for file1 vs file2 answers
                    def highlight_diff(a, b):
                        seqm = difflib.SequenceMatcher(None, a, b)
                        out1, out2 = '', ''
                        for opcode, a0, a1, b0, b1 in seqm.get_opcodes():
                            if opcode == 'equal':
                                out1 += a[a0:a1]
                                out2 += b[b0:b1]
                            elif opcode == 'replace':
                                out1 += f'<span style="background-color:#ffd6d6">{a[a0:a1]}</span>'
                                out2 += f'<span style="background-color:#ffd6d6">{b[b0:b1]}</span>'
                            elif opcode == 'insert':
                                out2 += f'<span style="background-color:#d6ffd6">{b[b0:b1]}</span>'
                            elif opcode == 'delete':
                                out1 += f'<span style="background-color:#ffd6d6">{a[a0:a1]}</span>'
                        return out1, out2

                    diff1, diff2 = zip(*(highlight_diff(a1, a2) for a1, a2 in zip(answers1, answers2)))

                    q_col_name = col1_q_val if col1_q_val is not None else "Question"
                    a1_col_name = col1_a_val if col1_a_val is not None else "Answer 1"
                    a2_col_name = col2_a_val if col2_a_val is not None else "Answer 2"
                    # Ensure column names are unique in the results DataFrame. If both files
                    # use the same column name, append a file-specific suffix so columns
                    # are not lost when creating a dict for the DataFrame.
                    if a1_col_name == a2_col_name:
                        a1_col_name = f"{a1_col_name} (File 1)"
                        a2_col_name = f"{a2_col_name} (File 2)"
                    # If the question column collides with answer names, disambiguate as well
                    if q_col_name == a1_col_name or q_col_name == a2_col_name:
                        q_col_name = f"{q_col_name} (Question)"

                    q2_col_name = col2_q_val if col2_q_val is not None else "Question 2"
                    if q2_col_name == q_col_name:
                        q2_col_name = f"{q2_col_name} (File 2)"
                    if q2_col_name == a1_col_name or q2_col_name == a2_col_name:
                        q2_col_name = f"{q2_col_name} (Match)"

                    sim_col_name = f"{a1_col_name} & {a2_col_name} Similarity"

                    results_payload = {
                        q_col_name: questions1,
                        a1_col_name: answers1,
                    }
                    if use_best_match_alignment:
                        results_payload[q2_col_name] = questions2
                    results_payload.update(
                        {
                            a2_col_name: answers2,
                            "Source File 1": file1_name,
                            "Source Sheet 1": sheet1_name,
                            "Source File 2": file2_name,
                            "Source Sheet 2": sheet2_name,
                            sim_col_name: final_percent_sim,
                        }
                    )
                    if use_best_match_alignment:
                        results_payload["Alignment Strategy"] = ["Best text match"] * len(final_percent_sim)
                        results_payload["Initial Alignment Score"] = alignment_seed_scores

                    results_df = pd.DataFrame(results_payload)

                    st.session_state['results_df'] = results_df.copy()
                    st.session_state['similarity_cols'] = [sim_col_name]
                    st.session_state['primary_sim_col'] = sim_col_name

                    diff_payload = {
                        q_col_name: questions1,
                        f"{a1_col_name} (diff)": diff1,
                    }
                    if use_best_match_alignment:
                        diff_payload[q2_col_name] = questions2
                    diff_payload.update(
                        {
                            f"{a2_col_name} (diff)": diff2,
                            "Source File 1": file1_name,
                            "Source Sheet 1": sheet1_name,
                            "Source File 2": file2_name,
                            "Source Sheet 2": sheet2_name,
                            sim_col_name: final_percent_sim,
                        }
                    )
                    if use_best_match_alignment:
                        diff_payload["Alignment Strategy"] = ["Best text match"] * len(final_percent_sim)
                        diff_payload["Initial Alignment Score"] = alignment_seed_scores

                    st.session_state['diff_table'] = pd.DataFrame(diff_payload).copy()

                    if use_best_match_alignment:
                        st.success(f"Compared {min_len} aligned text chunks using best-match pairing.")
                    else:
                        st.success(f"Compared {min_len} question-answer pairs.")
                else:
                    df2 = df1
                    # New mode: compare one base column with two target columns (no changes to other modes)
                    if comparison_mode == "Compare One Column with Two Targets (Same Excel)":
                        # Ensure at least three columns
                        if df1.shape[1] < 3:
                            st.error("Error: File must have at least three columns to compare (base + 2 targets). Please check your file.")
                            st.stop()

                        # Map selectors: base = col1_q_val, target A = col1_a_val, target B = col2_a_val
                        base_vals = df1[col1_q_val].astype(str).fillna("").tolist()
                        target_a_vals = df1[col1_a_val].astype(str).fillna("").tolist()
                        target_b_vals = df1[col2_a_val].astype(str).fillna("").tolist()

                        file1_name = getattr(uploaded_file1, 'name', 'file') if uploaded_file1 is not None else ''
                        try:
                            sheet1_name = pd.ExcelFile(uploaded_file1, engine='openpyxl').sheet_names[0]
                        except Exception:
                            sheet1_name = ''

                        min_len = min(len(base_vals), len(target_a_vals), len(target_b_vals))
                        base_vals = base_vals[:min_len]
                        target_a_vals = target_a_vals[:min_len]
                        target_b_vals = target_b_vals[:min_len]

                        def clean_answer(ans):
                            ans = ans.lower()
                            ans = re.sub(r'\[.*?\]', '', ans)
                            ans = re.sub(r'\bbased on the provided context,?\s*', '', ans)
                            ans = ans.translate(str.maketrans('', '', string.punctuation))
                            ans = ans.strip()
                            return ans

                        def aggressive_clean(ans):
                            ans = clean_answer(ans)
                            ans = _prepare_answer_for_similarity(ans)
                            # Preserve numeric tokens so quantity/value columns compare correctly.
                            context_phrases = [
                                r'based on the provided context', r'from the context', r'from context', r'context', r'see context',
                                r'as per context', r'per context', r'per the context', r'per the provided context', r'provided context', r'according to'
                            ]
                            for phrase in context_phrases:
                                ans = re.sub(rf'\b{phrase}\b', '', ans, flags=re.IGNORECASE)
                            ans = re.sub(r'\s+', ' ', ans)
                            return ans.strip()

                        cleaned_base = [aggressive_clean(a) for a in base_vals]
                        cleaned_a = [aggressive_clean(a) for a in target_a_vals]
                        cleaned_b = [aggressive_clean(a) for a in target_b_vals]
                        cleaned_base = list(map(str, cleaned_base))
                        cleaned_a = list(map(str, cleaned_a))
                        cleaned_b = list(map(str, cleaned_b))
                        prepared_base = [_prepare_answer_for_similarity(a) for a in base_vals]
                        prepared_target_a = [_prepare_answer_for_similarity(a) for a in target_a_vals]
                        prepared_target_b = [_prepare_answer_for_similarity(a) for a in target_b_vals]

                        explanations = [""] * min_len

                        # Use cloud provider or local model as appropriate, computing two similarity series
                        if matching_method != "Local Model" and api_key:
                            gpt_scores_a = []
                            gpt_scores_b = []
                            gpt_explanations = []
                            progress = st.progress(0, text=f"Comparing base->targetA and base->targetB with {matching_method}...")
                            for idx, (b, a, c) in enumerate(zip(base_vals, target_a_vals, target_b_vals)):
                                if st.session_state.get("cancel_requested", False):
                                    st.info("Comparison cancelled by user.")
                                    break
                                if not b.strip() or not a.strip():
                                    s_a, e_a = 0, "Empty base or target A"
                                else:
                                    # Get model-specific parameters
                                    model_params = get_model_parameters(matching_method)
                                    temp_key = f'{matching_method}_temperature'
                                    top_p_key = f'{matching_method}_top_p'
                                    max_tokens_key = f'{matching_method}_max_tokens'
                                    
                                    temp = st.session_state.get(temp_key, model_params.get('temperature', {}).get('default', 0.0))
                                    top_p = st.session_state.get(top_p_key, model_params.get('top_p', {}).get('default', 1.0))
                                    max_tokens = st.session_state.get(max_tokens_key, model_params.get('max_tokens', {}).get('default', 20))
                                    
                                    s_a, e_a = get_ai_similarity(
                                        matching_method,
                                        b,
                                        a,
                                        api_key,
                                        system_prompt=st.session_state.get('gpt_system_prompt', None),
                                        user_template=st.session_state.get('gpt_user_template', None),
                                        temperature=temp,
                                        top_p=top_p,
                                        max_tokens=max_tokens,
                                        model_name=provider_model_name,
                                    )
                                    if s_a is None:
                                        raise RuntimeError(
                                            f"{matching_method} ({provider_model_name}) failed at row {idx + 1} for Target A: {e_a}"
                                        )
                                if not b.strip() or not c.strip():
                                    s_b, e_b = 0, "Empty base or target B"
                                else:
                                    # Get model-specific parameters
                                    model_params = get_model_parameters(matching_method)
                                    temp_key = f'{matching_method}_temperature'
                                    top_p_key = f'{matching_method}_top_p'
                                    max_tokens_key = f'{matching_method}_max_tokens'
                                    
                                    temp = st.session_state.get(temp_key, model_params.get('temperature', {}).get('default', 0.0))
                                    top_p = st.session_state.get(top_p_key, model_params.get('top_p', {}).get('default', 1.0))
                                    max_tokens = st.session_state.get(max_tokens_key, model_params.get('max_tokens', {}).get('default', 20))
                                    
                                    s_b, e_b = get_ai_similarity(
                                        matching_method,
                                        b,
                                        c,
                                        api_key,
                                        system_prompt=st.session_state.get('gpt_system_prompt', None),
                                        user_template=st.session_state.get('gpt_user_template', None),
                                        temperature=temp,
                                        top_p=top_p,
                                        max_tokens=max_tokens,
                                        model_name=provider_model_name,
                                    )
                                    if s_b is None:
                                        raise RuntimeError(
                                            f"{matching_method} ({provider_model_name}) failed at row {idx + 1} for Target B: {e_b}"
                                        )
                                gpt_scores_a.append(s_a if s_a is not None else 0)
                                gpt_scores_b.append(s_b if s_b is not None else 0)
                                gpt_explanations.append(f"A:{e_a} | B:{e_b}")
                                progress.progress((idx+1)/min_len, text=f"Compared {idx+1}/{min_len} pairs")
                            progress.empty()
                            final_percent_sim_a = gpt_scores_a
                            final_percent_sim_b = gpt_scores_b
                            explanations = gpt_explanations
                            raw_sim_a = [None] * min_len
                            raw_sim_b = [None] * min_len
                        else:
                            if matching_method == "Local Model":
                                if st.session_state.get("cancel_requested", False):
                                    st.info("Comparison cancelled before model loading.")
                                    raise Exception("Cancelled")
                                main_model = load_main_model(selected_model)
                                _notify_local_backend(main_model)
                                cross_encoder = None
                                n = min_len
                                chunk_size = 64
                                progress = st.progress(0, text="Encoding and computing local similarities...")

                                if isinstance(main_model, _FallbackSentenceModel):
                                    percent_a = []
                                    percent_b = []
                                    raw_a_vals = []
                                    raw_b_vals = []
                                    for idx, (base_t, a_t, b_t) in enumerate(zip(cleaned_base, cleaned_a, cleaned_b)):
                                        if st.session_state.get("cancel_requested", False):
                                            st.info("Comparison cancelled by user.")
                                            break
                                        sa = _lexical_similarity_percent(base_t, a_t)
                                        sb = _lexical_similarity_percent(base_t, b_t)
                                        percent_a.append(sa)
                                        percent_b.append(sb)
                                        raw_a_vals.append(round(sa / 100.0, 4))
                                        raw_b_vals.append(round(sb / 100.0, 4))
                                        progress.progress((idx + 1) / max(1, n), text=f"Compared {idx+1}/{n} pairs")
                                    raw_a = np.array(raw_a_vals)
                                    raw_b = np.array(raw_b_vals)
                                    cross_scores_a = percent_a
                                    cross_scores_b = percent_b
                                    progress.empty()
                                else:
                                    cross_encoder = load_cross_encoder_model(selected_model) if selected_model in CROSS_ENCODER_MODELS else None
                                    sim_a_list = []
                                    sim_b_list = []
                                    raw_a_list = []
                                    raw_b_list = []
                                    processed = 0
                                    for i in range(0, n, chunk_size):
                                        if st.session_state.get("cancel_requested", False):
                                            st.info("Comparison cancelled by user.")
                                            break
                                        end = min(i + chunk_size, n)
                                        emb_b_chunk = main_model.encode(cleaned_base[i:end], convert_to_tensor=True)
                                        emb_a_chunk = main_model.encode(cleaned_a[i:end], convert_to_tensor=True)
                                        emb_c_chunk = main_model.encode(cleaned_b[i:end], convert_to_tensor=True)
                                        sim_a_chunk = util.cos_sim(emb_b_chunk, emb_a_chunk).diagonal().cpu().numpy()
                                        sim_b_chunk = util.cos_sim(emb_b_chunk, emb_c_chunk).diagonal().cpu().numpy()
                                        sim_a_list.extend(sim_a_chunk.tolist())
                                        sim_b_list.extend(sim_b_chunk.tolist())
                                        raw_a_list.extend(sim_a_chunk.round(4).tolist())
                                        raw_b_list.extend(sim_b_chunk.round(4).tolist())
                                        processed = end
                                        progress.progress(int(processed / n * 60), text=f"Encoded and compared {processed}/{n} pairs")

                                    sim_a = np.array(sim_a_list)
                                    sim_b = np.array(sim_b_list)
                                    percent_a = (sim_a * 100).round(2)
                                    percent_b = (sim_b * 100).round(2)
                                    raw_a = np.array(raw_a_list)
                                    raw_b = np.array(raw_b_list)
                                    cross_scores_a = None
                                    cross_scores_b = None
                                    if cross_encoder is not None:
                                        try:
                                            pairs_a = list(zip(cleaned_base, cleaned_a))
                                            pairs_b = list(zip(cleaned_base, cleaned_b))
                                            cross_scores_a_list = []
                                            cross_scores_b_list = []
                                            for i in range(0, n, chunk_size):
                                                if st.session_state.get("cancel_requested", False):
                                                    st.info("Comparison cancelled by user during cross-encoder step.")
                                                    break
                                                end = min(i + chunk_size, n)
                                                chunk_pairs_a = pairs_a[i:end]
                                                chunk_pairs_b = pairs_b[i:end]
                                                pred_a = cross_encoder.predict(chunk_pairs_a, show_progress_bar=False)
                                                pred_b = cross_encoder.predict(chunk_pairs_b, show_progress_bar=False)
                                                cross_scores_a_list.extend(pred_a.tolist())
                                                cross_scores_b_list.extend(pred_b.tolist())
                                                progress.progress(60 + int(end / n * 30), text=f"Cross-encoder processed {end}/{n} pairs")

                                            def normalize(arr):
                                                arr = np.array(arr)
                                                if np.max(arr) - np.min(arr) > 0:
                                                    arr = (arr - np.min(arr)) / (np.max(arr) - np.min(arr))
                                                return (arr * 100).round(2)

                                            cross_scores_a = normalize(cross_scores_a_list) if cross_scores_a_list else percent_a
                                            cross_scores_b = normalize(cross_scores_b_list) if cross_scores_b_list else percent_b
                                            progress.progress(100, text="Local model comparison complete")
                                        except Exception as e:
                                            st.warning(f"Cross-encoder failed: {e}")
                                            cross_scores_a = percent_a
                                            cross_scores_b = percent_b
                                    else:
                                        cross_scores_a = percent_a
                                        cross_scores_b = percent_b
                                        try:
                                            progress.progress(100, text="Local model comparison complete")
                                        except Exception:
                                            pass
                                from difflib import SequenceMatcher
                                def fuzzy_ratio(a, b):
                                    return int(SequenceMatcher(None, a, b).ratio() * 100)
                                fuzzy_a = [fuzzy_ratio(a,b) for a,b in zip(cleaned_base, cleaned_a)]
                                fuzzy_b = [fuzzy_ratio(a,b) for a,b in zip(cleaned_base, cleaned_b)]
                                lexical_a = [_lexical_similarity_percent(a, b) for a, b in zip(prepared_base, prepared_target_a)]
                                lexical_b = [_lexical_similarity_percent(a, b) for a, b in zip(prepared_base, prepared_target_b)]
                                final_percent_sim_a = [
                                    max(float(mpnet), float(fuzz), float(lex))
                                    for mpnet, fuzz, lex in zip(cross_scores_a, fuzzy_a, lexical_a)
                                ]
                                final_percent_sim_b = [
                                    max(float(mpnet), float(fuzz), float(lex))
                                    for mpnet, fuzz, lex in zip(cross_scores_b, fuzzy_b, lexical_b)
                                ]
                                raw_sim_a = raw_a
                                raw_sim_b = raw_b
                            else:
                                final_percent_sim_a = [None] * min_len
                                final_percent_sim_b = [None] * min_len
                                raw_sim_a = [None] * min_len
                                raw_sim_b = [None] * min_len

                        final_percent_sim_a = _calibrate_similarity_series(final_percent_sim_a, prepared_base, prepared_target_a)
                        final_percent_sim_b = _calibrate_similarity_series(final_percent_sim_b, prepared_base, prepared_target_b)

                        match_quality_a = ["High" if s and s > threshold else ("Medium" if s and s > 60 else "Low") for s in final_percent_sim_a]
                        match_quality_b = ["High" if s and s > threshold else ("Medium" if s and s > 60 else "Low") for s in final_percent_sim_b]

                        def highlight_diff(a, b):
                            seqm = difflib.SequenceMatcher(None, a, b)
                            out1, out2 = '', ''
                            for opcode, a0, a1, b0, b1 in seqm.get_opcodes():
                                if opcode == 'equal':
                                    out1 += a[a0:a1]
                                    out2 += b[b0:b1]
                                elif opcode == 'replace':
                                    out1 += f'<span style="background-color:#ffd6d6">{a[a0:a1]}</span>'
                                    out2 += f'<span style="background-color:#ffd6d6">{b[b0:b1]}</span>'
                                elif opcode == 'insert':
                                    out2 += f'<span style="background-color:#d6ffd6">{b[b0:b1]}</span>'
                                elif opcode == 'delete':
                                    out1 += f'<span style="background-color:#ffd6d6">{a[a0:a1]}</span>'
                            return out1, out2

                        diff_a1, diff_a2 = zip(*(highlight_diff(b, a) for b, a in zip(base_vals, target_a_vals)))
                        diff_b1, diff_b2 = zip(*(highlight_diff(b, c) for b, c in zip(base_vals, target_b_vals)))

                        # Build similarity column names from actual column names
                        sim1 = f"{col1_q_val} & {col1_a_val} Similarity" if col1_q_val else f"{col1_a_val} & {col1_a_val} Similarity"
                        sim2 = f"{col1_q_val} & {col2_a_val} Similarity" if col1_q_val else f"{col2_a_val} & {col2_a_val} Similarity"

                        results_df = pd.DataFrame({
                            col1_q_val: base_vals,
                            col1_a_val: target_a_vals,
                            col2_a_val: target_b_vals,
                            "Source File": file1_name,
                            "Source Sheet": sheet1_name,
                            sim1: final_percent_sim_a,
                            "Raw Similarity 1": raw_sim_a,
                            "Match Quality 1": match_quality_a,
                            sim2: final_percent_sim_b,
                            "Raw Similarity 2": raw_sim_b,
                            "Match Quality 2": match_quality_b,
                        })
                        st.session_state['diff_table'] = diff_table.copy()

                        # Persist similarity column names so downstream UI shows real names
                        st.session_state['similarity_cols'] = [sim1, sim2]
                        # For downstream metrics choose the first similarity as primary
                        st.session_state['primary_sim_col'] = sim1

                        st.success(f"Compared {min_len} base->target pairs.")
                        # Ensure consistent column name variables for downstream code
                        q_col_name = col1_q_val if col1_q_val is not None else "Question"
                        a1_col_name = col1_a_val if col1_a_val is not None else "Answer 1"
                        a2_col_name = col2_a_val if col2_a_val is not None else "Answer 2"
                        # Disambiguate identical column names so both target columns are preserved
                        if a1_col_name == a2_col_name:
                            a1_col_name = f"{a1_col_name} (File 1)"
                            a2_col_name = f"{a2_col_name} (File 2)"
                        if q_col_name == a1_col_name or q_col_name == a2_col_name:
                            q_col_name = f"{q_col_name} (Question)"
                        similarity_cols = []
                        if 'Similarity Match 1 (%)' in results_df.columns:
                            similarity_cols.append('Similarity Match 1 (%)')
                        if 'Similarity Match 2 (%)' in results_df.columns:
                            similarity_cols.append('Similarity Match 2 (%)')
                        st.session_state['similarity_cols'] = similarity_cols
                        # Normalize names used by downstream code
                        questions1 = base_vals
                        answers1 = target_a_vals
                        answers2 = target_b_vals
                        # diff_a2 and diff_b2 hold the target-side highlighted diffs
                        try:
                            diff1 = diff_a2
                        except NameError:
                            diff1 = [''] * len(questions1)
                        try:
                            diff2 = diff_b2
                        except NameError:
                            diff2 = [''] * len(questions1)
                    else:
                        # Existing same-file behavior: compare two columns (unchanged)
                        if df1.shape[1] < 3:
                            st.error("Error: File must have at least three columns to compare two answer columns. Please check your file.")
                            st.stop()

                        questions1 = df1[col1_q_val].astype(str).fillna("").tolist()
                        answers1 = df1[col1_a_val].astype(str).fillna("").tolist()
                        questions2 = df1[col1_q_val].astype(str).fillna("").tolist()
                        answers2 = df1[col2_a_val].astype(str).fillna("").tolist()
                        file1_name = getattr(uploaded_file1, 'name', 'file') if uploaded_file1 is not None else ''
                        try:
                            sheet1_name = pd.ExcelFile(uploaded_file1, engine='openpyxl').sheet_names[0]
                        except Exception:
                            sheet1_name = ''
                        min_len = min(len(questions1), len(questions2), len(answers1), len(answers2))
                        questions1 = questions1[:min_len]
                        questions2 = questions2[:min_len]
                        answers1 = answers1[:min_len]
                        answers2 = answers2[:min_len]

                        def clean_answer(ans):
                            ans = ans.lower()
                            ans = re.sub(r'\[.*?\]', '', ans)
                            ans = re.sub(r'\bbased on the provided context,?\s*', '', ans)
                            ans = ans.translate(str.maketrans('', '', string.punctuation))
                            ans = ans.strip()
                            return ans

                        def aggressive_clean(ans):
                            ans = clean_answer(ans)
                            ans = _prepare_answer_for_similarity(ans)
                            # Preserve numeric tokens so quantity/value columns compare correctly.
                            context_phrases = [
                                r'based on the provided context', r'from the context', r'from context', r'context', r'see context', r'as per context', r'per context', r'per the context', r'per the provided context', r'provided context', r'according to'
                            ]
                            for phrase in context_phrases:
                                ans = re.sub(rf'\b{phrase}\b', '', ans, flags=re.IGNORECASE)
                            ans = re.sub(r'\s+', ' ', ans)
                            return ans.strip()

                        cleaned1 = [aggressive_clean(a) for a in answers1]
                        cleaned2 = [aggressive_clean(a) for a in answers2]
                        cleaned1 = list(map(str, cleaned1))
                        cleaned2 = list(map(str, cleaned2))
                        prepared_answers1 = [_prepare_answer_for_similarity(a) for a in answers1]
                        prepared_answers2 = [_prepare_answer_for_similarity(a) for a in answers2]
                        question_contexts = [_prepare_question_for_similarity(q) for q in questions1]
                        semantic_texts1 = [
                            f"Question: {q}\nAnswer: {a}" if q else a
                            for q, a in zip(question_contexts, cleaned1)
                        ]
                        semantic_texts2 = [
                            f"Question: {q}\nAnswer: {a}" if q else a
                            for q, a in zip(question_contexts, cleaned2)
                        ]

                        explanations = [""] * min_len

                        if matching_method != "Local Model" and api_key:
                            gpt_scores = []
                            gpt_explanations = []
                            progress = st.progress(0, text=f"Comparing answers with {matching_method}...")
                            for idx, (a1, a2, q_ctx) in enumerate(zip(answers1, answers2, question_contexts)):
                                if st.session_state.get("cancel_requested", False):
                                    st.info("Comparison cancelled by user.")
                                    break
                                if not a1.strip() or not a2.strip():
                                    score, explanation = 0, "One or both answers are empty."
                                else:
                                    # Get model-specific parameters
                                    model_params = get_model_parameters(matching_method)
                                    temp_key = f'{matching_method}_temperature'
                                    top_p_key = f'{matching_method}_top_p'
                                    max_tokens_key = f'{matching_method}_max_tokens'

                                    temp = st.session_state.get(temp_key, model_params.get('temperature', {}).get('default', 0.0))
                                    top_p = st.session_state.get(top_p_key, model_params.get('top_p', {}).get('default', 1.0))
                                    max_tokens = st.session_state.get(max_tokens_key, model_params.get('max_tokens', {}).get('default', 20))

                                    score, explanation = get_ai_similarity(
                                        matching_method,
                                        a1,
                                        a2,
                                        api_key,
                                        system_prompt=st.session_state.get('gpt_system_prompt', None),
                                        user_template=st.session_state.get('gpt_user_template', None),
                                        temperature=temp,
                                        top_p=top_p,
                                        max_tokens=max_tokens,
                                        model_name=provider_model_name,
                                        question=q_ctx,
                                    )
                                    if score is None:
                                        raise RuntimeError(
                                            f"{matching_method} ({provider_model_name}) failed at row {idx + 1}: {explanation}"
                                        )
                                gpt_scores.append(score if score is not None else 0)
                                gpt_explanations.append(explanation)
                                progress.progress((idx+1)/min_len, text=f"Compared {idx+1}/{min_len} pairs")
                            progress.empty()
                            final_percent_sim = gpt_scores
                            explanations = gpt_explanations
                            raw_sim = [None] * min_len
                            fuzzy_scores = [None] * min_len
                        else:
                            if matching_method == "Local Model":
                                if st.session_state.get("cancel_requested", False):
                                    st.info("Comparison cancelled before model loading.")
                                    raise Exception("Cancelled")
                                main_model = load_main_model(selected_model)
                                _notify_local_backend(main_model)
                                cross_encoder = None
                                n = min_len
                                chunk_size = 64
                                progress = st.progress(0, text="Encoding and computing local similarities...")

                                if isinstance(main_model, _FallbackSentenceModel):
                                    percent_sim_mpnet = []
                                    raw_sims = []
                                    for idx, (lhs, rhs) in enumerate(zip(cleaned1, cleaned2)):
                                        if st.session_state.get("cancel_requested", False):
                                            st.info("Comparison cancelled by user.")
                                            break
                                        s = _lexical_similarity_percent(lhs, rhs)
                                        percent_sim_mpnet.append(s)
                                        raw_sims.append(round(s / 100.0, 4))
                                        progress.progress((idx + 1) / max(1, n), text=f"Compared {idx+1}/{n} pairs")
                                    raw_sim_mpnet = np.array(raw_sims)
                                    cross_scores = percent_sim_mpnet
                                    progress.empty()
                                else:
                                    cross_encoder = load_cross_encoder_model(selected_model) if selected_model in CROSS_ENCODER_MODELS else None
                                    sims = []
                                    raw_sims = []
                                    processed = 0
                                    for i in range(0, n, chunk_size):
                                        if st.session_state.get("cancel_requested", False):
                                            st.info("Comparison cancelled by user.")
                                            break
                                        end = min(i + chunk_size, n)
                                        emb1 = main_model.encode(semantic_texts1[i:end], convert_to_tensor=True)
                                        emb2 = main_model.encode(semantic_texts2[i:end], convert_to_tensor=True)
                                        sim_chunk = util.cos_sim(emb1, emb2).diagonal().cpu().numpy()
                                        sims.extend(sim_chunk.tolist())
                                        raw_sims.extend(sim_chunk.round(4).tolist())
                                        processed = end
                                        progress.progress(processed / n, text=f"Encoded and compared {processed}/{n} pairs")

                                    progress.progress(1.0, text=f"Encoded and compared {n}/{n} pairs")
                                    progress.empty()

                                    similarities = np.array(sims)
                                    percent_sim_mpnet = (similarities * 100).round(2)
                                    raw_sim_mpnet = np.array(raw_sims)
                                    cross_scores = None
                                    if cross_encoder is not None:
                                        try:
                                            pairs = list(zip(semantic_texts1, semantic_texts2))
                                            cross_sim_list = []
                                            for i in range(0, n, chunk_size):
                                                if st.session_state.get("cancel_requested", False):
                                                    st.info("Comparison cancelled by user during cross-encoder step.")
                                                    break
                                                end = min(i + chunk_size, n)
                                                pred = cross_encoder.predict(pairs[i:end], show_progress_bar=False)
                                                cross_sim_list.extend(pred.tolist())
                                                progress.progress(80 + int(end / n * 20), text=f"Cross-encoder processed {end}/{n} pairs")
                                            cross_sim = np.array(cross_sim_list)
                                            if cross_sim.size and np.max(cross_sim) - np.min(cross_sim) > 0:
                                                cross_sim = (cross_sim - np.min(cross_sim)) / (np.max(cross_sim) - np.min(cross_sim))
                                            cross_scores = (cross_sim * 100).round(2) if cross_sim.size else percent_sim_mpnet
                                        except Exception as e:
                                            st.warning(f"Cross-encoder failed: {e}")
                                            cross_scores = percent_sim_mpnet
                                    else:
                                        cross_scores = percent_sim_mpnet
                                        try:
                                            progress.progress(100, text="Local model comparison complete")
                                        except Exception:
                                            pass
                                from difflib import SequenceMatcher
                                def fuzzy_ratio(a, b):
                                    return int(SequenceMatcher(None, a, b).ratio() * 100)
                                fuzzy_scores = [fuzzy_ratio(a, b) for a, b in zip(cleaned1, cleaned2)]
                                lexical_scores = [_lexical_similarity_percent(a1, a2) for a1, a2 in zip(prepared_answers1, prepared_answers2)]
                                if not isinstance(main_model, _FallbackSentenceModel) and cross_encoder is not None and cross_scores is not None:
                                    final_percent_sim = cross_scores
                                else:
                                    final_percent_sim = [
                                        max(float(mpnet), float(fuzz), float(lex))
                                        for mpnet, fuzz, lex in zip(percent_sim_mpnet, fuzzy_scores, lexical_scores)
                                    ]
                                raw_sim = raw_sim_mpnet
                            else:
                                final_percent_sim = [None] * min_len
                                raw_sim = [None] * min_len
                                fuzzy_scores = [None] * min_len

                        final_percent_sim = _calibrate_similarity_series(final_percent_sim, prepared_answers1, prepared_answers2)

                        match_quality = [
                            "High" if s and s > threshold else ("Medium" if s and s > 60 else "Low")
                            for s in final_percent_sim
                        ]
                        def highlight_diff(a, b):
                            seqm = difflib.SequenceMatcher(None, a, b)
                            out1, out2 = '', ''
                            for opcode, a0, a1, b0, b1 in seqm.get_opcodes():
                                if opcode == 'equal':
                                    out1 += a[a0:a1]
                                    out2 += b[b0:b1]
                                elif opcode == 'replace':
                                    out1 += f'<span style="background-color:#ffd6d6">{a[a0:a1]}</span>'
                                    out2 += f'<span style="background-color:#ffd6d6">{b[b0:b1]}</span>'
                                elif opcode == 'insert':
                                    out2 += f'<span style="background-color:#d6ffd6">{b[b0:b1]}</span>'
                                elif opcode == 'delete':
                                    out1 += f'<span style="background-color:#ffd6d6">{a[a0:a1]}</span>'
                            return out1, out2

                        diff1, diff2 = zip(*(highlight_diff(a1, a2) for a1, a2 in zip(answers1, answers2)))

                        q_col_name = col1_q_val if col1_q_val is not None else "Question"
                        a1_col_name = col1_a_val if col1_a_val is not None else "Answer 1"
                        a2_col_name = col2_a_val if col2_a_val is not None else "Answer 2"
                        # Make sure column names are unique to avoid data being overwritten
                        if a1_col_name == a2_col_name:
                            a1_col_name = f"{a1_col_name} (File 1)"
                            a2_col_name = f"{a2_col_name} (File 2)"
                        if q_col_name == a1_col_name or q_col_name == a2_col_name:
                            q_col_name = f"{q_col_name} (Question)"

                        # Dynamic similarity column name (e.g., 'TruDiscovery & Open AI Similarity')
                        sim_col_name = f"{a1_col_name} & {a2_col_name} Similarity"
                        primary_sim_col = sim_col_name

                        results_df = pd.DataFrame({
                            q_col_name: questions1,
                            a1_col_name: answers1,
                            a2_col_name: answers2,
                            "Source File": file1_name,
                            "Source Sheet": sheet1_name,
                            sim_col_name: final_percent_sim
                        })

                        # Persist results and similarity column names
                        st.session_state['results_df'] = results_df.copy()
                        st.session_state['similarity_cols'] = [sim_col_name]
                        st.session_state['primary_sim_col'] = primary_sim_col

                        st.session_state['diff_table'] = pd.DataFrame({
                            q_col_name: questions1,
                            f"{a1_col_name} (diff)": diff1,
                            f"{a2_col_name} (diff)": diff2,
                            "Source File": file1_name,
                            "Source Sheet": sheet1_name,
                            sim_col_name: final_percent_sim
                        }).copy()

                        st.success(f"Compared {min_len} question-answer pairs.")
                    
                    # Normalize column and similarity names so downstream code is branch-agnostic
                    # Determine question/answer column names (exclude source/similarity/diff columns)
                    meta_cols = {'Source File','Source Sheet','Source File 1','Source Sheet 1','Source File 2','Source Sheet 2'}
                    non_sim_cols = [c for c in results_df.columns if c not in meta_cols and 'Similarity' not in c and '(diff)' not in c]
                    # Heuristic: first non-sim column is question, next two are answers/targets
                    q_col_name = non_sim_cols[0] if len(non_sim_cols) > 0 else 'Question'
                    a1_col_name = non_sim_cols[1] if len(non_sim_cols) > 1 else 'Answer 1'
                    a2_col_name = non_sim_cols[2] if len(non_sim_cols) > 2 else 'Answer 2'

                    # Ensure lists for display/diffs exist
                    try:
                        questions1 = results_df[q_col_name].astype(str).fillna("").tolist() if q_col_name in results_df.columns else []
                    except Exception:
                        questions1 = []
                    try:
                        answers1 = results_df[a1_col_name].astype(str).fillna("").tolist() if a1_col_name in results_df.columns else []
                    except Exception:
                        answers1 = []
                    try:
                        answers2 = results_df[a2_col_name].astype(str).fillna("").tolist() if a2_col_name in results_df.columns else []
                    except Exception:
                        answers2 = []

                    # Diffs if present
                    diff1_col = f"{a1_col_name} (diff)"
                    diff2_col = f"{a2_col_name} (diff)"
                    diff1 = results_df[diff1_col].tolist() if diff1_col in results_df.columns else [''] * len(questions1)
                    diff2 = results_df[diff2_col].tolist() if diff2_col in results_df.columns else [''] * len(questions1)

                    # Determine similarity columns
                    sim_cols_local = [c for c in results_df.columns if 'Similarity' in c]
                    # Prefer a stored primary_sim_col only if it exists in the current results; otherwise pick the first detected similarity column
                    stored_primary = st.session_state.get('primary_sim_col')
                    if stored_primary and stored_primary in results_df.columns:
                        primary_sim_col = stored_primary
                    else:
                        primary_sim_col = sim_cols_local[0] if sim_cols_local else None
                        st.session_state['primary_sim_col'] = primary_sim_col
                    st.session_state['similarity_cols'] = sim_cols_local

                    # final_percent_sim used in some display blocks; provide safe value
                    if primary_sim_col is not None and primary_sim_col in results_df.columns:
                        # coerce non-numeric to NaN then to None for downstream
                        try:
                            final_percent_sim = pd.to_numeric(results_df[primary_sim_col], errors='coerce').tolist()
                        except Exception:
                            final_percent_sim = results_df[primary_sim_col].tolist()
                    else:
                        final_percent_sim = [None] * len(questions1)

                    # Build filtered datasets safely only when the similarity column exists and is numeric
                    if primary_sim_col is not None and primary_sim_col in results_df.columns:
                        try:
                            numeric_sim = pd.to_numeric(results_df[primary_sim_col], errors='coerce')
                            results_below_80 = results_df[numeric_sim < 80].copy()
                            results_below_50 = results_df[numeric_sim < 50].copy()
                        except KeyError:
                            # Column disappeared between runs; clear stored primary and fallback to empty sets
                            st.session_state['primary_sim_col'] = None
                            results_below_80 = pd.DataFrame()
                            results_below_50 = pd.DataFrame()
                        except Exception:
                            results_below_80 = pd.DataFrame()
                            results_below_50 = pd.DataFrame()
                    else:
                        results_below_80 = pd.DataFrame()
                        results_below_50 = pd.DataFrame()

                    # Persist results so they survive reruns (e.g., when downloading)
                    st.session_state['results_df'] = results_df.copy()
                    st.session_state['diff_table'] = locals().get('diff_table', pd.DataFrame()).copy()
                    
                    # Display summary statistics (Total, Above threshold, Between 40-threshold, Below 40)
                    st.markdown("### Comparison Summary")
                    col_stat1, col_stat2, col_stat3, col_stat4 = st.columns(4)
                    with col_stat1:
                        # Count non-null similarity entries as total pairs
                        try:
                            if primary_sim_col is not None and primary_sim_col in results_df.columns:
                                numeric_sim_all = pd.to_numeric(results_df[primary_sim_col], errors='coerce')
                                total_pairs = int(numeric_sim_all.notna().sum())
                            else:
                                total_pairs = len(results_df)
                        except Exception:
                            total_pairs = len(results_df)
                        st.metric("Total Pairs", total_pairs)
                    with col_stat2:
                        # Above threshold count
                        try:
                            if primary_sim_col is not None and primary_sim_col in results_df.columns:
                                numeric_sim = pd.to_numeric(results_df[primary_sim_col], errors='coerce')
                                count_above = int((numeric_sim >= threshold).sum())
                            else:
                                count_above = 0
                        except Exception:
                            count_above = 0
                        st.metric(f"Above {threshold}%", count_above)
                    with col_stat3:
                        # Between 40% (inclusive) and threshold (exclusive)
                        try:
                            if primary_sim_col is not None and primary_sim_col in results_df.columns:
                                numeric_sim = pd.to_numeric(results_df[primary_sim_col], errors='coerce')
                                count_between = int(((numeric_sim >= 40) & (numeric_sim < threshold)).sum())
                            else:
                                count_between = 0
                        except Exception:
                            count_between = 0
                        st.metric(f"Between 40-{threshold}%", count_between)
                    with col_stat4:
                        # Below 40%
                        try:
                            if primary_sim_col is not None and primary_sim_col in results_df.columns:
                                numeric_sim = pd.to_numeric(results_df[primary_sim_col], errors='coerce')
                                count_below_40 = int((numeric_sim < 40).sum())
                            else:
                                count_below_40 = 0
                        except Exception:
                            count_below_40 = 0
                        st.metric("Below 40%", count_below_40)
                    
                    # Highlighted differences are hidden by default; user can expand to view them
                    # Build diff table using selected column names with a (diff) suffix and include source info
                    if IS_TWO_FILE_MODE:
                        diff_payload = {
                            q_col_name: questions1,
                            f"{a1_col_name} (diff)": diff1,
                            f"{a2_col_name} (diff)": diff2,
                            "Source File 1": file1_name,
                            "Source Sheet 1": sheet1_name,
                            "Source File 2": file2_name,
                            "Source Sheet 2": sheet2_name
                        }
                        if primary_sim_col is not None and primary_sim_col in results_df.columns:
                            diff_payload[primary_sim_col] = final_percent_sim
                        diff_table = pd.DataFrame(diff_payload)
                    else:
                        diff_payload = {
                            q_col_name: questions1,
                            f"{a1_col_name} (diff)": diff1,
                            f"{a2_col_name} (diff)": diff2,
                            "Source File": file1_name,
                            "Source Sheet": sheet1_name
                        }
                        if primary_sim_col is not None and primary_sim_col in results_df.columns:
                            diff_payload[primary_sim_col] = final_percent_sim
                        diff_table = pd.DataFrame(diff_payload)
                    # Comparison complete; results saved to session and shown below.
                    st.success("Comparison complete — results saved. Use the Download Options and expanders below to inspect or download results.")
    except Exception as e:
        logger.exception("Error processing files in similarity comparison flow.")
        st.error(f"Error processing files: {e}")
        # --- Error Logging for Debugging ---
        with open(ERROR_LOG_PATH, "a", encoding="utf-8") as logf:
            logf.write(traceback.format_exc() + "\n")
    # If results exist in session_state (e.g., after a rerun from download), show them
    if st.session_state.get('results_df') is not None:
        results_df = st.session_state['results_df']
        diff_table = st.session_state.get('diff_table')
        # Recompute filtered sets using detected similarity columns, with guards
        similarity_cols = [c for c in results_df.columns if 'Similarity' in c]
        primary_sim_col = similarity_cols[0] if similarity_cols else None
        if primary_sim_col is not None and primary_sim_col in results_df.columns:
            try:
                numeric_sim = pd.to_numeric(results_df[primary_sim_col], errors='coerce')
                results_below_80 = results_df[numeric_sim < 80].copy()
                results_below_50 = results_df[numeric_sim < 50].copy()
            except KeyError:
                st.session_state['primary_sim_col'] = None
                results_below_80 = pd.DataFrame()
                results_below_50 = pd.DataFrame()
            except Exception:
                results_below_80 = pd.DataFrame()
                results_below_50 = pd.DataFrame()
        else:
            results_below_80 = pd.DataFrame()
            results_below_50 = pd.DataFrame()

        st.markdown("---")

        with st.expander("Show full results table", expanded=False):
            st.markdown("**Full Results Table:**")
            st.dataframe(results_df, width="stretch")

        # Download options (drop source metadata)
        st.markdown("### Download Options")
        # Let user choose sheet naming for the exported Excel (only affects the written sheet name)
        sheet_name_choice = None
        if IS_TWO_FILE_MODE:
            sheet_name_choice = st.selectbox(
                "Export sheet name:",
                ["Sheet", "Merge file+sheet names", "Use first file sheet name (original)"],
                index=0,
                help="Choose how the sheet will be named inside the exported Excel file. 'Sheet' is a fixed name."
            )
        col_dl1, col_dl2, col_dl3 = st.columns(3)
        with col_dl1:
            # Use original uploaded columns for export (preserve original shape) and append similarity column(s)
            if IS_TWO_FILE_MODE:
                # For two-file comparisons, export the paired results (both answer columns + question + similarity)
                # This avoids only exporting file1's columns when the user expects both answers side-by-side.
                exclude_meta = ['Source File','Source Sheet','Source File 1','Source Sheet 1','Source File 2','Source Sheet 2']
                export_cols = [col for col in results_df.columns if col not in exclude_meta and '(diff)' not in col]
                export_df = results_df[export_cols].copy()
            else:
                if st.session_state.get('original_df') is not None:
                    export_df = st.session_state['original_df'].copy()
                else:
                    # Fallback: use columns from results_df minus source metadata and similarity columns
                    exclude_cols = ['Source File','Source Sheet','Source File 1','Source Sheet 1','Source File 2','Source Sheet 2']
                    orig_cols = [col for col in results_df.columns if col not in exclude_cols and ('Similarity' not in col and '(diff)' not in col)]
                    export_df = results_df[orig_cols].copy()
            # Format and add all similarity columns as numeric percentages (avoid Excel 'number stored as text' warning)
            similarity_cols_export = [c for c in results_df.columns if 'Similarity' in c]
            for col in similarity_cols_export:
                # Convert to numeric and store as a fraction (e.g., 0.83 for 83%) so Excel recognizes it as a number
                numeric = pd.to_numeric(results_df[col], errors='coerce')
                # Use integer percent precision to mirror previous behaviour, then divide by 100 for Excel percent format
                export_df[col] = numeric.apply(lambda v: (int(round(v)) / 100.0) if pd.notnull(v) else None)
            output_all = io.BytesIO()
            # Determine sheet name based on user choice
            sheet_name = "Results"
            try:
                if sheet_name_choice == "Sheet":
                    sheet_name = "Sheet"
                elif sheet_name_choice == "Merge file+sheet names":
                    # Try to read both sheet names and merge them; fall back to combined file names or 'Results'
                    s1 = s2 = None
                    try:
                        if uploaded_file1 is not None:
                            fn1 = getattr(uploaded_file1, 'name', '') or ''
                            if fn1.lower().endswith('.xlsx'):
                                s1 = pd.ExcelFile(uploaded_file1, engine='openpyxl').sheet_names[0]
                            else:
                                s1 = None
                        else:
                            s1 = None
                    except Exception:
                        s1 = None
                    try:
                        if uploaded_file2 is not None:
                            fn2 = getattr(uploaded_file2, 'name', '') or ''
                            if fn2.lower().endswith('.xlsx'):
                                s2 = pd.ExcelFile(uploaded_file2, engine='openpyxl').sheet_names[0]
                            else:
                                s2 = None
                        else:
                            s2 = None
                    except Exception:
                        s2 = None
                    if s1 and s2:
                        # sanitize and shorten names to avoid Excel sheet name limits
                        def clean_sheet(n):
                            return str(n)[:25].replace('/', '_').replace('\\', '_')
                        sheet_name = f"{clean_sheet(s1)}_{clean_sheet(s2)}"
                    else:
                        # fallback to merging file basenames
                        f1 = uploaded_file1.name.rsplit('.',1)[0] if uploaded_file1 is not None else ''
                        f2 = uploaded_file2.name.rsplit('.',1)[0] if uploaded_file2 is not None else ''
                        if f1 or f2:
                            sheet_name = (f1 + '_' + f2)[:31]
                        else:
                            sheet_name = "Results"
                else:
                    # Default/original behaviour: use first available uploaded file's first sheet
                    if uploaded_file1 is not None:
                        fn1 = getattr(uploaded_file1, 'name', '') or ''
                        if fn1.lower().endswith('.xlsx'):
                            sheet_name = pd.ExcelFile(uploaded_file1, engine='openpyxl').sheet_names[0]
                    elif uploaded_file2 is not None:
                        fn2 = getattr(uploaded_file2, 'name', '') or ''
                        if fn2.lower().endswith('.xlsx'):
                            sheet_name = pd.ExcelFile(uploaded_file2, engine='openpyxl').sheet_names[0]
            except Exception:
                sheet_name = "Results"
            with pd.ExcelWriter(output_all, engine='openpyxl') as writer:
                # If single-file modes, write all sheets from uploaded_file1 and inject export_df into selected sheet
                try:
                    sheet_results = st.session_state.get('sheet_results', {})
                    if not IS_TWO_FILE_MODE and uploaded_file1 is not None:
                        # selected sheet name stored earlier when reading
                        sel = st.session_state.get('selected_sheet_singlefile')
                        # Only treat uploaded_file1 as an Excel workbook if it appears to be .xlsx
                        fn1 = getattr(uploaded_file1, 'name', '') or ''
                        if fn1.lower().endswith('.xlsx'):
                            xls = pd.ExcelFile(uploaded_file1, engine='openpyxl')
                            for s in xls.sheet_names:
                                df_orig = read_uploaded_file(uploaded_file1, sheet_name=s)
                                if sel and s == sel:
                                    # Merge similarity columns into the original sheet (preserve original columns)
                                    try:
                                        res_df = sheet_results.get('file1', {}).get(s) if sheet_results else None
                                        if res_df is None:
                                            res_df = st.session_state.get('results_df') if st.session_state.get('results_df') is not None else export_df
                                        df_write = df_orig.copy()
                                        sim_cols = [c for c in res_df.columns if 'Similarity' in c]
                                        for c in sim_cols:
                                            numeric = pd.to_numeric(res_df[c], errors='coerce')
                                            df_write[c] = numeric.apply(lambda v: (int(round(v)) / 100.0) if pd.notnull(v) else None)
                                        df_write.to_excel(writer, index=False, sheet_name=s)
                                    except Exception:
                                        df_orig.to_excel(writer, index=False, sheet_name=s)
                                    # Immediately create & write a per-sheet Summary next to this sheet
                                    try:
                                        res_df = sheet_results.get('file1', {}).get(s, st.session_state.get('results_df'))
                                        if res_df is not None:
                                            # compute summary
                                            sim_col = st.session_state.get('primary_sim_col')
                                            if not sim_col or sim_col not in res_df.columns:
                                                sim_cols = [c for c in res_df.columns if 'Similarity' in c]
                                                sim_col = sim_cols[0] if sim_cols else None
                                            if sim_col and sim_col in res_df.columns:
                                                numeric_sim = pd.to_numeric(res_df[sim_col], errors='coerce')
                                                total_pairs = int(numeric_sim.notna().sum())
                                                above_thresh = int((numeric_sim >= threshold).sum())
                                                between_40_thresh = int(((numeric_sim >= 40) & (numeric_sim < threshold)).sum())
                                                below_40 = int((numeric_sim < 40).sum())
                                                avg_similarity = round(float(numeric_sim.mean()), 2) if total_pairs > 0 else 0
                                            else:
                                                total_pairs = len(res_df)
                                                above_thresh = between_40_thresh = below_40 = 0
                                                avg_similarity = 0
                                            summary_df = pd.DataFrame({
                                                "Metric": ["Total Pairs", f"Above {threshold}%", f"Between 40-{threshold}%", "Below 40%", "Average Similarity (%)", "High Threshold (%)"],
                                                "Value": [total_pairs, above_thresh, between_40_thresh, below_40, avg_similarity, threshold]
                                            })
                                            # sheet name for summary
                                            summary_name = (f"{s} Summary"[:28] + '...') if len(f"{s} Summary") > 31 else f"{s} Summary"
                                            summary_df.to_excel(writer, index=False, sheet_name=summary_name)
                                    except Exception:
                                        pass
                                else:
                                    df_orig.to_excel(writer, index=False, sheet_name=s)
                        else:
                            # uploaded_file1 is not an Excel workbook (likely CSV) - write the exported results sheet only
                            export_df.to_excel(writer, index=False, sheet_name=sel if sel else sheet_name)
                            if sel and s == sel:
                                # Merge similarity columns into the original sheet (preserve original columns)
                                try:
                                    res_df = sheet_results.get('file1', {}).get(s) if sheet_results else None
                                    if res_df is None:
                                        res_df = st.session_state.get('results_df') if st.session_state.get('results_df') is not None else export_df
                                    df_write = df_orig.copy()
                                    sim_cols = [c for c in res_df.columns if 'Similarity' in c]
                                    for c in sim_cols:
                                        numeric = pd.to_numeric(res_df[c], errors='coerce')
                                        df_write[c] = numeric.apply(lambda v: (int(round(v)) / 100.0) if pd.notnull(v) else None)
                                    df_write.to_excel(writer, index=False, sheet_name=s)
                                except Exception:
                                    df_orig.to_excel(writer, index=False, sheet_name=s)
                                # Immediately create & write a per-sheet Summary next to this sheet
                                try:
                                    res_df = sheet_results.get('file1', {}).get(s, st.session_state.get('results_df'))
                                    if res_df is not None:
                                        # compute summary
                                        sim_col = st.session_state.get('primary_sim_col')
                                        if not sim_col or sim_col not in res_df.columns:
                                            sim_cols = [c for c in res_df.columns if 'Similarity' in c]
                                            sim_col = sim_cols[0] if sim_cols else None
                                        if sim_col and sim_col in res_df.columns:
                                            numeric_sim = pd.to_numeric(res_df[sim_col], errors='coerce')
                                            total_pairs = int(numeric_sim.notna().sum())
                                            above_thresh = int((numeric_sim >= threshold).sum())
                                            between_40_thresh = int(((numeric_sim >= 40) & (numeric_sim < threshold)).sum())
                                            below_40 = int((numeric_sim < 40).sum())
                                            avg_similarity = round(float(numeric_sim.mean()), 2) if total_pairs > 0 else 0
                                        else:
                                            total_pairs = len(res_df)
                                            above_thresh = between_40_thresh = below_40 = 0
                                            avg_similarity = 0
                                        summary_df = pd.DataFrame({
                                            "Metric": ["Total Pairs", f"Above {threshold}%", f"Between 40-{threshold}%", "Below 40%", "Average Similarity (%)", "High Threshold (%)"],
                                            "Value": [total_pairs, above_thresh, between_40_thresh, below_40, avg_similarity, threshold]
                                        })
                                        # sheet name for summary
                                        summary_name = (f"{s} Summary"[:28] + '...') if len(f"{s} Summary") > 31 else f"{s} Summary"
                                        summary_df.to_excel(writer, index=False, sheet_name=summary_name)
                                except Exception:
                                    pass
                            else:
                                df_orig.to_excel(writer, index=False, sheet_name=s)
                        # end for sheets
                    elif IS_TWO_FILE_MODE and uploaded_file1 is not None and uploaded_file2 is not None:
                        # For two-file comparisons produce a single slim results sheet and one Summary sheet
                        res_df = st.session_state.get('results_df', results_df)
                        if res_df is None or res_df.empty:
                            # No comparison results available; write original workbooks as-is
                            x1 = pd.ExcelFile(uploaded_file1, engine='openpyxl')
                            # If uploaded files are Excel workbooks, mirror their sheets. If CSVs, skip mirroring and write results only.
                            fn1 = getattr(uploaded_file1, 'name', '') or ''
                            fn2 = getattr(uploaded_file2, 'name', '') or ''
                            if fn1.lower().endswith('.xlsx'):
                                x1 = pd.ExcelFile(uploaded_file1, engine='openpyxl')
                                for s in x1.sheet_names:
                                    df_orig = read_uploaded_file(uploaded_file1, sheet_name=s)
                                    df_orig.to_excel(writer, index=False, sheet_name=s)
                            if fn2.lower().endswith('.xlsx'):
                                x2 = pd.ExcelFile(uploaded_file2, engine='openpyxl')
                                for s in x2.sheet_names:
                                    df_orig = read_uploaded_file(uploaded_file2, sheet_name=s)
                                    out_name = s
                                    if out_name in writer.sheets:
                                        out_name = f"{s}_file2"
                                    df_orig.to_excel(writer, index=False, sheet_name=out_name)
                        else:
                            try:
                                meta_cols = {'Source File','Source Sheet','Source File 1','Source Sheet 1','Source File 2','Source Sheet 2'}
                                non_sim = [c for c in res_df.columns if c not in meta_cols and 'Similarity' not in c and '(diff)' not in c]
                                q_col = non_sim[0] if len(non_sim) > 0 else (res_df.columns[0] if len(res_df.columns) > 0 else 'Question')
                                a1_col = non_sim[1] if len(non_sim) > 1 else (res_df.columns[1] if len(res_df.columns) > 1 else 'Answer 1')
                                a2_col = non_sim[2] if len(non_sim) > 2 else (res_df.columns[2] if len(res_df.columns) > 2 else 'Answer 2')
                                sim_cols = [c for c in res_df.columns if 'Similarity' in c]
                                sim_col = st.session_state.get('primary_sim_col') if st.session_state.get('primary_sim_col') in res_df.columns else (sim_cols[0] if sim_cols else None)

                                slim = pd.DataFrame()
                                slim['Question'] = res_df[q_col] if q_col in res_df.columns else res_df.iloc[:,0]
                                slim['Answer File1'] = res_df[a1_col] if a1_col in res_df.columns else (res_df.iloc[:,1] if res_df.shape[1] > 1 else None)
                                slim['Answer File2'] = res_df[a2_col] if a2_col in res_df.columns else (res_df.iloc[:,2] if res_df.shape[1] > 2 else None)
                                if sim_col and sim_col in res_df.columns:
                                    numeric = pd.to_numeric(res_df[sim_col], errors='coerce')
                                    slim['Similarity'] = numeric.apply(lambda v: (int(round(v)) / 100.0) if pd.notnull(v) else None)
                                # write results and summary only
                                writer_sheet_name = sheet_name if 'sheet_name' in locals() else 'Results'
                                slim.to_excel(writer, index=False, sheet_name=writer_sheet_name)
                                try:
                                    if 'Similarity' in slim.columns:
                                        numeric_sim = pd.to_numeric((slim['Similarity'] * 100).round(2), errors='coerce')
                                        total_pairs = int(numeric_sim.notna().sum())
                                        above_thresh = int((numeric_sim >= threshold).sum())
                                        between_40_thresh = int(((numeric_sim >= 40) & (numeric_sim < threshold)).sum())
                                        below_40 = int((numeric_sim < 40).sum())
                                        avg_similarity = round(float(numeric_sim.mean()), 2) if total_pairs > 0 else 0
                                    else:
                                        total_pairs = len(slim)
                                        above_thresh = between_40_thresh = below_40 = 0
                                        avg_similarity = 0
                                    summary_df = pd.DataFrame({
                                        "Metric": ["Total Pairs", f"Above {threshold}%", f"Between 40-{threshold}%", "Below 40%", "Average Similarity (%)", "High Threshold (%)"],
                                        "Value": [total_pairs, above_thresh, between_40_thresh, below_40, avg_similarity, threshold]
                                    })
                                    summary_name = (f"{writer_sheet_name} Summary"[:28] + '...') if len(f"{writer_sheet_name} Summary") > 31 else f"{writer_sheet_name} Summary"
                                    summary_df.to_excel(writer, index=False, sheet_name=summary_name)
                                    st.session_state['all_results_summary_written'] = True
                                except Exception:
                                    pass
                            except Exception:
                                res_df.to_excel(writer, index=False, sheet_name=sheet_name)
                    else:
                        # Fallback: write the export_df to a single sheet
                        export_df.to_excel(writer, index=False, sheet_name=sheet_name)

                except Exception:
                    # If anything fails, fall back to single-sheet export
                    export_df.to_excel(writer, index=False, sheet_name=sheet_name)

                # After writing sheets, apply percent formatting to any similarity columns across sheets
                try:
                    for sname, ws in writer.sheets.items():
                        headers = [cell.value for cell in ws[1]]
                        for idx_h, h in enumerate(headers, start=1):
                            if h and 'Similarity' in str(h):
                                col_letter = get_column_letter(idx_h)
                                for row in range(2, ws.max_row + 1):
                                    cell = ws[f"{col_letter}{row}"]
                                    if cell.value is not None:
                                        cell.number_format = '0%'
                except Exception:
                    pass
                # Add a Summary sheet (only for the "All Results" export)
                # The user requested no Summary sheet for the "Compare Two Excel Files" mode,
                # so skip creating the global summary when that mode is active.
                if not IS_TWO_FILE_MODE and not st.session_state.get('all_results_summary_written', False):
                    try:
                        # Prefer results stored in session (if the app kept a copy), otherwise use local results_df
                        df_stats = st.session_state.get('results_df', results_df)
                        # Try to find the primary similarity column
                        sim_col = None
                        stored_primary = st.session_state.get('primary_sim_col')
                        if stored_primary and df_stats is not None and stored_primary in getattr(df_stats, 'columns', []):
                            sim_col = stored_primary
                        else:
                            sim_cols = [c for c in getattr(df_stats, 'columns', []) if 'Similarity' in c]
                            sim_col = sim_cols[0] if sim_cols else None

                        if df_stats is not None and sim_col and sim_col in df_stats.columns:
                            numeric_sim = pd.to_numeric(df_stats[sim_col], errors='coerce')
                            total_pairs = int(numeric_sim.notna().sum())
                            # Use the app's threshold slider for above/between/below counts
                            above_thresh = int((numeric_sim >= threshold).sum())
                            between_40_thresh = int(((numeric_sim >= 40) & (numeric_sim < threshold)).sum())
                            below_40 = int((numeric_sim < 40).sum())
                            avg_similarity = round(float(numeric_sim.mean()), 2) if total_pairs > 0 else 0
                            # Also include below 50% as an additional metric (kept for backward compatibility)
                            below_50 = int((numeric_sim < 50).sum())
                        else:
                            # Fallback counts when similarity column isn't available
                            total_pairs = len(df_stats) if df_stats is not None else 0
                            above_thresh = between_40_thresh = below_40 = below_50 = 0
                            avg_similarity = 0
                    except Exception:
                        total_pairs = above_thresh = between_40_thresh = below_40 = below_50 = avg_similarity = ""

                    summary_df = pd.DataFrame({
                        "Metric": [
                            "Total Pairs",
                            f"Above {threshold}%",
                            f"Between 40-{threshold}%",
                            "Below 40%",
                            "Average Similarity (%)",
                            "High Threshold (%)"
                        ],
                        "Value": [total_pairs, above_thresh, between_40_thresh, below_40, avg_similarity, threshold]
                    })
                    # Determine a descriptive summary sheet name that includes the compared sheet when possible
                    try:
                        summary_sheet_base = "Summary"
                        # Prefer single-file selected sheet
                        sel_single = st.session_state.get('selected_sheet_singlefile')
                        if sel_single:
                            summary_sheet_base = f"{sel_single} Summary"
                        else:
                            # Otherwise, try to pull a sheet name from sheet_results mapping
                            sheet_results = st.session_state.get('sheet_results', {})
                            # Prefer file1 mapping
                            if sheet_results.get('file1'):
                                first_sheet = next(iter(sheet_results['file1'].keys()))
                                if first_sheet:
                                    summary_sheet_base = f"{first_sheet} Summary"
                            elif sheet_results.get('file2'):
                                first_sheet = next(iter(sheet_results['file2'].keys()))
                                if first_sheet:
                                    summary_sheet_base = f"{first_sheet} Summary"
                        # Excel sheet name limit is 31 characters
                        summary_sheet_name = (summary_sheet_base[:28] + '...') if len(summary_sheet_base) > 31 else summary_sheet_base
                    except Exception:
                        summary_sheet_name = "Summary"

                    summary_df.to_excel(writer, index=False, sheet_name=summary_sheet_name)
            output_all.seek(0)
            # Use original filename for export
            base_filename = ''
            if uploaded_file1 is not None:
                base_filename = uploaded_file1.name.rsplit('.', 1)[0]
            elif uploaded_file2 is not None:
                base_filename = uploaded_file2.name.rsplit('.', 1)[0]
            else:
                base_filename = 'exported_results'
            export_filename = f"{base_filename}_similarity.xlsx"
            st.download_button(
                label=f"📊 All Results ({len(results_df)} pairs)",
                data=output_all,
                file_name=export_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                help="Download all comparison results",
                key="download_all_results_state"
            )
        with col_dl2:
            if len(results_below_80) > 0:
                if st.session_state.get('original_df') is not None:
                    export_80 = st.session_state['original_df'].loc[results_below_80.index].copy()
                else:
                    orig_cols_80 = [col for col in results_below_80.columns if col not in [
                        'Raw Similarity','Match Quality',
                        'Source File','Source Sheet','Source File 1','Source Sheet 1','Source File 2','Source Sheet 2'
                    ] and 'Similarity' not in col and '(diff)' not in col]
                    export_80 = results_below_80[orig_cols_80].copy()
                # Add similarity columns as numeric percentages
                sim_cols_80 = [c for c in results_below_80.columns if 'Similarity' in c]
                for col in sim_cols_80:
                    numeric = pd.to_numeric(results_below_80[col], errors='coerce')
                    export_80[col] = numeric.apply(lambda v: (int(round(v)) / 100.0) if pd.notnull(v) else None)
                output_80 = io.BytesIO()
                with pd.ExcelWriter(output_80, engine='openpyxl') as writer:
                    export_80.to_excel(writer, index=False, sheet_name="Below 80%")
                    # Apply percent formatting to similarity columns
                    try:
                        ws = writer.sheets["Below 80%"]
                        for col in sim_cols_80:
                            try:
                                col_idx = list(export_80.columns).index(col) + 1
                                col_letter = get_column_letter(col_idx)
                                for row in range(2, ws.max_row + 1):
                                    cell = ws[f"{col_letter}{row}"]
                                    if cell.value is not None:
                                        cell.number_format = '0%'
                            except Exception:
                                continue
                    except Exception:
                        pass
                output_80.seek(0)
                st.download_button(
                    label=f"⚠️ Below 80% ({len(results_below_80)} pairs)",
                    data=output_80,
                    file_name="similarity_match_below_80.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    help="Download pairs with similarity below 80%",
                    key="download_below_80_state"
                )
            else:
                st.info("No results below 80%")
        with col_dl3:
            if len(results_below_50) > 0:
                if st.session_state.get('original_df') is not None:
                    export_50 = st.session_state['original_df'].loc[results_below_50.index].copy()
                else:
                    orig_cols_50 = [col for col in results_below_50.columns if col not in [
                        'Raw Similarity','Match Quality',
                        'Source File','Source Sheet','Source File 1','Source Sheet 1','Source File 2','Source Sheet 2'
                    ] and 'Similarity' not in col and '(diff)' not in col]
                    export_50 = results_below_50[orig_cols_50].copy()
                # Add similarity columns as numeric percentages
                sim_cols_50 = [c for c in results_below_50.columns if 'Similarity' in c]
                for col in sim_cols_50:
                    numeric = pd.to_numeric(results_below_50[col], errors='coerce')
                    export_50[col] = numeric.apply(lambda v: (int(round(v)) / 100.0) if pd.notnull(v) else None)
                output_50 = io.BytesIO()
                with pd.ExcelWriter(output_50, engine='openpyxl') as writer:
                    export_50.to_excel(writer, index=False, sheet_name="Below 50%")
                    # Apply percent formatting to similarity columns
                    try:
                        ws = writer.sheets["Below 50%"]
                        for col in sim_cols_50:
                            try:
                                col_idx = list(export_50.columns).index(col) + 1
                                col_letter = get_column_letter(col_idx)
                                for row in range(2, ws.max_row + 1):
                                    cell = ws[f"{col_letter}{row}"]
                                    if cell.value is not None:
                                        cell.number_format = '0%'
                            except Exception:
                                continue
                    except Exception:
                        pass
                output_50.seek(0)
                st.download_button(
                    label=f"❌ Below 50% ({len(results_below_50)} pairs)",
                    data=output_50,
                    file_name="similarity_match_below_50.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    help="Download pairs with similarity below 50%",
                    key="download_below_50_state"
                )
            else:
                st.info("No results below 50%")

        has_non_spreadsheet_input = (
            comparison_mode == "Compare Any Two Files"
            and (
                (uploaded_file1 is not None and not _is_spreadsheet_upload(uploaded_file1))
                or (uploaded_file2 is not None and not _is_spreadsheet_upload(uploaded_file2))
            )
        )
        if has_non_spreadsheet_input:
            alt_export_df = results_df.drop(columns=[c for c in results_df.columns if "(diff)" in str(c)], errors="ignore")
            alt_display_df = _build_non_excel_export_df(alt_export_df)
            alt_summary = _build_export_summary(
                results_df,
                threshold,
                st.session_state.get('primary_sim_col'),
            )
            alt_base_filename = ''
            if uploaded_file1 is not None:
                alt_base_filename = uploaded_file1.name.rsplit('.', 1)[0]
            elif uploaded_file2 is not None:
                alt_base_filename = uploaded_file2.name.rsplit('.', 1)[0]
            else:
                alt_base_filename = 'exported_results'

            csv_bytes = alt_display_df.to_csv(index=False).encode('utf-8-sig')
            json_payload = _build_non_excel_json_payload(
                alt_export_df,
                alt_summary,
                uploaded_file1=uploaded_file1,
                uploaded_file2=uploaded_file2,
            )
            html_payload = _build_non_excel_html_payload(
                alt_display_df,
                alt_summary,
                f"{alt_base_filename} Similarity Report",
            )

            st.markdown("#### Additional Exports For Non-Spreadsheet Inputs")
            alt_col1, alt_col2, alt_col3 = st.columns(3)
            with alt_col1:
                st.download_button(
                    label="CSV Report",
                    data=csv_bytes,
                    file_name=f"{alt_base_filename}_similarity.csv",
                    mime="text/csv",
                    help="Download the results table as CSV.",
                    key="download_non_excel_csv_state",
                )
            with alt_col2:
                st.download_button(
                    label="JSON Report",
                    data=json_payload,
                    file_name=f"{alt_base_filename}_similarity.json",
                    mime="application/json",
                    help="Download the results plus summary metadata as JSON.",
                    key="download_non_excel_json_state",
                )
            with alt_col3:
                st.download_button(
                    label="HTML Report",
                    data=html_payload,
                    file_name=f"{alt_base_filename}_similarity.html",
                    mime="text/html",
                    help="Download a readable HTML report for PDF/TXT/DOC-style comparisons.",
                    key="download_non_excel_html_state",
                )
else:
    if IS_TWO_FILE_MODE:
        st.info("Please upload both files to begin.")
    else:
        st.info("Please upload a file to begin column comparison.")
